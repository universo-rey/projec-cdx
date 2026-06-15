from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.properties import CalcProperties


BASE_DIR = Path(r"C:\Users\enzo1\PROJEC CDX\outputs\dataverse_blocker_frontier_20260614")
XLSX_PATH = BASE_DIR / "dataverse_blocker_frontier.xlsx"
README_PATH = BASE_DIR / "README.md"
MAPA_PATH = BASE_DIR / "MAPA.md"
READBACK_PATH = BASE_DIR / "READBACK.md"
VALIDATION_PATH = BASE_DIR / "validation.json"


ROWS = [
    {
        "id": "BFR-001",
        "categoria": "Identidad",
        "frontera": "Identidad duplicada o clave inestable",
        "ejemplo": "sdu_asset_key o external IDs duplicados en el paquete SDU",
        "senal": "duplicate identity scan falla; la clave no es row-stable",
        "resolucion": "mixta",
        "accion": "Reformular clave compuesta, regenerar seed y repetir el gate de identidad",
        "rollback": "Revertir el paquete generado y volver al modelo previo",
        "estado": "bloqueante",
        "fuente": "memory + workspace",
        "notas": "No usar el external ID crudo como única identidad cuando el origen permite duplicados.",
    },
    {
        "id": "BFR-002",
        "categoria": "Entorno",
        "frontera": "Entorno equivocado o default",
        "ejemplo": "org o tenant no coincide con el target real",
        "senal": "Get-MgContext, pac env who o metadata muestran otro entorno",
        "resolucion": "humana",
        "accion": "Confirmar tenant y environment exactos antes de seguir",
        "rollback": "Cerrar la sesión incorrecta y volver a autenticarse en el target correcto",
        "estado": "bloqueante",
        "fuente": "memory + live-governance canon",
        "notas": "La frontera correcta es la verificación del target, no la suposición por nombre.",
    },
    {
        "id": "BFR-003",
        "categoria": "Evidencia",
        "frontera": "Metadata only sin filas vivas",
        "ejemplo": "hay metadata o solución, pero no se confirmaron live rows",
        "senal": "entity metadata visible; live row result ausente o ambiguo",
        "resolucion": "gate",
        "accion": "Pedir el gate de lectura live exacto o mantener el carril metadata-only",
        "rollback": "No promover la lectura a dato vivo sin una fila confirmada",
        "estado": "observado",
        "fuente": "memory",
        "notas": "La metadata no se debe elevar a evidencia viva sin row result.",
    },
    {
        "id": "BFR-004",
        "categoria": "Aprobación",
        "frontera": "Review o aprobación stale",
        "ejemplo": "PR con comentarios outdated o review requerida tras nuevos commits",
        "senal": "REVIEW_REQUIRED, outdated thread, approval vieja",
        "resolucion": "humana",
        "accion": "Pedir aprobación nueva o comentario actualizado",
        "rollback": "No mergear hasta revalidar el review actual",
        "estado": "bloqueante",
        "fuente": "workspace",
        "notas": "Este fue el patrón real en los PRs de Torre y Cabina.",
    },
    {
        "id": "BFR-005",
        "categoria": "Cobertura",
        "frontera": "coverage_equivalence falsa",
        "ejemplo": "Local governance validators fallan aunque no haya tests requeridos rotos",
        "senal": "all_required_passed=true pero coverage_equivalence=false",
        "resolucion": "mixta",
        "accion": "Reconciliar manifest, grafo de cobertura o equivalencias antes de empujar de nuevo",
        "rollback": "Restaurar la referencia canónica si la equivalencia estaba mal declarada",
        "estado": "bloqueante",
        "fuente": "workspace",
        "notas": "El bloqueo observado fue una política de cobertura, no un test roto.",
    },
    {
        "id": "BFR-006",
        "categoria": "Seguridad",
        "frontera": "Secretos o credenciales visibles",
        "ejemplo": "secret_like_hits, .env.local trackeado o material sensible detectado",
        "senal": "scan de secretos o inventario con material sensible",
        "resolucion": "humana",
        "accion": "Detener, clasificar y sanear antes de continuar",
        "rollback": "Eliminar la exposición o moverla a una superficie protegida",
        "estado": "bloqueante",
        "fuente": "memory + workspace canon",
        "notas": "Si aparece, no se avanza por inferencia.",
    },
    {
        "id": "BFR-007",
        "categoria": "Gate",
        "frontera": "Live write sin orden exacta",
        "ejemplo": "falta target, owner, rollback, postcheck o evidencia",
        "senal": "approval_required=true sin orden atómica explícita",
        "resolucion": "humana",
        "accion": "Emitir una orden atómica con objetivo, superficie, owner y postcheck",
        "rollback": "Definir reversión antes de cualquier write",
        "estado": "bloqueante",
        "fuente": "atomic contract + memory",
        "notas": "La autorización general no se debe leer como permiso de escritura live.",
    },
    {
        "id": "BFR-008",
        "categoria": "Validación",
        "frontera": "Validador local falla",
        "ejemplo": "workflow termina con exit code 1 o validation FAIL",
        "senal": "exit code no cero o reporte de fallo local",
        "resolucion": "automatica",
        "accion": "Reejecutar el validador correcto tras corregir la causa concreta",
        "rollback": "Vuelta al último estado verificable",
        "estado": "observado",
        "fuente": "workspace",
        "notas": "No todo fallo de validación exige humano, pero sí detiene el avance.",
    },
    {
        "id": "BFR-009",
        "categoria": "Rollback",
        "frontera": "Rollback o compensación indefinida",
        "ejemplo": "acción live propuesta sin salida de reversión",
        "senal": "operating contract sin rollback/postcheck",
        "resolucion": "humana",
        "accion": "No ejecutar hasta definir reversión y readback",
        "rollback": "Diseñar compensación antes del write",
        "estado": "bloqueante",
        "fuente": "atomic contract",
        "notas": "Esta es una frontera fuerte para cualquier superficie gobernada.",
    },
    {
        "id": "BFR-010",
        "categoria": "Ambigüedad",
        "frontera": "Target ambiguo o entidad inexistente",
        "ejemplo": "el nombre no resuelve a una superficie exacta",
        "senal": "target_ambiguous o entity_not_found",
        "resolucion": "gate",
        "accion": "Reducir el alcance hasta nombre exacto y surface exacta",
        "rollback": "No hay acción hasta que el target deje de ser ambiguo",
        "estado": "bloqueante",
        "fuente": "dataverse discovery canon",
        "notas": "La búsqueda debe parar antes de inventar un target.",
    },
    {
        "id": "BFR-011",
        "categoria": "Importabilidad",
        "frontera": "Manifest o índice desalineado",
        "ejemplo": "schema, manifest o equivalencia de cobertura no coinciden",
        "senal": "header mismatch, missing required test o registro obsoleto",
        "resolucion": "automatica",
        "accion": "Regenerar índice/manifest y volver a validar",
        "rollback": "Restaurar el último manifiesto bueno",
        "estado": "mitigable",
        "fuente": "workspace",
        "notas": "Suele ser reparable sin intervención humana si el canon está claro.",
    },
    {
        "id": "BFR-012",
        "categoria": "Superficie",
        "frontera": "Write live externa o costed API",
        "ejemplo": "Microsoft, Graph, Power Platform o Dataverse write sin gate",
        "senal": "acción mutante contra una superficie gobernada",
        "resolucion": "gate",
        "accion": "Mantener lectura gobernada hasta que exista una orden explícita",
        "rollback": "Definir compensación y postcheck antes del write",
        "estado": "bloqueante",
        "fuente": "repo + memory canon",
        "notas": "La superficie puede leerse; la escritura necesita orden atómica.",
    },
]


def setup_sheet_common(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[4].height = 20


def style_range(ws, cell_range, fill=None, font=None, border=None, alignment=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def main() -> None:
    BASE_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    try:
        wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    except Exception:
        pass

    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_registry = wb.create_sheet("Registro")
    ws_lists = wb.create_sheet("Listas")

    setup_sheet_common(ws_summary)
    setup_sheet_common(ws_registry)
    setup_sheet_common(ws_lists)

    dark_teal = "0F766E"
    deep_blue = "155E75"
    light_bg = "F8FAFC"
    pale_bg = "ECFEFF"
    line = Side(style="thin", color="CBD5E1")
    border = Border(left=line, right=line, top=line, bottom=line)
    title_fill = PatternFill("solid", fgColor=dark_teal)
    section_fill = PatternFill("solid", fgColor=deep_blue)
    card_fill = PatternFill("solid", fgColor=pale_bg)
    body_fill = PatternFill("solid", fgColor="FFFFFF")
    note_fill = PatternFill("solid", fgColor=light_bg)
    white_font = Font(color="FFFFFF", bold=True)
    dark_font = Font(color="0F172A", bold=True)
    body_font = Font(color="0F172A")
    muted_font = Font(color="334155")
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Resumen
    ws_summary.merge_cells("A1:L1")
    ws_summary["A1"] = "Dataverse Blocker Frontier"
    style_range(ws_summary, "A1:L1", fill=title_fill, font=white_font, border=border, alignment=center)
    ws_summary.row_dimensions[1].height = 34

    ws_summary.merge_cells("A2:L2")
    ws_summary["A2"] = (
        "Matriz operativa para detectar fronteras que detienen automatización y "
        "separar qué se resuelve solo, qué exige gate y qué necesita intervención humana."
    )
    style_range(ws_summary, "A2:L2", fill=note_fill, font=muted_font, border=border, alignment=wrap)
    ws_summary.row_dimensions[2].height = 36

    ws_summary.merge_cells("A4:B4")
    ws_summary["A4"] = "Resumen ejecutivo"
    style_range(ws_summary, "A4:B4", fill=section_fill, font=white_font, border=border, alignment=center)

    summary_rows = [
        ("Total de fronteras", '=COUNTA(Registro!$A$6:$A$200)'),
        ("Resolución humana", '=COUNTIF(Registro!$F$6:$F$200,"humana")'),
        ("Resolución gate", '=COUNTIF(Registro!$F$6:$F$200,"gate")'),
        ("Resolución mixta", '=COUNTIF(Registro!$F$6:$F$200,"mixta")'),
        ("Resolución automática", '=COUNTIF(Registro!$F$6:$F$200,"automatica")'),
    ]
    ws_summary["A5"] = "Métrica"
    ws_summary["B5"] = "Valor"
    style_range(ws_summary, "A5:B5", fill=section_fill, font=white_font, border=border, alignment=center)
    for idx, (label, formula) in enumerate(summary_rows, start=6):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"B{idx}"] = formula
        style_range(ws_summary, f"A{idx}:B{idx}", fill=card_fill if idx % 2 == 0 else body_fill, font=body_font, border=border, alignment=wrap)

    ws_summary.merge_cells("D4:H4")
    ws_summary["D4"] = "Lectura rápida"
    style_range(ws_summary, "D4:H4", fill=section_fill, font=white_font, border=border, alignment=center)

    ws_summary.merge_cells("D5:H9")
    ws_summary["D5"] = (
        "1. Si hay target ambiguo, entorno equivocado o metadata-only, no se promociona.\n"
        "2. Si aparece coverage_equivalence=false, se reconcilia el canon antes de avanzar.\n"
        "3. Si hay secretos, approvals stale o live write sin orden exacta, la intervención humana manda.\n"
        "4. Lo reparable queda marcado como automática o mixta, pero sigue siendo frontera."
    )
    style_range(ws_summary, "D5:H9", fill=note_fill, font=muted_font, border=border, alignment=wrap)
    ws_summary.row_dimensions[5].height = 54
    ws_summary.row_dimensions[6].height = 54
    ws_summary.row_dimensions[7].height = 54
    ws_summary.row_dimensions[8].height = 54
    ws_summary.row_dimensions[9].height = 54

    ws_summary.merge_cells("D11:H11")
    ws_summary["D11"] = "Distribución por tipo de resolución"
    style_range(ws_summary, "D11:H11", fill=section_fill, font=white_font, border=border, alignment=center)
    ws_summary["D12"] = "Tipo"
    ws_summary["E12"] = "Cantidad"
    style_range(ws_summary, "D12:E12", fill=section_fill, font=white_font, border=border, alignment=center)
    types = ["humana", "gate", "mixta", "automatica"]
    for offset, value in enumerate(types, start=13):
        ws_summary[f"D{offset}"] = value
        ws_summary[f"E{offset}"] = f'=COUNTIF(Registro!$F$6:$F$200,D{offset})'
        style_range(ws_summary, f"D{offset}:E{offset}", fill=body_fill, font=body_font, border=border, alignment=wrap)

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Fronteras por tipo de resolución"
    chart.y_axis.title = "Tipo"
    chart.x_axis.title = "Cantidad"
    data = Reference(ws_summary, min_col=5, min_row=12, max_row=16)
    cats = Reference(ws_summary, min_col=4, min_row=13, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7.2
    chart.width = 11
    chart.legend = None
    ws_summary.add_chart(chart, "J4")

    for col, width in {
        "A": 26,
        "B": 14,
        "C": 28,
        "D": 18,
        "E": 18,
        "F": 16,
        "G": 22,
        "H": 22,
        "I": 14,
        "J": 22,
        "K": 18,
        "L": 36,
    }.items():
        ws_summary.column_dimensions[col].width = width

    # Registro
    ws_registry.merge_cells("A1:K1")
    ws_registry["A1"] = "Registro de fronteras"
    style_range(ws_registry, "A1:K1", fill=title_fill, font=white_font, border=border, alignment=center)
    ws_registry.row_dimensions[1].height = 32

    ws_registry.merge_cells("A2:K2")
    ws_registry["A2"] = "Una fila por frontera o bloqueo. El objetivo es separar lo humano, lo automático y lo que requiere gate."
    style_range(ws_registry, "A2:K2", fill=note_fill, font=muted_font, border=border, alignment=wrap)
    ws_registry.row_dimensions[2].height = 30

    headers = [
        "ID",
        "Categoria",
        "Frontera",
        "Ejemplo real",
        "Señal de detección",
        "Tipo de resolución",
        "Acción mínima",
        "Rollback / salida",
        "Estado",
        "Fuente",
        "Notas",
    ]
    for idx, header in enumerate(headers, start=1):
        cell = ws_registry.cell(row=5, column=idx, value=header)
        cell.fill = section_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = center

    start_row = 6
    for row_idx, item in enumerate(ROWS, start=start_row):
        values = [
            item["id"],
            item["categoria"],
            item["frontera"],
            item["ejemplo"],
            item["senal"],
            item["resolucion"],
            item["accion"],
            item["rollback"],
            item["estado"],
            item["fuente"],
            item["notas"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_registry.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = wrap
            cell.font = body_font
            cell.fill = body_fill if row_idx % 2 else card_fill
        # Highlight the high-friction cells
        if item["resolucion"] == "humana":
            ws_registry.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="FDE68A")
        elif item["resolucion"] == "gate":
            ws_registry.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="BFDBFE")
        elif item["resolucion"] == "mixta":
            ws_registry.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="DDD6FE")
        else:
            ws_registry.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="BBF7D0")

        if item["estado"] == "bloqueante":
            ws_registry.cell(row=row_idx, column=9).fill = PatternFill("solid", fgColor="FCA5A5")
        elif item["estado"] == "mitigable":
            ws_registry.cell(row=row_idx, column=9).fill = PatternFill("solid", fgColor="FDE68A")
        else:
            ws_registry.cell(row=row_idx, column=9).fill = PatternFill("solid", fgColor="BFDBFE")

    end_row = start_row + len(ROWS) - 1
    table = Table(displayName="BlockerFrontierTable", ref=f"A5:K{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_registry.add_table(table)
    ws_registry.freeze_panes = "A6"

    # Data validations
    dv_resolution = DataValidation(type="list", formula1="=Listas!$A$2:$A$5", allow_blank=False)
    dv_state = DataValidation(type="list", formula1="=Listas!$B$2:$B$4", allow_blank=False)
    ws_registry.add_data_validation(dv_resolution)
    ws_registry.add_data_validation(dv_state)
    dv_resolution.add(f"F6:F{end_row}")
    dv_state.add(f"I6:I{end_row}")

    # Column widths
    widths = {
        "A": 12,
        "B": 16,
        "C": 28,
        "D": 30,
        "E": 30,
        "F": 18,
        "G": 34,
        "H": 24,
        "I": 14,
        "J": 18,
        "K": 42,
    }
    for col, width in widths.items():
        ws_registry.column_dimensions[col].width = width

    # Lists
    ws_lists.merge_cells("A1:D1")
    ws_lists["A1"] = "Listas de validación"
    style_range(ws_lists, "A1:D1", fill=title_fill, font=white_font, border=border, alignment=center)
    ws_lists.row_dimensions[1].height = 28

    ws_lists["A2"] = "Tipo de resolución"
    ws_lists["B2"] = "Estado"
    ws_lists["C2"] = "Categoria"
    ws_lists["D2"] = "Ejemplo de uso"
    style_range(ws_lists, "A2:D2", fill=section_fill, font=white_font, border=border, alignment=center)

    resolution_values = ["humana", "gate", "mixta", "automatica"]
    state_values = ["bloqueante", "mitigable", "observado"]
    category_values = [
        "Identidad",
        "Entorno",
        "Evidencia",
        "Aprobación",
        "Cobertura",
        "Seguridad",
        "Gate",
        "Validación",
        "Rollback",
        "Ambigüedad",
        "Importabilidad",
        "Superficie",
    ]
    use_values = [
        "Intervención humana obligatoria",
        "Requiere gate explícito",
        "Se puede reparar y revalidar",
        "Se automatiza sin abrir live",
    ]

    for idx, value in enumerate(resolution_values, start=2):
        ws_lists[f"A{idx}"] = value
    for idx, value in enumerate(state_values, start=2):
        ws_lists[f"B{idx}"] = value
    for idx, value in enumerate(category_values, start=2):
        ws_lists[f"C{idx}"] = value
    for idx, value in enumerate(use_values, start=2):
        ws_lists[f"D{idx}"] = value

    style_range(ws_lists, f"A2:D{max(len(category_values), len(use_values), len(resolution_values), len(state_values)) + 1}",
                fill=body_fill, font=body_font, border=border, alignment=wrap)
    for col, width in {"A": 18, "B": 16, "C": 18, "D": 28}.items():
        ws_lists.column_dimensions[col].width = width

    # File outputs
    wb.save(XLSX_PATH)

    counts = Counter(item["resolucion"] for item in ROWS)
    state_counts = Counter(item["estado"] for item in ROWS)
    notes = [
        "# Dataverse Blocker Frontier",
        "",
        "Este workbook concentra las fronteras que más probablemente detienen automatización y exigen intervención humana, gate explícito o reconciliación local.",
        "",
        "## Resumen",
        f"- Total de fronteras: {len(ROWS)}",
        f"- Resolución humana: {counts['humana']}",
        f"- Resolución gate: {counts['gate']}",
        f"- Resolución mixta: {counts['mixta']}",
        f"- Resolución automática: {counts['automatica']}",
        "",
        "## Estado",
        f"- bloqueante: {state_counts['bloqueante']}",
        f"- mitigable: {state_counts['mitigable']}",
        f"- observado: {state_counts['observado']}",
        "",
        "## Lectura",
        "- La frontera más fuerte sigue siendo la de autorización humana para live writes y aprobación.",
        "- La frontera más frágil para el motor es la identidad duplicada o la cobertura desalineada.",
        "- Metadata-only sirve para descubrir, no para afirmar estado vivo.",
        "",
        "## Siguiente paso recomendado",
        "Conectar este workbook a una tabla Dataverse o a un tracker de PRs para registrar bloqueos reales por superficie, owner y readback.",
    ]
    README_PATH.write_text("\n".join(notes) + "\n", encoding="utf-8")

    mapa = [
        "# Mapa del paquete",
        "",
        "- `dataverse_blocker_frontier.xlsx`: workbook principal con Resumen, Registro y Listas.",
        "- `README.md`: lectura rápida y cifras del paquete.",
        "- `MAPA.md`: guía de las hojas y uso esperado.",
        "- `READBACK.md`: conclusión corta sobre la frontera detectada.",
        "",
        "## Hojas",
        "",
        "- **Resumen**: KPIs y lectura ejecutiva.",
        "- **Registro**: una fila por frontera o bloqueo.",
        "- **Listas**: vocabularios para validación.",
    ]
    MAPA_PATH.write_text("\n".join(mapa) + "\n", encoding="utf-8")

    readback = {
        "total_fronteras": len(ROWS),
        "resolucion": dict(counts),
        "estado": dict(state_counts),
        "conclusion": "La frontera dominante es humana/gate: identidad, approvals, live writes y target exacto. La frontera automática existe, pero no reemplaza el gate cuando hay escritura, aprobación o seguridad.",
    }
    READBACK_PATH.write_text(json.dumps(readback, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
