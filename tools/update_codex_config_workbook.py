from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    import tomllib
except ModuleNotFoundError:  # pragma: no cover - Python <3.11 fallback
    import tomli as tomllib


ROOT = Path(r"C:\Users\enzo1\PROJEC CDX")
CODEX_ROOT = Path(r"C:\Users\enzo1\.codex")
AGENTS_ROOT = Path(r"C:\Users\enzo1\.agents")
WORKBOOK = ROOT / "workbooks" / "CODEX_GLOBAL_STATE_DECISION_WORKBOOK_20260617.xlsx"
BACKUP_DIR = ROOT / "workbooks" / "_backups"
STATE_PATH = CODEX_ROOT / ".codex-global-state.json"
STATE_BAK = CODEX_ROOT / ".codex-global-state.json.bak"
CONFIG_PATH = CODEX_ROOT / "config.toml"
VHDX_PATH = Path(
    r"C:\Users\enzo1\Documents\Maquina de Trabajo\C_ROOT_20260615\SGIN_SANDBOX"
    r"\DISCO_VIRTUAL\00_DISCO_DESARROLLO.vhdx"
)
EVIDENCE_LOG = Path(
    r"C:\Users\enzo1\CodexLocal\OPTIMIZACION_PC"
    r"\D_DRIVE_RECOVERY_20260617_210929\diskpart_attach_readonly.log"
)
GITHUB_ROOT = Path(r"C:\Users\enzo1\Documents\GitHub")
HOME_CODEX_LOCAL = Path(r"C:\Users\enzo1\CodexLocal")
DOCUMENTS_CODEX_LOCAL = Path(r"C:\Users\enzo1\Documents\CodexLocal")
DOCUMENTS_CODEX = Path(r"C:\Users\enzo1\Documents\Codex")
DOCUMENTS_CODEX_ARCHIVES = Path(r"C:\Users\enzo1\Documents\CodexArchives")
D_ROOT = Path(r"D:\\")
D_CODEX_ROOT = D_ROOT / ".agents" / "codex"
D_SKILLS_ROOT = D_ROOT / ".agents" / "skills"
D_AUTHORITY_CANON = D_ROOT / "02_AUTHORITY_CANON"
D_GOVERNANCE_REGISTRY = D_ROOT / "01_GOVERNANCE_REGISTRY"
D_MATRICES_ROOT = D_ROOT / "matrices"
D_DATAVERSE_DATA = D_ROOT / "dataverse" / "data"
D_VALIDATION_ROOT = D_ROOT / "validation"
D_ACTIVE_REPOS = Path(r"D:\03_CORTE_EJECUTORA_DEL_REY\10_REPOS\02_ACTIVE")
D_MATRIX_INDEX = D_CODEX_ROOT / "matrices" / "MATRIX_INDEX.csv"
D_RECIPE_INDEX = D_CODEX_ROOT / "recipes" / "RECIPE_INDEX.csv"
D_SKILL_USAGE_MATRIX = D_CODEX_ROOT / "skills" / "SKILL_USAGE_MATRIX.csv"
D_TOOL_INDEX = D_CODEX_ROOT / "tools" / "TOOL_INDEX.csv"
D_AGENTS_INDEX = D_CODEX_ROOT / "AGENTS_INDEX.csv"
D_AGENTS_JSON = D_CODEX_ROOT / "agents.json"
D_ROUTING_JSON = D_CODEX_ROOT / "routing.json"
D_LEVELS_CSV = D_CODEX_ROOT / "agents" / "LEVELS.csv"
D_AGENT_WORKPAPERS_MATRIX = D_CODEX_ROOT / "matrices" / "AGENT_WORKPAPERS_MATRIX.csv"
D_GOVERNED_ASSET_INVENTORY = D_CODEX_ROOT / "matrices" / "GOVERNED_ASSET_CANONICAL_INVENTORY.csv"
SKILLS_UNIFIED_TABLE = ROOT / "inventarios" / "SKILLS_UNIFIED_TABLE.csv"
UNIVERSE_AUDIT = ROOT / "outputs" / "universe_relationship_audit_20260614" / "UNIVERSE_RELATIONSHIP_AUDIT.csv"
CONNECTION_MATRIX_DIR = CODEX_ROOT / "matrices" / "matrices" / "connections"
DATAVERSE_CONNECTION_DATA = CODEX_ROOT / "matrices" / "dataverse" / "data"
DATAVERSE_LIVE_SUMMARY = ROOT / "operativa" / "DATAVERSE_POWER_PLATFORM_LIVE_SUMMARY_20260616.json"
DATAVERSE_LIVE_INVENTORY = ROOT / "operativa" / "DATAVERSE_POWER_PLATFORM_LIVE_INVENTORY_20260616.json"
DATAVERSE_BOTS_CSV = ROOT / "inventarios" / "DATAVERSE_BOTS_LIVE_20260616.csv"
DATAVERSE_SOLUTIONS_CSV = ROOT / "inventarios" / "DATAVERSE_SOLUTIONS_LIVE_20260616.csv"
DATAVERSE_WORKQUEUES_CSV = ROOT / "inventarios" / "DATAVERSE_WORKQUEUES_LIVE_20260616.csv"
ATOMIC_ACTION_MATRIX = ROOT / "atomic" / "CODEX_ATOMIC_ACTION_MATRIX.csv"
MASTER_ATOMIC_MATRIX = ROOT / "operativa" / "MATRIZ_PLAN_MAESTRO_ATOMICO_20260617.csv"
METADATA_HYDRATION_MATRIX = ROOT / "hitos" / "20260616-sdu-dataverse-metadata-wave-v1" / "METADATA_HYDRATION_MATRIX.csv"
PROMPT_CHUNK_SIZE = 25000

TITLE_FILL = "1F4E79"
SUB_FILL = "D9EAF7"
HEADER_FILL = "244062"
ALT_FILL = "F7FBFF"
OK_FILL = "E2F0D9"
WARN_FILL = "FFF2CC"
BAD_FILL = "FCE4D6"
WHITE = "FFFFFF"
BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def sha256(path: Path) -> str | None:
    return hashlib.sha256(path.read_bytes()).hexdigest().upper() if path.exists() else None


def kb(value: int | float | None) -> float | None:
    return round(float(value) / 1024, 2) if value is not None else None


def gb(value: int | float | None) -> float | None:
    return round(float(value) / (1024**3), 2) if value is not None else None


def ps_json(command: str):
    try:
        result = subprocess.run(
            ["powershell.exe", "-NoProfile", "-Command", command],
            capture_output=True,
            text=True,
            timeout=25,
        )
        if result.returncode or not result.stdout.strip():
            return None
        return json.loads(result.stdout)
    except Exception:
        return None


def as_list(value):
    if value is None:
        return []
    return value if isinstance(value, list) else [value]


def json_size(value) -> int:
    return len(json.dumps(value, ensure_ascii=False, sort_keys=True))


def safe(value):
    if value is None:
        return "null"
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


SIMPLE_PATTERNS = [
    "OPENAI_API_KEY",
    "api_key",
    "client_secret",
    "BEGIN PRIVATE KEY",
    "Bearer ",
    "password",
    "access_token",
    "refresh_token",
]
TOKEN_PATTERNS = {
    "openai_sk_like": re.compile(r"sk-[A-Za-z0-9]{20,}"),
    "github_token_like": re.compile(r"gh[pousr]_[A-Za-z0-9_]{20,}"),
    "jwt_like": re.compile(
        r"eyJ[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}"
    ),
    "aws_access_key_like": re.compile(r"AKIA[0-9A-Z]{16}"),
    "private_key_block": re.compile(
        r"-----BEGIN [A-Z ]*PRIVATE KEY-----.*?-----END [A-Z ]*PRIVATE KEY-----",
        re.DOTALL,
    ),
}


def count_mentions(text: str) -> dict[str, int]:
    lower = text.lower()
    counts = {pattern: lower.count(pattern.lower()) for pattern in SIMPLE_PATTERNS}
    counts["Bearer "] = text.count("Bearer ")
    counts.update({name: len(regex.findall(text)) for name, regex in TOKEN_PATTERNS.items()})
    return counts


def sanitize_prompt_text(text: str) -> tuple[str, int]:
    sanitized = text
    redactions = 0
    for name, regex in TOKEN_PATTERNS.items():
        sanitized, count = regex.subn(f"[REDACTED_{name.upper()}]", sanitized)
        redactions += count
    assignment_regex = re.compile(
        r"(?i)\b(OPENAI_API_KEY|api_key|client_secret|access_token|refresh_token|password)\s*[:=]\s*([^\s,;]+)"
    )
    sanitized, count = assignment_regex.subn(r"\1=[REDACTED_VALUE]", sanitized)
    redactions += count
    bearer_regex = re.compile(r"(?i)\bBearer\s+[A-Za-z0-9._~+/=-]{12,}")
    sanitized, count = bearer_regex.subn("Bearer [REDACTED_VALUE]", sanitized)
    redactions += count
    return sanitized, redactions


def chunk_text(text: str, size: int = PROMPT_CHUNK_SIZE) -> list[str]:
    return [text[i : i + size] for i in range(0, len(text), size)] or [""]


def git_value(path: Path, args: list[str]) -> str:
    try:
        result = subprocess.run(
            ["git", "-C", str(path), *args],
            capture_output=True,
            text=True,
            timeout=10,
        )
        return result.stdout.strip() if result.returncode == 0 else ""
    except Exception:
        return ""


def command_json(args: list[str], timeout: int = 25):
    try:
        result = subprocess.run(args, capture_output=True, text=True, timeout=timeout)
        if result.returncode != 0 or not result.stdout.strip():
            return None
        return json.loads(result.stdout)
    except Exception:
        return None


def command_text(args: list[str], timeout: int = 25) -> tuple[int, str, str]:
    try:
        result = subprocess.run(args, capture_output=True, text=True, timeout=timeout)
        return result.returncode, result.stdout.strip(), result.stderr.strip()
    except Exception as exc:
        return 1, "", str(exc)


def repo_slug_from_remote(remote: str) -> str:
    if not remote or remote == "N/D":
        return ""
    cleaned = remote.strip()
    cleaned = re.sub(r"^git@github\.com:", "", cleaned)
    cleaned = re.sub(r"^https://github\.com/", "", cleaned)
    cleaned = re.sub(r"^http://github\.com/", "", cleaned)
    cleaned = cleaned[:-4] if cleaned.endswith(".git") else cleaned
    return cleaned.strip("/")


def parse_ahead_behind(status_line: str) -> tuple[str, str]:
    ahead = ""
    behind = ""
    match = re.search(r"\[(.*?)\]", status_line or "")
    if not match:
        return ahead, behind
    for part in match.group(1).split(","):
        token = part.strip()
        if token.startswith("ahead "):
            ahead = token.replace("ahead ", "")
        elif token.startswith("behind "):
            behind = token.replace("behind ", "")
    return ahead, behind


def repo_row(path: Path, source: str, note: str) -> list:
    exists = path.exists()
    git_root = git_value(path, ["rev-parse", "--show-toplevel"]) if exists else ""
    is_git = bool(git_root)
    git_path = Path(git_root) if git_root else path
    remote = git_value(git_path, ["remote", "get-url", "origin"]) if is_git else ""
    branch = git_value(git_path, ["branch", "--show-current"]) if is_git else ""
    head = git_value(git_path, ["rev-parse", "--short", "HEAD"]) if is_git else ""
    status = git_value(git_path, ["status", "--short"]) if is_git else ""
    dirty_count = len([line for line in status.splitlines() if line.strip()])
    return [
        source,
        path.name or str(path),
        str(path),
        "SI" if exists else "NO",
        "SI" if is_git else "NO",
        git_root or "N/D",
        branch or "N/D",
        head or "N/D",
        remote or "N/D",
        dirty_count if is_git else None,
        "NO_TOCAR" if is_git else "OBSERVAR",
        note,
    ]


def discover_repositories(projects: list[str]) -> list[list]:
    candidates: list[tuple[Path, str, str]] = [
        (ROOT, "WORKSPACE_ACTUAL", "Repo/workspace actual de la sesión."),
        (Path(r"D:\\"), "D_ROOT_WRAPPER", "Cabina D wrapper repo recuperada read-only."),
    ]
    if GITHUB_ROOT.exists():
        for child in sorted([p for p in GITHUB_ROOT.iterdir() if p.is_dir()], key=lambda p: p.name.lower()):
            candidates.append((child, "GITHUB_LOCAL", "Directorio bajo Documents/GitHub."))
    if D_ACTIVE_REPOS.exists():
        for child in sorted([p for p in D_ACTIVE_REPOS.iterdir() if p.is_dir()], key=lambda p: p.name.lower()):
            candidates.append((child, "D_ACTIVE_REPOS", "Repo activo dentro de Corte Ejecutora D."))
    for project in projects:
        candidates.append((Path(project), "CODEX_TRUSTED_PROJECT", "Entrada [projects] en config.toml."))

    seen = set()
    rows = []
    for path, source, note in candidates:
        key = os.path.normcase(os.path.abspath(str(path)))
        if key in seen:
            continue
        seen.add(key)
        rows.append(repo_row(path, source, note))
    return rows


def unique_repo_contexts(repo_rows: list[list]) -> list[dict[str, str]]:
    contexts: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in repo_rows:
        git_root = str(row[5] or "")
        if not git_root or git_root == "N/D":
            continue
        key = os.path.normcase(os.path.abspath(git_root))
        if key in seen:
            continue
        seen.add(key)
        remote = str(row[8] or "")
        contexts.append(
            {
                "source": str(row[0] or ""),
                "name": str(row[1] or ""),
                "path": str(row[2] or ""),
                "git_root": git_root,
                "branch": str(row[6] or ""),
                "head": str(row[7] or ""),
                "remote": remote,
                "slug": repo_slug_from_remote(remote),
                "dirty_count": str(row[9] or ""),
            }
        )
    contexts.sort(key=lambda item: (item.get("slug") or item.get("git_root") or "").lower())
    return contexts


def discover_branch_rows(repo_rows: list[list]) -> list[list]:
    rows: list[list] = []
    live_slugs_seen: set[str] = set()
    for ctx in unique_repo_contexts(repo_rows):
        git_root = Path(ctx["git_root"])
        current = git_value(git_root, ["branch", "--show-current"])
        status_head = git_value(git_root, ["status", "--short", "--branch"]).splitlines()
        ahead, behind = parse_ahead_behind(status_head[0] if status_head else "")
        ref_output = git_value(
            git_root,
            [
                "for-each-ref",
                "--format=%(refname)|%(refname:short)|%(objectname:short)|%(committerdate:iso8601)|%(authorname)|%(upstream:short)",
                "refs/heads",
                "refs/remotes/origin",
            ],
        )
        for line in ref_output.splitlines():
            parts = line.split("|")
            if len(parts) < 6:
                continue
            refname, short, commit, committed_at, author, upstream = parts[:6]
            if short == "origin/HEAD":
                continue
            scope = "LOCAL_BRANCH" if refname.startswith("refs/heads/") else "REMOTE_TRACKING"
            rows.append(
                [
                    scope,
                    ctx["slug"] or ctx["name"],
                    ctx["name"],
                    ctx["git_root"],
                    short,
                    "SI" if scope == "LOCAL_BRANCH" and short == current else "NO",
                    upstream,
                    ahead,
                    behind,
                    commit,
                    committed_at,
                    author,
                    "",
                    "",
                    "LOCAL_REF",
                    "Referencia Git local; no implica que GitHub remoto ya esté fresco.",
                ]
            )

        if ctx["slug"] and ctx["slug"] not in live_slugs_seen:
            live_slugs_seen.add(ctx["slug"])
            branches = command_json(["gh", "api", f"repos/{ctx['slug']}/branches?per_page=100"], timeout=35)
            if isinstance(branches, list):
                for branch in branches:
                    if not isinstance(branch, dict):
                        continue
                    rows.append(
                        [
                            "GITHUB_LIVE_BRANCH",
                            ctx["slug"],
                            ctx["name"],
                            ctx["git_root"],
                            branch.get("name"),
                            "N/D",
                            "",
                            "",
                            "",
                            (branch.get("commit") or {}).get("sha", "")[:12],
                            "",
                            "",
                            branch.get("protected"),
                            "",
                            "GITHUB_LIVE_READ",
                            "Lectura GitHub API via gh; primeras 100 ramas.",
                        ]
                    )
            else:
                rows.append(
                    [
                        "GITHUB_LIVE_UNAVAILABLE",
                        ctx["slug"],
                        ctx["name"],
                        ctx["git_root"],
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "REVISAR",
                        "No se pudo leer ramas live con gh para este repo.",
                    ]
                )
    return rows


def discover_pr_rows(repo_rows: list[list]) -> list[list]:
    rows: list[list] = []
    slugs_seen: set[str] = set()
    for ctx in unique_repo_contexts(repo_rows):
        slug = ctx["slug"]
        if not slug:
            continue
        if slug in slugs_seen:
            continue
        slugs_seen.add(slug)
        pr_data = command_json(
            [
                "gh",
                "pr",
                "list",
                "--repo",
                slug,
                "--state",
                "all",
                "--limit",
                "50",
                "--json",
                "number,title,state,isDraft,headRefName,baseRefName,url,updatedAt,createdAt,author,mergeStateStatus,reviewDecision",
            ],
            timeout=45,
        )
        if isinstance(pr_data, list):
            if not pr_data:
                rows.append(
                    [
                        "GITHUB_LIVE_PR_LIST",
                        slug,
                        "",
                        "SIN_PRS_VISIBLES",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "OK",
                        "gh pr list no devolvio PRs dentro del limite solicitado.",
                    ]
                )
            for pr in pr_data:
                if not isinstance(pr, dict):
                    continue
                state = pr.get("state")
                is_draft = pr.get("isDraft")
                action = "REVISAR_ABIERTO" if state == "OPEN" else "HISTORICO"
                if state == "OPEN" and is_draft:
                    action = "REVISAR_DRAFT"
                author = pr.get("author") or {}
                rows.append(
                    [
                        "GITHUB_LIVE_PR_LIST",
                        slug,
                        pr.get("number"),
                        state,
                        is_draft,
                        pr.get("headRefName"),
                        pr.get("baseRefName"),
                        pr.get("title"),
                        author.get("login") if isinstance(author, dict) else "",
                        pr.get("createdAt"),
                        pr.get("updatedAt"),
                        pr.get("mergeStateStatus"),
                        pr.get("reviewDecision"),
                        action,
                        pr.get("url"),
                    ]
                )
        else:
            rows.append(
                [
                    "GITHUB_LIVE_UNAVAILABLE",
                    slug,
                    "",
                    "UNKNOWN",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "REVISAR",
                    "No se pudo leer PRs con gh para este repo.",
                ]
            )
    rows.sort(key=lambda item: (str(item[1]).lower(), 0 if item[3] == "OPEN" else 1, str(item[2])))
    return rows


def discover_cloud_environment_rows(env: dict, prompt_text: str) -> list[list]:
    rows = []
    repo_map = env.get("repo_map") or {}
    repo_names = []
    if isinstance(repo_map, dict):
        for repo in repo_map.values():
            if isinstance(repo, dict):
                repo_names.append(repo.get("repository_full_name") or repo.get("full_name") or repo.get("name"))
    rows.append(
        [
            "CONFIRMADO_LOCAL",
            env.get("id"),
            env.get("label"),
            env.get("workspace_dir"),
            ", ".join([r for r in repo_names if r]) or safe(env.get("repos")),
            safe((env.get("agent_network_access") or {}).get("mode")),
            env.get("task_count"),
            len(env.get("secrets") or {}),
            safe((env.get("cache_settings") or {}).get("post_setup_cache_enabled")),
            "\n".join(as_list(env.get("setup"))),
            "\n".join(as_list(env.get("maintenance_setup"))),
            "NO_TOCAR",
            "Objeto estructurado vigente en global-state local.",
        ]
    )

    active_id = str(env.get("id") or "")
    env_url_regex = re.compile(r"https://chatgpt\.com/codex/cloud/settings/environment/([a-f0-9]{32})", re.I)
    for env_id in sorted(set(env_url_regex.findall(prompt_text))):
        if env_id == active_id:
            continue
        rows.append(
            [
                "MENCION_PROMPT_URL",
                env_id,
                "N/D",
                "N/D",
                "N/D",
                "N/D",
                "N/D",
                "N/D",
                "N/D",
                "N/D",
                "N/D",
                "VERIFICAR_CLOUD_LIVE",
                "URL encontrada en prompt-history; no hay objeto local estructurado.",
            ]
        )

    repo_regex = re.compile(r"\b(?:universo-rey|SeshatSgin|SeshatEB)/[A-Za-z0-9_.-]+\b")
    active_label = str(env.get("label") or "")
    for repo in sorted(set(repo_regex.findall(prompt_text))):
        if repo == active_label:
            continue
        rows.append(
            [
                "MENCION_PROMPT_REPO",
                "N/D",
                repo,
                "N/D",
                repo,
                "N/D",
                "N/D",
                "N/D",
                "N/D",
                "N/D",
                "N/D",
                "VERIFICAR_CLOUD_LIVE",
                "Repo/label mencionado en prompts; puede corresponder a environment Cloud histórico.",
            ]
        )
    return rows


def read_csv_dicts(path: Path, limit: int | None = None) -> list[dict[str, str]]:
    if not path.exists():
        return []
    rows: list[dict[str, str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for index, row in enumerate(reader, 1):
            if limit is not None and index > limit:
                break
            rows.append({str(k): "" if v is None else str(v) for k, v in row.items()})
    return rows


def read_json_obj(path: Path):
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def path_from_cell(value: str) -> Path | None:
    if not value:
        return None
    candidate = value.strip().strip('"')
    if "|" in candidate:
        candidate = candidate.split("|", 1)[0].strip()
    if re.match(r"^[A-Za-z]:[\\/]", candidate):
        return Path(candidate.replace("/", "\\"))
    return None


def csv_shape(path: Path | None) -> tuple[int | None, str]:
    if path is None or not path.exists() or path.suffix.lower() != ".csv":
        return None, ""
    rows = read_csv_dicts(path)
    columns = list(rows[0].keys()) if rows else []
    return len(rows), "|".join(columns)


def dir_counts(path: Path, recursive: bool = True) -> tuple[int | None, int | None, str]:
    if not path.exists() or not path.is_dir():
        return None, None, ""
    dirs = 0
    files = 0
    extensions: dict[str, int] = {}
    iterator = path.rglob("*") if recursive else path.iterdir()
    try:
        for item in iterator:
            if item.is_dir():
                dirs += 1
            elif item.is_file():
                files += 1
                suffix = item.suffix.lower() or "[none]"
                extensions[suffix] = extensions.get(suffix, 0) + 1
    except Exception:
        return dirs, files, ""
    ext_text = ", ".join(f"{key}:{value}" for key, value in sorted(extensions.items(), key=lambda item: (-item[1], item[0]))[:8])
    return dirs, files, ext_text


def dir_file_stats(path: Path, recursive: bool = True) -> tuple[int | None, int | None, int | None, str]:
    if not path.exists() or not path.is_dir():
        return None, None, None, ""
    dirs = 0
    files = 0
    bytes_total = 0
    extensions: dict[str, int] = {}
    iterator = path.rglob("*") if recursive else path.iterdir()
    try:
        for item in iterator:
            if item.is_dir():
                dirs += 1
            elif item.is_file():
                files += 1
                try:
                    bytes_total += item.stat().st_size
                except OSError:
                    pass
                suffix = item.suffix.lower() or "[none]"
                extensions[suffix] = extensions.get(suffix, 0) + 1
    except Exception:
        return dirs, files, bytes_total, ""
    ext_text = ", ".join(
        f"{key}:{value}"
        for key, value in sorted(extensions.items(), key=lambda item: (-item[1], item[0]))[:8]
    )
    return dirs, files, bytes_total, ext_text


def file_modified(path: Path | None) -> str:
    if path is None or not path.exists():
        return ""
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def file_label(path: Path) -> str:
    try:
        resolved = path.resolve()
    except Exception:
        resolved = path
    prefixes = [
        (ROOT, "PROJEC_CDX"),
        (CODEX_ROOT, "CODEX_ROOT"),
        (AGENTS_ROOT, "AGENTS_ROOT"),
        (GITHUB_ROOT, "GITHUB_ROOT"),
        (Path(r"D:\\"), "D_ROOT"),
    ]
    for root, label in prefixes:
        try:
            if resolved == root.resolve() or root.resolve() in resolved.parents:
                return label
        except Exception:
            continue
    return "LOCAL"


def read_text_sample(path: Path, limit: int = 6000) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="replace")[:limit]
    except Exception:
        return ""


def first_markdown_heading(path: Path) -> str:
    for line in read_text_sample(path).splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            return stripped.lstrip("#").strip()
    return path.stem


def first_meaningful_line(path: Path) -> str:
    for line in read_text_sample(path).splitlines():
        stripped = line.strip()
        if stripped and not stripped.startswith("---") and not stripped.startswith("#"):
            return stripped[:500]
    return ""


def skill_metadata(path: Path) -> dict[str, str]:
    text = read_text_sample(path)
    meta: dict[str, str] = {}
    if text.startswith("---"):
        parts = text.split("---", 2)
        if len(parts) >= 3:
            for line in parts[1].splitlines():
                if ":" not in line:
                    continue
                key, value = line.split(":", 1)
                meta[key.strip()] = value.strip().strip('"')
    if "name" not in meta:
        meta["name"] = path.parent.name
    if "description" not in meta:
        meta["description"] = first_meaningful_line(path)
    return meta


def parse_md_table(path: Path, expected_headers: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    headers: list[str] | None = None
    for line in read_text_sample(path, 20000).splitlines():
        stripped = line.strip()
        if not stripped.startswith("|") or stripped.count("|") < 2:
            continue
        cells = [cell.strip().strip("`") for cell in stripped.strip("|").split("|")]
        if all(set(cell.replace(" ", "")) <= {"-"} for cell in cells):
            continue
        if headers is None:
            headers = cells
            continue
        if len(cells) != len(headers):
            continue
        row = dict(zip(headers, cells))
        if all(header in row for header in expected_headers):
            rows.append(row)
    return rows


def discover_agent_rows() -> list[list]:
    rows: list[list] = []

    for row in read_csv_dicts(D_AGENTS_INDEX):
        rows.append(
            [
                "D_AGENTS_INDEX",
                row.get("agent_id") or row.get("name"),
                row.get("name"),
                row.get("primary_surface"),
                row.get("agent_id"),
                row.get("status"),
                row.get("workpaper_path"),
                row.get("level_id"),
                "Usar perfil, workpaper y routing D antes de delegar trabajo de cabina.",
                str(D_AGENTS_INDEX),
            ]
        )

    fan_in = ROOT / "operativa" / "FAN_IN_AGENTES_MAXIMO_ESTADO_REAL_20260617.csv"
    for row in read_csv_dicts(fan_in):
        rows.append(
            [
                "FAN_IN_MAXIMO_ESTADO_REAL",
                row.get("agent"),
                row.get("role"),
                row.get("surface"),
                "",
                row.get("max_real_state"),
                row.get("evidence"),
                row.get("risk"),
                row.get("next_lane"),
                str(fan_in),
            ]
        )

    sdu_map = ROOT / "dataverse" / "MAPA_AGENTES_SDU.md"
    for row in parse_md_table(sdu_map, ["Agente", "Canonical ID", "Estado"]):
        rows.append(
            [
                "DATAVERSE_SDU_ROSTER",
                row.get("Agente"),
                "sdu_runtime",
                "HUBDesarrollo / SDUCapabilityControlPlane",
                row.get("Canonical ID"),
                row.get("Estado"),
                "mon_sdu_agent_connection_mapping",
                "",
                "Actualizar mapa si aparece nuevo agente SDU.",
                str(sdu_map),
            ]
        )

    live_json = ROOT / "operativa" / "REVISION_AGENTES_LIVE_6_20260616.json"
    if live_json.exists():
        try:
            data = json.loads(live_json.read_text(encoding="utf-8"))
        except Exception:
            data = []
        for item in as_list(data):
            if not isinstance(item, dict):
                continue
            evidence = item.get("evidencia")
            rows.append(
                [
                    "REVISION_AGENTES_LIVE_6",
                    item.get("agente"),
                    item.get("agente_visible"),
                    "Corte ejecutora / workpapers",
                    "",
                    "PENDIENTE_REAL=False" if item.get("pendiente_real") is False else safe(item.get("pendiente_real")),
                    " | ".join(as_list(evidence)),
                    item.get("riesgo"),
                    item.get("siguiente_paso"),
                    str(live_json),
                ]
            )

    dv_agents = ROOT / "dataverse" / "MAPA_AGENTES.md"
    for row in parse_md_table(dv_agents, ["Component Id", "Component Type", "Resolucion"]):
        rows.append(
            [
                "DATAVERSE_COMPONENTE_RESUELTO",
                row.get("Resolucion"),
                f"componenttype:{row.get('Component Type')}",
                "PowerVirtualAgentsProd",
                row.get("Component Id"),
                "RESUELTO",
                "ApplicationUser/ApplicationUserRole mapeado.",
                "",
                "Tipificar componenttype antes de registrar nuevos ids.",
                str(dv_agents),
            ]
        )

    return rows


def discover_universe_rows() -> list[list]:
    rows = [
        ["SUPERFICIE_CANONICA", "PROJEC_CDX", str(ROOT), "workspace", "Workbook y workbench actual", "", "ACTIVO", str(ROOT / "AGENTS.md")],
        ["SUPERFICIE_CANONICA", "CODEX_ROOT", str(CODEX_ROOT), "codex_config", "Cabina global local", "", "ACTIVO", str(CODEX_ROOT / "README.md")],
        ["SUPERFICIE_CANONICA", "AGENTS_ROOT", str(AGENTS_ROOT), "agent_layer", "Skills, recipes, plugins y scripts locales", "", "ACTIVO", str(AGENTS_ROOT / "README.md")],
        ["SUPERFICIE_CANONICA", "GITHUB_ROOT", str(GITHUB_ROOT), "repos_root", "Repositorios canonicos locales", "", "ACTIVO", str(GITHUB_ROOT)],
        ["SUPERFICIE_CANONICA", "D_CABINA_UNIVERSAL", r"D:\\", "consola_rectora_gobernada", "Cabina D / consola rectora local", "", "MONTADA_GOBERNADA" if Path(r"D:\\").exists() else "OBSERVAR", r"D:\AGENTS.md|D:\MANIFEST.yaml|D:\REPO_SCOPE.md"],
        ["SUPERFICIE_CANONICA", "CODEX_CLOUD_PROJEC_CDX", "/workspace/projec-cdx", "codex_cloud", "Environment local confirmado por global-state", "", "CONFIRMADO_LOCAL", str(STATE_PATH)],
        ["SUPERFICIE_CANONICA", "DATAVERSE_SDU", str(ROOT / "dataverse"), "dataverse_metadata", "Mapas y conexiones Dataverse", "", "METADATA_ONLY", str(ROOT / "dataverse" / "INDICE_DATAVERSE.md")],
        ["SUPERFICIE_CANONICA", "SESHAT_SHAREPOINT_CANON", "https://escribaniabitsch.sharepoint.com/sites/SeshatHubRegistroN.8/SitePages/Home.aspx", "sharepoint_canon", "Canon Seshat indicado por owner", "", "GOBERNADO", "orden conversacional / workbook"],
    ]
    for row in read_csv_dicts(UNIVERSE_AUDIT):
        rows.append(
            [
                "UNIVERSE_RELATIONSHIP_AUDIT",
                row.get("Universo"),
                row.get("Origen"),
                row.get("Tipo"),
                row.get("Destino"),
                row.get("Dependencia"),
                row.get("Estado"),
                row.get("Fuente"),
            ]
        )
    return rows


def discover_skill_rows() -> list[list]:
    rows: list[list] = []
    seen_paths: set[str] = set()
    for row in read_csv_dicts(SKILLS_UNIFIED_TABLE):
        source_path = row.get("SourcePath", "")
        if source_path:
            seen_paths.add(os.path.normcase(os.path.abspath(source_path)))
        rows.append(
            [
                row.get("RootLabel"),
                row.get("Kind") or "skill",
                row.get("Family"),
                row.get("Canonical"),
                row.get("Alias"),
                row.get("Purpose"),
                source_path,
                "INVENTARIO_UNIFICADO",
                row.get("Notes"),
            ]
        )

    skill_roots = [
        CODEX_ROOT / "skills",
        AGENTS_ROOT / "skills",
        CODEX_ROOT / "plugins" / "cache",
        D_CODEX_ROOT / "skills",
        D_SKILLS_ROOT,
    ]
    for root in skill_roots:
        if not root.exists():
            continue
        for path in sorted(root.rglob("SKILL.md"), key=lambda item: str(item).lower()):
            key = os.path.normcase(os.path.abspath(str(path)))
            if key in seen_paths:
                continue
            seen_paths.add(key)
            meta = skill_metadata(path)
            rows.append(
                [
                    file_label(path),
                    "skill",
                    "DESCUBIERTO_SCAN",
                    meta.get("name"),
                    path.parent.name,
                    meta.get("description"),
                    str(path),
                    "DESCUBIERTO_SCAN",
                    "Encontrado por escaneo local de SKILL.md.",
                ]
            )
    rows.sort(key=lambda item: (str(item[0]).lower(), str(item[3]).lower(), str(item[6]).lower()))
    return rows


def discover_recipe_rows() -> list[list]:
    rows: list[list] = []
    seen: set[str] = set()
    index_paths = [
        AGENTS_ROOT / "recipes" / "RECIPE_INDEX.csv",
        AGENTS_ROOT / "codex" / "recipes" / "RECIPE_INDEX.csv",
        D_RECIPE_INDEX,
    ]
    for index_path in index_paths:
        for row in read_csv_dicts(index_path):
            raw_path = row.get("path", "")
            recipe_path = path_from_cell(raw_path) or (index_path.parent / raw_path)
            key = os.path.normcase(os.path.abspath(str(recipe_path)))
            seen.add(key)
            rows.append(
                [
                    file_label(index_path),
                    row.get("recipe_id") or row.get("name") or Path(raw_path).stem,
                    row.get("kind") or row.get("level_id") or "indexed_recipe",
                    str(recipe_path),
                    row.get("size_bytes"),
                    "",
                    "INDEXADO",
                    row.get("output") or "RECIPE_INDEX.csv",
                    str(index_path),
                ]
            )

    recipe_roots = [ROOT / "recipes", AGENTS_ROOT / "recipes", AGENTS_ROOT / "codex" / "recipes", D_CODEX_ROOT / "recipes"]
    for root in recipe_roots:
        if not root.exists():
            continue
        for path in sorted(root.glob("*.md"), key=lambda item: item.name.lower()):
            key = os.path.normcase(os.path.abspath(str(path)))
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                [
                    file_label(path),
                    path.stem,
                    "markdown_recipe",
                    str(path),
                    path.stat().st_size,
                    path.stat().st_mtime,
                    "DESCUBIERTO_SCAN",
                    first_markdown_heading(path) or first_meaningful_line(path),
                    str(path),
                ]
            )
    rows.sort(key=lambda item: (str(item[0]).lower(), str(item[1]).lower(), str(item[3]).lower()))
    return rows


def discover_tool_rows(config: dict) -> list[list]:
    rows: list[list] = []
    tool_roots = [
        ROOT / "tools",
        AGENTS_ROOT / "scripts",
        CODEX_ROOT / "environment",
        Path(r"D:\.agents\codex\tools"),
    ]
    allowed_suffixes = {".ps1", ".py", ".sh", ".mjs", ".js", ".cmd", ".bat"}
    seen: set[str] = set()
    for root in tool_roots:
        if not root.exists():
            continue
        for path in sorted(root.rglob("*"), key=lambda item: str(item).lower()):
            if not path.is_file() or path.suffix.lower() not in allowed_suffixes:
                continue
            key = os.path.normcase(os.path.abspath(str(path)))
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                [
                    file_label(path),
                    path.name,
                    path.suffix.lower().lstrip("."),
                    str(path),
                    path.stat().st_size,
                    path.stat().st_mtime,
                    "LOCAL_TOOL",
                    first_meaningful_line(path),
                ]
            )

    for name in sorted((config.get("mcp_servers") or {}).keys()):
        rows.append(["config.toml", name, "mcp_server", str(CONFIG_PATH), "", "", "CONFIGURADO_LOCAL", "Servidor MCP configurado."])
    for name, item in sorted((config.get("plugins") or {}).items()):
        if isinstance(item, dict) and item.get("enabled") is True:
            rows.append(["config.toml", name, "plugin_enabled", str(CONFIG_PATH), "", "", "ENABLED_LOCAL", "Plugin habilitado en Codex local."])
    rows.sort(key=lambda item: (str(item[2]).lower(), str(item[0]).lower(), str(item[1]).lower()))
    return rows


def discover_connection_rows(config: dict, env: dict) -> list[list]:
    rows: list[list] = []
    surface_path = CONNECTION_MATRIX_DIR / "CONNECTION_SURFACE_INVENTORY.csv"
    for row in read_csv_dicts(surface_path):
        rows.append(
            [
                "CONNECTION_SURFACE_INVENTORY",
                "surface",
                row.get("surface"),
                row.get("provider"),
                row.get("gate_required"),
                row.get("status"),
                row.get("live_enabled"),
                "",
                "",
                f"instances={row.get('instance_count')} scopes={row.get('tenant_scopes')} risk={row.get('highest_risk')}",
                "",
                str(surface_path),
            ]
        )

    gate_path = CONNECTION_MATRIX_DIR / "CONNECTION_GATE_MATRIX.csv"
    for row in read_csv_dicts(gate_path):
        rows.append(
            [
                "CONNECTION_GATE_MATRIX",
                "gate",
                row.get("surface"),
                row.get("provider"),
                row.get("gate_id"),
                row.get("status"),
                "no",
                "",
                "",
                f"allowed={row.get('allowed_actions')} blocked={row.get('blocked_actions')}",
                row.get("stop_condition"),
                str(gate_path),
            ]
        )

    mapping_path = CONNECTION_MATRIX_DIR / "AGENT_CONNECTION_MAPPING.csv"
    for row in read_csv_dicts(mapping_path):
        rows.append(
            [
                "AGENT_CONNECTION_MAPPING",
                "agent_repo_mapping",
                row.get("repo"),
                "mixed",
                row.get("gate"),
                row.get("status"),
                "no",
                row.get("owner_agent"),
                row.get("reviewer_agent"),
                row.get("surfaces"),
                row.get("stop_condition"),
                str(mapping_path),
            ]
        )

    seed_surface_path = DATAVERSE_CONNECTION_DATA / "seed_connection_surfaces.csv"
    for row in read_csv_dicts(seed_surface_path):
        rows.append(
            [
                "DATAVERSE_SEED_CONNECTION_SURFACES",
                "dataverse_seed_surface",
                row.get("surface") or row.get("surface_id"),
                row.get("provider"),
                row.get("gate_required") or row.get("gate"),
                row.get("status"),
                row.get("live_enabled"),
                "",
                "",
                safe(row),
                "",
                str(seed_surface_path),
            ]
        )

    for name in sorted((config.get("mcp_servers") or {}).keys()):
        rows.append(["config.toml", "mcp_server", name, "Codex/local", "GATE_CONNECTOR_REGISTRY_NO_SECRET", "CONFIGURADO_LOCAL", "no", "", "", "MCP server local configurado.", "connection_gate_missing", str(CONFIG_PATH)])
    for name, item in sorted((config.get("plugins") or {}).items()):
        if isinstance(item, dict) and item.get("enabled") is True:
            rows.append(["config.toml", "plugin_surface", name, "Codex UI", "GATE_CONNECTOR_REGISTRY_NO_SECRET", "ENABLED_LOCAL", "no", "", "", "Plugin habilitado; surface seleccionada por tarea.", "", str(CONFIG_PATH)])

    env_label = env.get("label")
    if env_label:
        rows.append(["global-state", "codex_cloud_env", env_label, "OpenAI Codex Cloud", "GATE_OPENAI_LIVE_GOVERNED_ORDER", "CONFIRMADO_LOCAL", safe((env.get("agent_network_access") or {}).get("mode")), "", "", f"id={env.get('id')} workspace={env.get('workspace_dir')}", "", str(STATE_PATH)])

    return rows


def discover_dataverse_source_rows() -> list[list]:
    rows: list[list] = []
    source_map = ROOT / "dataverse" / "DATAVERSE_OPERATIONAL_CHAIN_SOURCE_MAP.csv"
    for row in read_csv_dicts(source_map):
        rows.append(
            [
                "DATAVERSE_OPERATIONAL_CHAIN_SOURCE_MAP",
                row.get("canonical_role"),
                row.get("logical_surface"),
                row.get("entity_set_name"),
                row.get("solution_unique_name"),
                row.get("live_table_status"),
                row.get("row_status"),
                row.get("functional_source_table"),
                row.get("evidence_source"),
                row.get("notes"),
                str(source_map),
            ]
        )

    live_read = ROOT / "operativa" / "DATAVERSE_REHIDRATACION_LIVE_READ_20260617.json"
    if live_read.exists():
        try:
            data = json.loads(live_read.read_text(encoding="utf-8"))
        except Exception:
            data = {}
        for item in as_list(data.get("nomenclature")):
            if not isinstance(item, dict):
                continue
            rows.append(
                [
                    "DATAVERSE_REHIDRATACION_LIVE_READ_NOMENCLATURE",
                    item.get("role"),
                    item.get("logical_name"),
                    item.get("entity_set_name"),
                    data.get("environment_url"),
                    data.get("live_state"),
                    f"object_type_code={item.get('object_type_code')}",
                    item.get("schema_name"),
                    item.get("primary_id_attribute"),
                    f"primary_name={item.get('primary_name_attribute')} ownership={item.get('ownership_type')}",
                    str(live_read),
                ]
            )
        for pair in as_list(data.get("pairs")):
            if not isinstance(pair, dict):
                continue
            rows.append(
                [
                    "DATAVERSE_REHIDRATACION_LIVE_READ_PAIR_SOURCE",
                    pair.get("label"),
                    pair.get("source_entity_logical_name"),
                    pair.get("source_entity_set_name"),
                    data.get("environment_url"),
                    pair.get("source_mon_status"),
                    f"count={pair.get('source_count')} id={pair.get('source_id')}",
                    pair.get("source_canonical"),
                    pair.get("source_statecode_label"),
                    "live GET read-only; source row confirmed 1/1" if pair.get("source_count") == 1 else "source row count requires review",
                    str(live_read),
                ]
            )
            rows.append(
                [
                    "DATAVERSE_REHIDRATACION_LIVE_READ_PAIR_EVIDENCE",
                    pair.get("label"),
                    pair.get("evidence_entity_logical_name"),
                    pair.get("evidence_entity_set_name"),
                    data.get("environment_url"),
                    pair.get("evidence_mon_status"),
                    f"count={pair.get('evidence_count')} id={pair.get('evidence_id')}",
                    pair.get("evidence_canonical"),
                    pair.get("evidence_statecode_label"),
                    "live GET read-only; evidence row confirmed 1/1" if pair.get("evidence_count") == 1 else "evidence row count requires review",
                    str(live_read),
                ]
            )

    for selection_path in sorted((ROOT / "operativa").glob("DATAVERSE_LIVE_ROWS_CONSUMER_SELECTED_*.csv")):
        for row in read_csv_dicts(selection_path):
            rows.append(
                [
                    "DATAVERSE_LIVE_ROWS_CONSUMER_SELECTED",
                    row.get("decision_id"),
                    row.get("consumer"),
                    "workbook",
                    row.get("target"),
                    row.get("estado"),
                    row.get("evidence"),
                    row.get("source"),
                    row.get("postcheck"),
                    (
                        f"owner={row.get('owner')} next_delta={row.get('next_delta')} "
                        f"live_writes={row.get('dataverse_write_executed')}/"
                        f"{row.get('microsoft_live_write_executed')}"
                    ),
                    str(selection_path),
                ]
            )

    summary_path = ROOT / "operativa" / "DATAVERSE_POWER_PLATFORM_LIVE_SUMMARY_20260616.json"
    if summary_path.exists():
        try:
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
        except Exception:
            summary = {}
        for key in [
            "solutions_count",
            "workflows_count",
            "bots_count",
            "workqueues_count",
            "workqueueitems_count",
        ]:
            if key in summary:
                rows.append(
                    [
                        "DATAVERSE_POWER_PLATFORM_LIVE_SUMMARY",
                        key,
                        "",
                        "",
                        summary.get("environment_name") or summary.get("environment_url"),
                        "OBSERVED_READ_ONLY",
                        summary.get(key),
                        "",
                        "",
                        "Resumen vivo previo; no refrescado en esta corrida.",
                        str(summary_path),
                    ]
                )
        for solution in as_list(summary.get("custom_or_operational_solutions")):
            if not isinstance(solution, dict):
                continue
            rows.append(
                [
                    "DATAVERSE_POWER_PLATFORM_SOLUTIONS",
                    solution.get("solution_unique_name"),
                    "solution",
                    "",
                    summary.get("environment_name") or summary.get("environment_url"),
                    "OBSERVED_READ_ONLY",
                    solution.get("version"),
                    solution.get("solution_id"),
                    solution.get("friendly_name"),
                    f"is_managed={solution.get('is_managed')}",
                    str(summary_path),
                ]
            )
    return rows


def load_json_file(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def discover_environment_solution_rows() -> list[list]:
    rows: list[list] = []
    summary = load_json_file(DATAVERSE_LIVE_SUMMARY)
    inventory = load_json_file(DATAVERSE_LIVE_INVENTORY)
    environment = summary.get("environment_name") or inventory.get("environment") or "HUBDesarrollo"
    environment_url = summary.get("environment_url") or inventory.get("environment") or ""
    environment_id = summary.get("environment_id") or ""

    if summary:
        rows.append(
            [
                "DATAVERSE_LIVE_SUMMARY",
                environment,
                environment_url,
                environment_id,
                "ENVIRONMENT",
                "HUBDesarrollo",
                "",
                "",
                "",
                "",
                "OBSERVED_READ_ONLY",
                f"solutions={summary.get('solutions_count')} bots={summary.get('bots_count')} queues={summary.get('sdu_queue_count')}",
                str(DATAVERSE_LIVE_SUMMARY),
            ]
        )

    seen: set[str] = set()
    for row in read_csv_dicts(DATAVERSE_SOLUTIONS_CSV):
        key = row.get("solution_unique_name") or row.get("solution_id")
        if key:
            seen.add(key)
        rows.append(
            [
                "DATAVERSE_SOLUTIONS_LIVE_CSV",
                environment,
                environment_url,
                environment_id,
                "SOLUTION",
                row.get("solution_unique_name"),
                row.get("friendly_name"),
                row.get("version"),
                row.get("is_managed"),
                "",
                row.get("classification"),
                f"created={row.get('created_on')} modified={row.get('modified_on')}",
                str(DATAVERSE_SOLUTIONS_CSV),
            ]
        )

    for solution in as_list(inventory.get("solutions")):
        if not isinstance(solution, dict):
            continue
        key = solution.get("uniquename")
        if key in seen:
            continue
        rows.append(
            [
                "DATAVERSE_LIVE_INVENTORY_SOLUTIONS",
                environment,
                environment_url,
                environment_id,
                "SOLUTION",
                key,
                solution.get("friendlyname"),
                solution.get("version"),
                solution.get("ismanaged"),
                solution.get("component_count"),
                "OBSERVED_READ_ONLY",
                "Solucion presente en inventario live amplio.",
                str(DATAVERSE_LIVE_INVENTORY),
            ]
        )

    for item in as_list(inventory.get("table_counts")):
        if not isinstance(item, dict):
            continue
        rows.append(
            [
                "DATAVERSE_LIVE_INVENTORY_TABLE_COUNTS",
                environment,
                environment_url,
                environment_id,
                "TABLE_COUNT",
                item.get("entity_set"),
                "",
                "",
                "",
                item.get("count"),
                item.get("status"),
                "Conteo de tabla confirmado por lectura previa.",
                str(DATAVERSE_LIVE_INVENTORY),
            ]
        )
    return rows


def discover_environment_agent_rows() -> list[list]:
    rows: list[list] = []
    summary = load_json_file(DATAVERSE_LIVE_SUMMARY)
    inventory = load_json_file(DATAVERSE_LIVE_INVENTORY)
    environment = summary.get("environment_name") or inventory.get("environment") or "HUBDesarrollo"
    environment_url = summary.get("environment_url") or inventory.get("environment") or ""
    environment_id = summary.get("environment_id") or ""

    for item in as_list(inventory.get("agent_rows")):
        if not isinstance(item, dict):
            continue
        rows.append(
            [
                "DATAVERSE_AGENT_CONNECTION_MAPPING",
                environment,
                environment_url,
                environment_id,
                "SDUCapabilityControlPlane",
                "mon_sdu_agent_connection_mappings",
                item.get("mon_display_name"),
                item.get("mon_canonical_id"),
                item.get("mon_owner_agent"),
                item.get("mon_reviewer_agent"),
                item.get("mon_surfaces"),
                item.get("mon_status"),
                item.get("mon_last_reconciled_at"),
                item.get("mon_sdu_agent_connection_mappingid"),
                item.get("mon_risk_level"),
                str(DATAVERSE_LIVE_INVENTORY),
            ]
        )

    for row in read_csv_dicts(DATAVERSE_BOTS_CSV):
        rows.append(
            [
                "DATAVERSE_BOTS_LIVE_CSV",
                environment,
                environment_url,
                environment_id,
                "PowerPlatformBots",
                "bots",
                row.get("name"),
                row.get("schema_name"),
                "",
                "",
                "Copilot/Bot",
                row.get("classification"),
                row.get("modified_on"),
                row.get("botid"),
                f"state={row.get('statecode')} status={row.get('statuscode')}",
                str(DATAVERSE_BOTS_CSV),
            ]
        )

    return rows


def discover_workqueue_rows() -> list[list]:
    rows: list[list] = []
    summary = load_json_file(DATAVERSE_LIVE_SUMMARY)
    environment = summary.get("environment_name") or "HUBDesarrollo"
    environment_url = summary.get("environment_url") or ""
    for row in read_csv_dicts(DATAVERSE_WORKQUEUES_CSV):
        backlog = row.get("backlog_total_items")
        try:
            backlog_number = int(backlog or 0)
        except ValueError:
            backlog_number = 0
        action = "DELTA_PRIORIZAR_BACKLOG" if backlog_number > 0 else "OK_SIN_BACKLOG"
        rows.append(
            [
                "DATAVERSE_WORKQUEUES_LIVE_CSV",
                environment,
                environment_url,
                row.get("queue_name"),
                row.get("workqueueid"),
                row.get("statecode"),
                row.get("statuscode"),
                row.get("retry_limit"),
                row.get("requeue_limit"),
                row.get("prioritytype"),
                row.get("workqueuetype"),
                backlog_number,
                row.get("classification"),
                action,
                str(DATAVERSE_WORKQUEUES_CSV),
            ]
        )
    return rows


def discover_atomic_action_rows() -> list[list]:
    rows: list[list] = []
    for row in read_csv_dicts(ATOMIC_ACTION_MATRIX):
        rows.append(
            [
                row.get("action_type"),
                row.get("purpose"),
                row.get("target_surface"),
                row.get("mutating"),
                row.get("requires_lock"),
                row.get("rollback_or_compensation"),
                row.get("validation_focus"),
                row.get("default_terminal_state"),
                row.get("notes"),
                str(ATOMIC_ACTION_MATRIX),
            ]
        )
    return rows


def discover_master_atomic_matrix_rows() -> list[list]:
    rows: list[list] = []
    for row in read_csv_dicts(MASTER_ATOMIC_MATRIX):
        rows.append(
            [
                row.get("wave_id"),
                row.get("agente_owner"),
                row.get("agente_reviewer"),
                row.get("owner_humano"),
                row.get("decision_humana_required"),
                row.get("approval_status"),
                row.get("execution_actor"),
                row.get("surface"),
                row.get("read_scope"),
                row.get("write_scope"),
                row.get("lock_key"),
                row.get("estado_operativo"),
                row.get("audit_status"),
                row.get("environment_classification"),
                row.get("permission_scope"),
                row.get("candidate_count"),
                row.get("allowed_action"),
                row.get("blocked_action"),
                row.get("evidencia"),
                row.get("evidence_sink"),
                row.get("validador"),
                row.get("postcheck"),
                row.get("rollback"),
                row.get("stop_condition"),
                row.get("next_delta"),
                str(MASTER_ATOMIC_MATRIX),
            ]
        )
    return rows


def discover_metadata_hydration_rows() -> list[list]:
    rows: list[list] = []
    for row in read_csv_dicts(METADATA_HYDRATION_MATRIX):
        rows.append(
            [
                row.get("wave_id"),
                row.get("metadata_id"),
                row.get("tenant_id"),
                row.get("environment_name"),
                row.get("surface"),
                row.get("source_artifact"),
                row.get("target_entity_set"),
                row.get("target_logical_name"),
                row.get("canonical_id"),
                row.get("display_name"),
                row.get("classification"),
                row.get("status"),
                row.get("owner"),
                row.get("evidence_ref"),
                row.get("rollback"),
                row.get("postcheck"),
                row.get("validator"),
                row.get("stop_condition"),
                row.get("notes"),
                str(METADATA_HYDRATION_MATRIX),
            ]
        )
    return rows


def discover_d_surface_rows(scan_time: str) -> list[list]:
    surfaces = [
        (D_ROOT, "D root wrapper", False),
        (D_CODEX_ROOT, "Capa local de agentes Codex", True),
        (D_AUTHORITY_CANON, "Autoridad y canon", True),
        (D_CODEX_ROOT / "matrices", "Matrices operativas", True),
        (D_CODEX_ROOT / "maps", "Mapas", True),
        (D_CODEX_ROOT / "agents", "Perfiles y subniveles de agentes", True),
        (D_CODEX_ROOT / "workpapers", "Workpapers por agente", True),
        (D_CODEX_ROOT / "readbacks", "Readbacks y evidencia", True),
        (D_CODEX_ROOT / "orders", "Órdenes gobernadas", True),
        (D_CODEX_ROOT / "recipes", "Recetas", True),
        (D_CODEX_ROOT / "skills", "Skills ligadas a Codex D", True),
        (D_SKILLS_ROOT, "Skills portables D", True),
        (D_CODEX_ROOT / "tools", "Tools, validators y runners", True),
        (D_CODEX_ROOT / "logs", "Logs esperados/no presentes", True),
    ]
    rows: list[list] = []
    for path, notes, recursive in surfaces:
        exists = path.exists()
        dirs, files, extensions = dir_counts(path, recursive=recursive) if exists else (None, None, "")
        if exists:
            status = "AVAILABLE_READONLY"
        elif path.name.lower() == "logs":
            status = "MISSING_EXPECTED_OR_NOT_APPLICABLE"
        else:
            status = "MISSING"
        rows.append(
            [
                str(path),
                "SI" if exists else "NO",
                dirs,
                files,
                extensions,
                scan_time,
                status,
                notes,
            ]
        )
    return rows


def discover_local_surface_rows(scan_time: str, d_status: str) -> list[list]:
    surfaces = [
        (
            D_ROOT,
            "D_CABINA_RECTOR_LOCAL",
            "consola_rectora",
            "Cabina Universal D; autoridad/gobierno heredado y wrapper repo.",
            "Leer AGENTS/MANIFEST/REPO_SCOPE; no escribir sin gate.",
            False,
        ),
        (
            ROOT,
            "PROJEC_CDX_WORKBENCH",
            "workbench_actual",
            "Workbench actual de cierre, decision y workbook.",
            "Entrada viva de este hilo; versionar por rama codex/*.",
            True,
        ),
        (
            CODEX_ROOT,
            "CODEX_RUNTIME_CONFIG",
            "runtime_config_memoria",
            "Config local, skills, memoria, sesiones y runtime Codex.",
            "No tocar secretos ni runtime interno por inferencia.",
            False,
        ),
        (
            HOME_CODEX_LOCAL,
            "HOME_CODEXLOCAL_LIVE_ENTRY",
            "entrada_viva_liviana",
            "Nueva entrada liviana/puente local; todavia incompleta.",
            "Debe tener README/MAPA/AGENTS/indice antes de operar como raiz principal.",
            True,
        ),
        (
            DOCUMENTS_CODEX_LOCAL,
            "DOCUMENTS_CODEXLOCAL_LEGACY_HEAVY",
            "workspace_legado_pesado",
            "Workspace local pesado historico con indices y evidencia.",
            "No mover 14GB+ sin paquete de migracion y rollback.",
            True,
        ),
        (
            DOCUMENTS_CODEX,
            "DOCUMENTS_CODEX_CHRONOLOGY",
            "cronologia_documental",
            "Cronologia documental por fecha, handoffs y rastros de sesiones.",
            "Fuente de recuperacion historica; no canon operativo.",
            True,
        ),
        (
            DOCUMENTS_CODEX_ARCHIVES,
            "DOCUMENTS_CODEX_ARCHIVES",
            "archivo_zip_reversible",
            "Archivo frio reversible de zips y snapshots.",
            "Conservar; no extraer ni borrar sin orden.",
            True,
        ),
        (
            GITHUB_ROOT,
            "DOCUMENTS_GITHUB_CANONICAL_REPOS",
            "repositorios_canonicos",
            "Raiz canonica local de repos GitHub.",
            "Los repos conservan sus propios remotos y ramas.",
            False,
        ),
    ]
    rows: list[list] = []
    for path, surface_id, surface_class, role, rule, recursive in surfaces:
        exists = path.exists()
        dirs, files, bytes_total, extensions = dir_file_stats(path, recursive=recursive) if exists else (None, None, None, "")
        has_readme = (path / "README.md").exists() if exists else False
        has_agents = (path / "AGENTS.md").exists() if exists else False
        has_map = any((path / name).exists() for name in ("MAPA.md", "MAPA_MAESTRO.md")) if exists else False
        has_index = any((path / name).exists() for name in ("INDICE.csv", "CODEX_INDEX.csv", "CODEXLOCAL_INDEX.csv")) if exists else False
        has_git = (path / ".git").exists() if exists else False
        if not exists:
            status = "MISSING"
        elif path == D_ROOT:
            status = d_status
        elif path == HOME_CODEX_LOCAL and not has_readme:
            status = "ENTRYPOINT_INCOMPLETE"
        elif path == DOCUMENTS_CODEX_LOCAL:
            status = "LEGACY_HEAVY_INDEXED"
        elif path == DOCUMENTS_CODEX:
            status = "CHRONOLOGY_INDEXED"
        elif path == DOCUMENTS_CODEX_ARCHIVES:
            status = "ARCHIVE_REVERSIBLE"
        elif has_readme or has_map or has_index:
            status = "INDEXED"
        else:
            status = "OBSERVED"
        rows.append(
            [
                surface_id,
                str(path),
                surface_class,
                status,
                "SI" if exists else "NO",
                "SI" if has_readme else "NO",
                "SI" if has_agents else "NO",
                "SI" if has_map else "NO",
                "SI" if has_index else "NO",
                "SI" if has_git else "NO",
                dirs,
                files,
                kb(bytes_total) if bytes_total is not None else None,
                extensions,
                scan_time,
                role,
                rule,
            ]
        )
    return rows


def discover_workspace_rows(scan_time: str, d_status: str) -> list[list]:
    branch = git_value(ROOT, ["branch", "--show-current"]) or "DETACHED_OR_UNKNOWN"
    head = git_value(ROOT, ["rev-parse", "--short", "HEAD"]) or "UNKNOWN"
    remote = git_value(ROOT, ["remote", "get-url", "origin"]) or "NO_REMOTE"
    upstream = git_value(ROOT, ["rev-parse", "--abbrev-ref", "--symbolic-full-name", "@{u}"]) or "NO_UPSTREAM"
    status = git_value(ROOT, ["status", "--short", "--branch"]) or "UNKNOWN"
    dirty = "SI" if any(line and not line.startswith("##") for line in status.splitlines()) else "NO"
    rows = [
        ["workspace.root", str(ROOT), "ACTUAL", "Workspace operativo de este hilo."],
        ["workspace.cwd_when_built", str(Path.cwd()), "INFO", "Directorio desde donde corre el builder."],
        ["workspace.branch", branch, "ACTUAL", "Rama de trabajo vigente."],
        ["workspace.head", head, "INFO", "Commit corto del workspace."],
        ["workspace.remote", remote, "INFO", "Remoto origin del repo actual."],
        ["workspace.upstream", upstream, "INFO", "Upstream de la rama actual."],
        ["workspace.dirty", dirty, "OBSERVAR" if dirty == "SI" else "PASS", "Estado Git antes/despues de regenerar workbook."],
        ["workspace.status", status.replace("\n", " | "), "INFO", "Resumen git status --short --branch."],
        ["workspace.workbook", str(WORKBOOK), "ACTUAL", "Workbook de decision vigente."],
        ["workspace.builder", str(ROOT / "tools" / "update_codex_config_workbook.py"), "ACTUAL", "Builder idempotente del workbook."],
        ["workspace.current", str(ROOT / "operativa" / "CURRENT.md"), "ACTUAL", "Estado vivo de PROJEC CDX."],
        ["workspace.next", str(ROOT / "operativa" / "NEXT.md"), "ACTUAL", "Movimiento unico."],
        ["workspace.trace", str(ROOT / "operativa" / "TRACE.md"), "ACTUAL", "Trazabilidad activa."],
        ["workspace.codex_root", str(CODEX_ROOT), "REFERENCIA", "Runtime/config/memoria local Codex."],
        ["workspace.d_root", str(D_ROOT), d_status, "Consola rectora local D."],
        ["workspace.home_codexlocal", str(HOME_CODEX_LOCAL), "REFERENCIA", "Entrada viva liviana nueva."],
        ["workspace.documents_codexlocal", str(DOCUMENTS_CODEX_LOCAL), "REFERENCIA", "Workspace legado pesado."],
        ["workspace.documents_codex", str(DOCUMENTS_CODEX), "REFERENCIA", "Cronologia documental."],
        ["workspace.documents_codex_archives", str(DOCUMENTS_CODEX_ARCHIVES), "REFERENCIA", "Archivo reversible ZIP."],
        ["workspace.github_root", str(GITHUB_ROOT), "REFERENCIA", "Repositorios canonicos locales."],
        ["workspace.codex_worktree_env", os.environ.get("CODEX_WORKTREE_PATH", "N/D"), "INFO", "Variable de worktree si existe."],
        ["workspace.updated_at", scan_time, "INFO", "Timestamp local de regeneracion."],
    ]
    return rows


def discover_d_indexes_rows() -> list[list]:
    rows: list[list] = []
    roots = [
        D_CODEX_ROOT,
        D_CODEX_ROOT / "matrices",
        D_CODEX_ROOT / "recipes",
        D_CODEX_ROOT / "skills",
        D_CODEX_ROOT / "tools",
        D_CODEX_ROOT / "workpapers",
        D_AUTHORITY_CANON,
    ]
    seen: set[str] = set()
    for root in roots:
        if not root.exists():
            continue
        for path in sorted(root.rglob("*.csv"), key=lambda item: str(item).lower()):
            if not any(token in path.name.upper() for token in ["INDEX", "MATRIX", "REGISTRY", "MAP", "CANON", "POLICY", "GATE", "ORDER"]):
                continue
            key = os.path.normcase(os.path.abspath(str(path)))
            if key in seen:
                continue
            seen.add(key)
            shape_rows, columns = csv_shape(path)
            consumed = "NO"
            if path in {
                D_MATRIX_INDEX,
                D_RECIPE_INDEX,
                D_SKILL_USAGE_MATRIX,
                D_TOOL_INDEX,
                D_AGENTS_INDEX,
                D_LEVELS_CSV,
                D_AGENT_WORKPAPERS_MATRIX,
                D_GOVERNED_ASSET_INVENTORY,
            }:
                consumed = "SI"
            if path.name in {
                "WORKPAPER_INDEX.csv",
                "ORDER_PREPARATION_ASSIGNMENT_MATRIX.csv",
                "EVIDENCE_READBACK_REGISTRY_20260603.csv",
                "AGENT_TOOL_RECIPE_SKILL_MATRIX.csv",
                "SUBSKILL_USAGE_MATRIX.csv",
                "SUBRECIPE_INDEX.csv",
                "SKILL_METADATA_QUALITY_MATRIX.csv",
            }:
                consumed = "SI"
            rows.append(
                [
                    str(root),
                    path.name,
                    str(path),
                    "csv_index_or_matrix",
                    shape_rows,
                    columns,
                    path.stat().st_size,
                    file_modified(path),
                    consumed,
                    "MODELED_OR_NAVIGABLE" if consumed == "SI" else "PENDING_MODELING",
                ]
            )
    return rows


def discover_d_matrix_index_rows() -> list[list]:
    rows: list[list] = []
    for row in read_csv_dicts(D_MATRIX_INDEX):
        raw_path = row.get("path", "")
        path = path_from_cell(raw_path)
        shape_rows, columns = csv_shape(path)
        rows.append(
            [
                row.get("matrix_id"),
                raw_path,
                "SI" if path and path.exists() else "NO",
                shape_rows,
                columns,
                row.get("scope"),
                row.get("primary_reader"),
                row.get("update_rule"),
                "D_MATRIX_INDEX",
                str(D_MATRIX_INDEX),
            ]
        )
    return rows


def discover_d_skill_registry_rows() -> list[list]:
    rows: list[list] = []
    physical: dict[str, list[str]] = {}
    for root in [D_CODEX_ROOT / "skills", D_SKILLS_ROOT]:
        if not root.exists():
            continue
        for path in sorted(root.rglob("SKILL.md"), key=lambda item: str(item).lower()):
            physical.setdefault(path.parent.name, []).append(str(path))
    for row in read_csv_dicts(D_SKILL_USAGE_MATRIX):
        skill_id = row.get("skill_id") or ""
        paths = physical.get(skill_id, [])
        rows.append(
            [
                row.get("source"),
                skill_id,
                skill_id,
                " | ".join(paths),
                "SI" if paths else "NO",
                row.get("assigned_level"),
                row.get("assigned_agents"),
                row.get("use_when"),
                row.get("live_boundary"),
                "DUPLICATE_PHYSICAL" if len(paths) > 1 else "",
                "MODELED_FROM_D_SKILL_USAGE_MATRIX",
                str(D_SKILL_USAGE_MATRIX),
            ]
        )
    for skill_id, paths in sorted(physical.items()):
        if any(row[1] == skill_id for row in rows):
            continue
        rows.append(
            [
                "d_physical_scan",
                skill_id,
                skill_id,
                " | ".join(paths),
                "SI",
                "",
                "",
                first_meaningful_line(Path(paths[0])),
                "local_only",
                "UNINDEXED_PHYSICAL_SKILL",
                "PENDING_REGISTRY_ALIGNMENT",
                paths[0],
            ]
        )
    return rows


def discover_d_subskill_rows() -> list[list]:
    rows: list[list] = []
    subskill_index = D_CODEX_ROOT / "skills" / "SUBSKILL_USAGE_MATRIX.csv"
    for row in read_csv_dicts(subskill_index):
        rows.append(
            [
                row.get("subskill_id") or row.get("skill_id"),
                row.get("parent_skill") or row.get("parent_skill_id"),
                row.get("assigned_agents"),
                row.get("use_when"),
                row.get("required_recipe"),
                row.get("required_tool"),
                row.get("validator"),
                row.get("stop_condition"),
                str(subskill_index),
                "MODELED_FROM_D_SUBSKILL_USAGE_MATRIX",
            ]
        )
    metadata_quality = D_CODEX_ROOT / "skills" / "SKILL_METADATA_QUALITY_MATRIX.csv"
    for row in read_csv_dicts(metadata_quality):
        rows.append(
            [
                row.get("skill_id"),
                "metadata_quality",
                row.get("assigned_agents") or row.get("owner_agent"),
                row.get("use_when") or row.get("quality_status") or row.get("status"),
                row.get("required_recipe"),
                row.get("required_tool"),
                row.get("validator"),
                row.get("stop_condition"),
                str(metadata_quality),
                "MODELED_FROM_D_SKILL_METADATA_QUALITY",
            ]
        )
    return rows


def discover_d_recipe_registry_rows() -> list[list]:
    rows: list[list] = []
    for index_path in [
        D_RECIPE_INDEX,
        D_CODEX_ROOT / "recipes" / "SUBRECIPE_INDEX.csv",
        D_CODEX_ROOT / "recipes" / "SOURCE_TCU_RECIPES_INDEX.csv",
    ]:
        for row in read_csv_dicts(index_path):
            raw_path = row.get("path", "")
            path = path_from_cell(raw_path) or (index_path.parent / raw_path if raw_path else None)
            rows.append(
                [
                    row.get("recipe_id") or row.get("subrecipe_id") or row.get("name") or row.get("id"),
                    row.get("level_id") or row.get("parent_recipe") or row.get("source"),
                    row.get("primary_agent") or row.get("assigned_agents") or row.get("owner_agent"),
                    str(path) if path else raw_path,
                    row.get("output") or row.get("purpose") or row.get("use_when"),
                    path.stat().st_size if path and path.exists() and path.is_file() else "",
                    file_modified(path),
                    str(index_path),
                    "MODELED_FROM_D_RECIPE_REGISTRY",
                ]
            )
    return rows


def discover_d_tools_registry_rows() -> list[list]:
    rows: list[list] = []
    for index_path in [
        D_TOOL_INDEX,
        D_CODEX_ROOT / "tools" / "SOURCE_TCU_TOOLS_INDEX.csv",
        D_CODEX_ROOT / "tools" / "SOURCE_TCU_TOOL_PERMISSIONS_INDEX.csv",
    ]:
        for row in read_csv_dicts(index_path):
            raw_path = row.get("path_or_command") or row.get("path") or row.get("command") or ""
            path = path_from_cell(raw_path)
            rows.append(
                [
                    row.get("tool_id") or row.get("name") or row.get("id"),
                    row.get("level_id") or row.get("source"),
                    row.get("tool_type") or row.get("kind") or row.get("permission_scope"),
                    raw_path,
                    "SI" if path and path.exists() else "NO" if path else "N/D",
                    row.get("allowed_surface") or row.get("allowed_actions"),
                    row.get("blocked_surface") or row.get("blocked_actions"),
                    str(index_path),
                    "MODELED_FROM_D_TOOL_REGISTRY",
                ]
            )
    return rows


def discover_d_validator_rows() -> list[list]:
    tool_rows = read_csv_dicts(D_TOOL_INDEX)
    linked_by_path: dict[str, list[str]] = {}
    for row in tool_rows:
        tool_id = row.get("tool_id", "")
        raw_path = row.get("path_or_command", "")
        path = path_from_cell(raw_path)
        if path:
            linked_by_path.setdefault(os.path.normcase(os.path.abspath(str(path))), []).append(tool_id)
    rows: list[list] = []
    tools_dir = D_CODEX_ROOT / "tools"
    if tools_dir.exists():
        for path in sorted(tools_dir.iterdir(), key=lambda item: item.name.lower()):
            if not path.is_file() or path.suffix.lower() not in {".ps1", ".py", ".sh", ".cmd", ".bat", ".js", ".mjs"}:
                continue
            linked = linked_by_path.get(os.path.normcase(os.path.abspath(str(path))), [])
            kind = "validator" if "validate" in path.name.lower() else "runner" if "run" in path.name.lower() else "tool_script"
            rows.append(
                [
                    path.name,
                    str(path),
                    "|".join(linked),
                    "",
                    "",
                    "SI" if linked else "NO",
                    "NO",
                    "",
                    "NOT_EXECUTED_READONLY_INVENTORY",
                    kind,
                ]
            )
    return rows


def discover_d_agent_registry_rows() -> list[list]:
    rows: list[list] = []
    agent_json = read_json_obj(D_AGENTS_JSON) or {}
    json_agents = {item.get("id"): item for item in as_list(agent_json.get("agents") if isinstance(agent_json, dict) else []) if isinstance(item, dict)}
    for row in read_csv_dicts(D_AGENTS_INDEX):
        agent_id = row.get("agent_id")
        item = json_agents.get(agent_id, {})
        rows.append(
            [
                agent_id,
                row.get("name"),
                row.get("level_id") or item.get("level_id"),
                row.get("layer") or item.get("layer"),
                row.get("primary_surface"),
                row.get("profile_path") or item.get("home_path"),
                row.get("workpaper_path") or item.get("workpapers_path"),
                item.get("mission"),
                "|".join(as_list(item.get("default_skills"))),
                "|".join(as_list(item.get("default_recipes"))),
                "|".join(as_list(item.get("default_tools"))),
                "|".join(as_list(item.get("default_plugins"))),
                row.get("status") or item.get("runtime_alignment_mode"),
                str(D_AGENTS_INDEX),
            ]
        )
    return rows


def discover_d_levels_rows() -> list[list]:
    rows: list[list] = []
    for row in read_csv_dicts(D_LEVELS_CSV):
        rows.append(
            [
                row.get("level_id"),
                row.get("name"),
                row.get("folder"),
                row.get("primary_role"),
                row.get("agent_count"),
                row.get("default_read"),
                str(D_LEVELS_CSV),
            ]
        )
    return rows


def discover_d_routing_rows() -> list[list]:
    rows: list[list] = []
    data = read_json_obj(D_ROUTING_JSON)
    if not isinstance(data, dict):
        return rows
    default_agent = data.get("default_agent")
    for item in as_list(data.get("routes")):
        if not isinstance(item, dict):
            continue
        rows.append(
            [
                "route",
                item.get("order_class"),
                "|".join(as_list(item.get("signals"))),
                "|".join(as_list(item.get("agents"))),
                default_agent,
                "",
                "",
                data.get("status"),
                str(D_ROUTING_JSON),
            ]
        )
    for item in as_list(data.get("handoff_rules")):
        if not isinstance(item, dict):
            continue
        rows.append(
            [
                "handoff_rule",
                item.get("when"),
                "",
                item.get("assigned_agent"),
                default_agent,
                "|".join(as_list(item.get("must_include"))),
                "",
                data.get("status"),
                str(D_ROUTING_JSON),
            ]
        )
    return rows


def discover_d_workpaper_rows() -> list[list]:
    rows: list[list] = []
    workpaper_index = D_CODEX_ROOT / "workpapers" / "WORKPAPER_INDEX.csv"
    canonical_agents = {row.get("agent_id") for row in read_csv_dicts(D_AGENTS_INDEX)}
    file_counts: dict[str, int] = {}
    root = D_CODEX_ROOT / "workpapers"
    if root.exists():
        for child in sorted([item for item in root.iterdir() if item.is_dir()], key=lambda item: item.name.lower()):
            file_counts[child.name] = len([item for item in child.rglob("*") if item.is_file()])
    index_rows = read_csv_dicts(workpaper_index) or read_csv_dicts(D_AGENT_WORKPAPERS_MATRIX)
    for row in index_rows:
        agent_id = row.get("agent_id")
        path_text = row.get("workpaper_path") or row.get("workpapers_path")
        path = path_from_cell(path_text or "")
        rows.append(
            [
                agent_id,
                row.get("level_id"),
                path_text,
                file_counts.get(agent_id, 0),
                "SI" if agent_id in canonical_agents else "NO",
                row.get("status"),
                row.get("primary_surface"),
                row.get("purpose"),
                row.get("owner_agent"),
                row.get("required_matrices"),
                row.get("required_recipes"),
                row.get("required_tools"),
                row.get("required_validators"),
                row.get("evidence_policy"),
                row.get("validator"),
                row.get("stop_condition"),
                str(workpaper_index if workpaper_index.exists() else D_AGENT_WORKPAPERS_MATRIX),
            ]
        )
    for folder, count in sorted(file_counts.items()):
        if any(row[0] == folder for row in rows):
            continue
        rows.append(
            [
                folder,
                "",
                str(root / folder),
                count,
                "NO",
                "EXTRA_WORKPAPER_FOLDER",
                "",
                "Carpeta existe en workpapers pero no está en AGENTS_INDEX.",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "noncanonical_workpaper_folder",
                str(root / folder),
            ]
        )
    return rows


def discover_d_order_rows() -> list[list]:
    rows: list[list] = []
    order_matrix = D_CODEX_ROOT / "matrices" / "ORDER_PREPARATION_ASSIGNMENT_MATRIX.csv"
    for row in read_csv_dicts(order_matrix):
        rows.append(
            [
                row.get("order_class"),
                row.get("preparer_agent"),
                row.get("reviewer_agent"),
                row.get("approver_role"),
                row.get("canon_as_of"),
                row.get("source_authority"),
                row.get("required_fields"),
                row.get("allowed_actions"),
                row.get("blocked_actions"),
                row.get("recipe"),
                row.get("tool"),
                row.get("evidence"),
                row.get("validator"),
                row.get("expiration_rule"),
                row.get("stop_condition"),
                str(order_matrix),
            ]
        )
    orders_root = D_CODEX_ROOT / "orders"
    if orders_root.exists():
        for path in sorted([item for item in orders_root.iterdir() if item.is_file()], key=lambda item: item.name.lower()):
            rows.append(
                [
                    path.stem,
                    "",
                    "",
                    "",
                    "",
                    str(orders_root),
                    "",
                    "",
                    "",
                    "",
                    "",
                    path.name,
                    "",
                    "",
                    "order_file_inventory",
                    str(path),
                ]
            )
    return rows


def discover_d_readback_rows() -> list[list]:
    rows: list[list] = []
    registry = D_CODEX_ROOT / "matrices" / "EVIDENCE_READBACK_REGISTRY_20260603.csv"
    for row in read_csv_dicts(registry):
        rows.append(
            [
                row.get("evidence_id"),
                row.get("readback_path"),
                row.get("source_repo"),
                row.get("repo_id"),
                row.get("lane_id"),
                row.get("agent_id"),
                row.get("validator"),
                row.get("status"),
                row.get("proves"),
                row.get("does_not_prove"),
                row.get("next_gate"),
                row.get("evidence_category"),
                row.get("canonical_weight"),
                str(registry),
            ]
        )
    readbacks_root = D_CODEX_ROOT / "readbacks"
    if readbacks_root.exists():
        registered = {str(row[1]).replace("/", "\\").lower() for row in rows}
        for path in sorted([item for item in readbacks_root.rglob("*") if item.is_file()], key=lambda item: str(item).lower()):
            normalized = str(path).replace("/", "\\").lower()
            if normalized in registered:
                continue
            rows.append(
                [
                    path.stem,
                    str(path),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "FILE_ONLY_NOT_IN_REGISTRY",
                    first_markdown_heading(path) if path.suffix.lower() == ".md" else path.name,
                    "",
                    "registry_alignment_review",
                    "SUPPORT",
                    "low",
                    str(path),
                ]
            )
    return rows


def discover_d_agent_tool_skill_rows() -> list[list]:
    rows: list[list] = []
    path = D_CODEX_ROOT / "matrices" / "AGENT_TOOL_RECIPE_SKILL_MATRIX.csv"
    for row in read_csv_dicts(path):
        rows.append(
            [
                row.get("agent_id"),
                row.get("level_id"),
                row.get("purpose"),
                row.get("surface"),
                row.get("skill_refs"),
                row.get("recipe_refs"),
                row.get("tool_refs"),
                row.get("plugin_refs"),
                row.get("workpaper_path"),
                row.get("validator"),
                row.get("default_mode"),
                row.get("stop_condition"),
                str(path),
            ]
        )
    return rows


def discover_d_authority_rows() -> list[list]:
    rows: list[list] = []
    if not D_AUTHORITY_CANON.exists():
        return rows
    for path in sorted([item for item in D_AUTHORITY_CANON.rglob("*") if item.is_file()], key=lambda item: str(item).lower()):
        shape_rows, columns = csv_shape(path)
        kind = "csv_matrix" if path.suffix.lower() == ".csv" else "markdown_canon" if path.suffix.lower() == ".md" else path.suffix.lower().lstrip(".")
        rows.append(
            [
                path.name,
                kind,
                str(path),
                path.stat().st_size,
                file_modified(path),
                shape_rows,
                columns,
                first_markdown_heading(path) if path.suffix.lower() == ".md" else first_meaningful_line(path),
                "D_AUTHORITY_CANON",
            ]
        )
    return rows


def discover_d_governed_asset_rows() -> list[list]:
    rows: list[list] = []
    for row in read_csv_dicts(D_GOVERNED_ASSET_INVENTORY):
        rows.append(
            [
                row.get("asset_class"),
                row.get("asset_id"),
                row.get("owner_agent"),
                row.get("authority_level"),
                row.get("governing_matrix"),
                row.get("required_recipe"),
                row.get("required_tool"),
                row.get("evidence"),
                row.get("validator"),
                row.get("coverage_status"),
                row.get("stop_condition"),
                str(D_GOVERNED_ASSET_INVENTORY),
            ]
        )
    return rows


def d_candidate_artifacts() -> list[Path]:
    roots = [
        D_CODEX_ROOT / "matrices",
        D_MATRICES_ROOT,
        D_GOVERNANCE_REGISTRY,
        D_DATAVERSE_DATA,
        D_VALIDATION_ROOT,
    ]
    suffixes = {".csv", ".md", ".json", ".yaml", ".yml", ".txt"}
    candidates: list[Path] = []
    seen: set[str] = set()
    for root in roots:
        if not root.exists():
            continue
        for path in sorted(root.rglob("*"), key=lambda item: str(item).lower()):
            if not path.is_file() or path.suffix.lower() not in suffixes:
                continue
            lowered = str(path).lower()
            if any(part in lowered for part in ["\\.git\\", "\\node_modules\\", "\\.venv\\", "\\__pycache__\\"]):
                continue
            key = os.path.normcase(os.path.abspath(str(path)))
            if key in seen:
                continue
            seen.add(key)
            candidates.append(path)
    return candidates


def classify_d_artifact(path: Path) -> tuple[str, str, str, str]:
    name = path.name.upper()
    raw = str(path)
    lowered = raw.lower()
    if "SOURCE_" in name or path.parent.name.lower().startswith("source"):
        artifact_type = "source_ref"
    elif "LEDGER" in name:
        artifact_type = "ledger"
    elif "REGISTRY" in name:
        artifact_type = "registry"
    elif "INDEX" in name:
        artifact_type = "index"
    elif "MATRIX" in name:
        artifact_type = "matrix"
    elif "SEED" in name:
        artifact_type = "seed"
    elif "VALIDATION" in name or "\\validation\\" in lowered:
        artifact_type = "validation_report"
    else:
        artifact_type = "source_ref" if path.suffix.lower() in {".md", ".json", ".yaml", ".yml"} else "csv_source"

    if D_MATRICES_ROOT.exists() and (path == D_MATRICES_ROOT or D_MATRICES_ROOT in path.parents):
        try:
            relative = path.relative_to(D_MATRICES_ROOT)
            domain = relative.parts[0] if len(relative.parts) > 1 else "matrices_root"
        except Exception:
            domain = "matrices"
    elif D_GOVERNANCE_REGISTRY.exists() and (path == D_GOVERNANCE_REGISTRY or D_GOVERNANCE_REGISTRY in path.parents):
        domain = "governance_registry"
    elif D_DATAVERSE_DATA.exists() and (path == D_DATAVERSE_DATA or D_DATAVERSE_DATA in path.parents):
        domain = "dataverse_data"
    elif D_VALIDATION_ROOT.exists() and (path == D_VALIDATION_ROOT or D_VALIDATION_ROOT in path.parents):
        domain = "validation"
    elif D_CODEX_ROOT in path.parents:
        domain = "codex_agent_matrices"
    else:
        domain = "d_root"

    critical_names = {
        "MATRIX_INDEX.csv",
        "AGENT_DEFAULT_SKILL_ASSIGNMENT_MATRIX.csv",
        "CAPABILITY_USE_HARDENING_MATRIX.csv",
        "OPERATIONAL_CHAIN_GOVERNANCE_MATRIX.csv",
        "LIVE_SURFACE_GATE_MATRIX_20260603.csv",
        "EVIDENCE_READBACK_REGISTRY_20260603.csv",
        "PARALLEL_ISSUE_LANE_QUEUE.csv",
        "CONNECTION_INSTANCE_INVENTORY.csv",
        "CONNECTION_SECRET_BOUNDARY_MATRIX.csv",
        "DATAVERSE_DEV_ENVIRONMENT_BINDING_MATRIX.csv",
        "DATAVERSE_QUEUE_BINDING_UPDATE_RESULT.csv",
        "WORK_QUEUE_ENVIRONMENT_BINDING_MATRIX.csv",
        "WORK_QUEUE_FLOW_MATRIX.csv",
        "DEV_RUNTIME_STATE_MATRIX.csv",
        "seed_matrices.csv",
        "seed_environments.csv",
        "DATAVERSE_VALIDATION_REPORT.md",
    }
    priority = "P0" if path.name in critical_names else "P1" if artifact_type in {"matrix", "registry", "ledger", "validation_report"} else "P2"

    if artifact_type == "source_ref":
        recommended_sheet = "D Source Refs"
    elif artifact_type in {"registry", "ledger", "index"}:
        recommended_sheet = "D Registries Ledgers"
    elif artifact_type in {"matrix", "seed", "csv_source"}:
        recommended_sheet = "D Domain Matrices"
    elif artifact_type == "validation_report":
        recommended_sheet = "D Coverage Audit"
    else:
        recommended_sheet = "D Coverage Audit"
    return artifact_type, domain, recommended_sheet, priority


def discover_d_source_ref_rows() -> list[list]:
    rows: list[list] = []
    for path in d_candidate_artifacts():
        artifact_type, domain, recommended_sheet, priority = classify_d_artifact(path)
        if artifact_type != "source_ref":
            continue
        shape_rows, columns = csv_shape(path)
        rows.append(
            [
                path.stem,
                str(path),
                domain,
                artifact_type,
                recommended_sheet,
                priority,
                shape_rows,
                columns,
                path.stat().st_size,
                file_modified(path),
                first_markdown_heading(path) if path.suffix.lower() == ".md" else first_meaningful_line(path),
                "MODELED_CURRENT_RUN",
            ]
        )
    return rows


def discover_d_registries_ledgers_rows() -> list[list]:
    rows: list[list] = []
    for path in d_candidate_artifacts():
        artifact_type, domain, recommended_sheet, priority = classify_d_artifact(path)
        if artifact_type not in {"registry", "ledger", "index"}:
            continue
        shape_rows, columns = csv_shape(path)
        owner_agent = ""
        reviewer_agent = ""
        stop_condition = ""
        sample_rows = read_csv_dicts(path, limit=1)
        if sample_rows:
            sample = sample_rows[0]
            owner_agent = sample.get("owner_agent") or sample.get("primary_reader") or sample.get("agent_id") or ""
            reviewer_agent = sample.get("reviewer_agent") or ""
            stop_condition = sample.get("stop_condition") or ""
        rows.append(
            [
                path.stem,
                artifact_type,
                str(path),
                domain,
                owner_agent,
                reviewer_agent,
                "",
                "OBSERVED_READ_ONLY",
                stop_condition,
                shape_rows,
                columns,
                priority,
                recommended_sheet,
            ]
        )
    return rows


def discover_d_domain_matrix_rows() -> list[list]:
    rows: list[list] = []
    for path in d_candidate_artifacts():
        artifact_type, domain, recommended_sheet, priority = classify_d_artifact(path)
        if artifact_type not in {"matrix", "seed", "csv_source"}:
            continue
        shape_rows, columns = csv_shape(path)
        sample_rows = read_csv_dicts(path, limit=1)
        sample = sample_rows[0] if sample_rows else {}
        surface = sample.get("surface") or sample.get("target_surface") or sample.get("asset_id") or sample.get("path") or ""
        owner = sample.get("owner_agent") or sample.get("primary_reader") or sample.get("agent_id") or ""
        gate = sample.get("gate") or sample.get("gate_required") or sample.get("required_gate") or ""
        postcheck = sample.get("postcheck") or sample.get("validator") or ""
        boundary = sample.get("live_boundary") or sample.get("allowed_surface") or sample.get("blocked_surface") or ""
        rows.append(
            [
                domain,
                path.name,
                artifact_type,
                surface,
                boundary,
                owner,
                gate,
                postcheck,
                shape_rows,
                columns,
                priority,
                str(path),
            ]
        )
    return rows


def discover_d_coverage_audit_rows() -> list[list]:
    rows: list[list] = []
    directly_modeled = {
        os.path.normcase(os.path.abspath(str(path)))
        for path in [
            D_MATRIX_INDEX,
            D_RECIPE_INDEX,
            D_SKILL_USAGE_MATRIX,
            D_TOOL_INDEX,
            D_AGENTS_INDEX,
            D_LEVELS_CSV,
            D_AGENT_WORKPAPERS_MATRIX,
            D_GOVERNED_ASSET_INVENTORY,
            D_CODEX_ROOT / "workpapers" / "WORKPAPER_INDEX.csv",
            D_CODEX_ROOT / "matrices" / "ORDER_PREPARATION_ASSIGNMENT_MATRIX.csv",
            D_CODEX_ROOT / "matrices" / "EVIDENCE_READBACK_REGISTRY_20260603.csv",
            D_CODEX_ROOT / "matrices" / "AGENT_TOOL_RECIPE_SKILL_MATRIX.csv",
            D_CODEX_ROOT / "skills" / "SUBSKILL_USAGE_MATRIX.csv",
            D_CODEX_ROOT / "skills" / "SKILL_METADATA_QUALITY_MATRIX.csv",
            D_CODEX_ROOT / "recipes" / "SUBRECIPE_INDEX.csv",
            D_CODEX_ROOT / "recipes" / "SOURCE_TCU_RECIPES_INDEX.csv",
            D_CODEX_ROOT / "tools" / "SOURCE_TCU_TOOLS_INDEX.csv",
            D_CODEX_ROOT / "tools" / "SOURCE_TCU_TOOL_PERMISSIONS_INDEX.csv",
        ]
    }
    for path in d_candidate_artifacts():
        artifact_type, domain, recommended_sheet, priority = classify_d_artifact(path)
        key = os.path.normcase(os.path.abspath(str(path)))
        matched_by_path = "SI" if key in directly_modeled else "NO"
        matched_by_file = "SI" if path.name in {D_MATRIX_INDEX.name, D_RECIPE_INDEX.name, D_SKILL_USAGE_MATRIX.name, D_TOOL_INDEX.name, D_AGENTS_INDEX.name} else "NO"
        matched_by_id = "SI" if artifact_type in {"matrix", "index", "registry", "ledger"} else "NO" if priority == "P0" else ""
        if matched_by_path == "SI":
            missing_reason = "MODELED_DIRECTLY"
            next_delta = "Mantener sincronizado."
        elif recommended_sheet in {"D Source Refs", "D Registries Ledgers", "D Domain Matrices"}:
            missing_reason = "MODELED_IN_COVERAGE_SHEET"
            next_delta = "Si pasa a operativo, crear hoja especializada o vincular a matriz maestra."
        else:
            missing_reason = "PENDING_FINE_MODEL"
            next_delta = "Revisar si debe entrar en validación o cierre."
        rows.append(
            [
                str(path),
                artifact_type,
                domain,
                matched_by_path,
                matched_by_file,
                matched_by_id,
                missing_reason,
                next_delta,
                recommended_sheet,
                priority,
                path.stat().st_size,
                file_modified(path),
            ]
        )
    rows.sort(key=lambda row: (row[9], row[6], row[0].lower()))
    return rows


def discover_d_workbook_gap_rows() -> list[list]:
    rows = [
        ["W_D_INDEX_LAYER", "D Indexes", "index_registry", str(D_CODEX_ROOT), "ADDED_CURRENT_RUN", "All D indexes visible with row/column counts", "Mantener como landing sheet para nuevas matrices."],
        ["W_D_MATRIX_LAYER", "D Matrix Index", "matrix_registry", str(D_MATRIX_INDEX), "ADDED_CURRENT_RUN", "159-row matrix index modeled", "Usar para decidir próxima absorción fina."],
        ["W_D_COVERAGE_LAYER", "D Source Refs", "source_refs", str(D_ROOT), "ADDED_CURRENT_RUN", "SOURCE_* and markdown/json refs modeled as coverage", "Escalar solo los P0/P1 que pasen a operación."],
        ["W_D_COVERAGE_LAYER", "D Registries Ledgers", "registry_ledger", str(D_ROOT), "ADDED_CURRENT_RUN", "Indexes, registries and ledgers modeled across D", "Usar owner/reviewer/stop condition when present."],
        ["W_D_COVERAGE_LAYER", "D Domain Matrices", "domain_matrices", str(D_MATRICES_ROOT), "ADDED_CURRENT_RUN", "Domain matrices from D:/matrices and governance registry visible", "Promover a matriz maestra si afecta decisiones."],
        ["W_D_COVERAGE_LAYER", "D Coverage Audit", "coverage_audit", str(D_ROOT), "ADDED_CURRENT_RUN", "Path/file/id coverage audit modeled", "Resolver P0/P1 pending fine model first."],
        ["W_D_AGENT_WORKPAPER_READBACK_LAYER", "D Agent Registry", "agent_roster", str(D_AGENTS_INDEX), "ADDED_CURRENT_RUN", "14 canonical D agents modeled", "Seleccionar owner/reviewer desde esta hoja."],
        ["W_D_AGENT_WORKPAPER_READBACK_LAYER", "D Routing", "routing", str(D_ROUTING_JSON), "ADDED_CURRENT_RUN", "25 routes + handoff rules modeled", "Usar señales de routing para despachos."],
        ["W_D_AGENT_WORKPAPER_READBACK_LAYER", "D Workpapers", "workpapers", str(D_CODEX_ROOT / "workpapers"), "ADDED_CURRENT_RUN", "Canonical + extra workpaper folders visible", "Revisar carpetas no canónicas antes de limpiar."],
        ["W_D_AGENT_WORKPAPER_READBACK_LAYER", "D Orders", "orders", str(D_CODEX_ROOT / "orders"), "ADDED_CURRENT_RUN", "Order classes + files modeled", "Preparar órdenes desde matriz, no desde memoria conversacional."],
        ["W_D_AGENT_WORKPAPER_READBACK_LAYER", "D Readbacks", "readbacks", str(D_CODEX_ROOT / "readbacks"), "ADDED_CURRENT_RUN", "Registry + file-only readbacks visible", "Alinear readbacks no registrados si son vigentes."],
        ["W_D_CAPABILITY_LAYER", "D Skills Registry", "skill_usage", str(D_SKILL_USAGE_MATRIX), "ADDED_CURRENT_RUN", "Skill ids, levels, agents and live boundary modeled", "Resolver duplicados físicos cuando se active cleanup."],
        ["W_D_CAPABILITY_LAYER", "D Recipes Registry", "recipe_registry", str(D_RECIPE_INDEX), "ADDED_CURRENT_RUN", "Recipe ids modeled independently from physical markdown", "Usar recipe_id como clave, no nombre de archivo."],
        ["W_D_CAPABILITY_LAYER", "D Tools Registry", "tool_registry", str(D_TOOL_INDEX), "ADDED_CURRENT_RUN", "Tool ids and allowed/blocked surfaces modeled", "No confundir tool inventariado con validador ejecutado."],
        ["W_D_CAPABILITY_LAYER", "D Validators", "validator_execution", str(D_CODEX_ROOT / "tools"), "PENDING_EXECUTION", "Validators listed but not run in this workbook pass", "Ejecutar solo con gate read-only/postcheck explícito."],
        ["W_D_AUTHORITY_LAYER", "D Authority Canon", "authority_canon", str(D_AUTHORITY_CANON), "ADDED_CURRENT_RUN", "Authority files and matrices modeled", "Usar como fuente de jerarquía y cierre."],
        ["W_D_AUTHORITY_LAYER", "D Governed Assets", "asset_inventory", str(D_GOVERNED_ASSET_INVENTORY), "ADDED_CURRENT_RUN", "Governed assets modeled", "Usar coverage_status para detectar ruido o brechas."],
        ["W_D_LOGS_LAYER", "D Surface Summary", "logs", str(D_CODEX_ROOT / "logs"), "MISSING_EXPECTED_OR_NOT_APPLICABLE", "Logs folder absent", "No tratar como evidencia disponible."],
    ]
    validation_coverage = D_CODEX_ROOT / "matrices" / "VALIDATION_COVERAGE_MATRIX.csv"
    for row in read_csv_dicts(validation_coverage):
        rows.append(
            [
                f"W_D_VALIDATION_COVERAGE::{row.get('artifact_class')}",
                "D Workbook Gaps",
                "validation_coverage",
                row.get("required_index"),
                row.get("coverage_status"),
                f"validator={row.get('required_validator')} owner={row.get('owner_agent')}",
                row.get("stop_condition"),
            ]
        )
    source_gap = D_CODEX_ROOT / "matrices" / "SOURCE_CAPABILITY_ADOPTION_GAP_MATRIX.csv"
    for row in read_csv_dicts(source_gap):
        rows.append(
            [
                f"W_D_SOURCE_CAPABILITY_GAP::{row.get('source_id')}",
                "D Workbook Gaps",
                row.get("capability_class"),
                row.get("source_path"),
                row.get("current_state"),
                row.get("gap"),
                f"{row.get('next_action')} | validator={row.get('validator')} | stop={row.get('stop_condition')}",
            ]
        )
    return rows


def discover_atomic_delta_rows(
    repo_rows: list[list],
    pr_rows: list[list],
    branch_rows: list[list],
    env_agent_rows: list[list],
    workqueue_rows: list[list],
    cloud_rows: list[list],
    dataverse_source_rows: list[list],
    d_gap_rows: list[list] | None = None,
) -> list[list]:
    rows: list[list] = []
    sequence = 1

    open_prs = [row for row in pr_rows if str(row[3]) == "OPEN"]
    for row in open_prs:
        rows.append(
            [
                sequence,
                "GITHUB_PR",
                row[1],
                f"PR #{row[2]} {row[7]}",
                "REVISAR_DRAFT" if row[4] is True else "REVISAR_ABIERTO",
                "Medio",
                "Confirmar mergeabilidad, reviewer y siguiente acción.",
                row[14],
                "No merge automatico; no force push.",
                "PR con estado actualizado en GitHub.",
            ]
        )
        sequence += 1

    dirty_repos = [row for row in repo_rows if isinstance(row[9], int) and row[9] > 0]
    for row in dirty_repos:
        rows.append(
            [
                sequence,
                "REPO_DIRTY",
                row[1],
                row[2],
                "DELTA_TRIAGE_CAMBIOS",
                "Medio",
                f"Revisar {row[9]} cambios locales antes de publicar o limpiar.",
                row[2],
                "No reset; no checkout destructivo.",
                "Repo con cambios locales.",
            ]
        )
        sequence += 1

    for row in workqueue_rows:
        backlog = row[11] if len(row) > 11 else 0
        if isinstance(backlog, int) and backlog > 0:
            rows.append(
                [
                    sequence,
                    "DATAVERSE_QUEUE",
                    row[3],
                    f"backlog={backlog}",
                    "DELTA_PRIORIZAR_COLA",
                    "Medio",
                    "Elegir una cola y consumir/depurar con gate explícito.",
                    row[14],
                    "No ejecutar flows ni procesar items sin owner/rollback.",
                    "Cola viva con backlog.",
                ]
            )
            sequence += 1

    for row in cloud_rows:
        if row[11] == "VERIFICAR_CLOUD_LIVE":
            rows.append(
                [
                    sequence,
                    "CODEX_CLOUD",
                    row[2] or row[1],
                    row[1],
                    "DELTA_VERIFICAR_ENV_CLOUD",
                    "Bajo/Medio",
                    "Separar environment confirmado de mención histórica.",
                    row[12],
                    "No tocar secrets ni setup hasta confirmar target.",
                    "Mención detectada en prompt-history sin objeto local.",
                ]
            )
            sequence += 1

    live_confirmed = any(row[5] == "live_rows_confirmed" for row in dataverse_source_rows)
    rows.append(
        [
            sequence,
            "DATAVERSE_SOURCE",
            "mon_sdu_source_artifacts / mon_sdu_evidences",
            "5 pares source/evidence",
            "OK_BASE_REHIDRATADA",
            "Bajo",
            "Usar estas tablas como fuente larga del workbook.",
            str(ROOT / "operativa" / "DATAVERSE_REHIDRATACION_LIVE_READ_20260617.json"),
            "No convertir metadata en permiso de escritura general.",
            "live_rows_confirmed" if live_confirmed else "requiere refresh",
        ]
    )
    sequence += 1

    if env_agent_rows:
        rows.append(
            [
                sequence,
                "AGENTES_ENTORNO",
                "HUBDesarrollo",
                f"{len(env_agent_rows)} agentes/bots mapeados",
                "DELTA_SELECCIONAR_AGENT_OWNER",
                "Bajo/Medio",
                "Tomar owner/reviewer desde Agentes Entorno antes de despachar waves.",
                str(DATAVERSE_LIVE_INVENTORY),
                "No delegar a agente sin superficie, gate y evidencia.",
                "Tabla operativa lista.",
            ]
        )
        sequence += 1

    if d_gap_rows:
        pending = [row for row in d_gap_rows if str(row[4]).startswith("PENDING") or str(row[4]).startswith("MISSING")]
        rows.append(
            [
                sequence,
                "D_CABINA_AGENT_LAYER",
                "W_D_AGENT_WORKPAPER_READBACK_LAYER",
                f"{len(d_gap_rows)} gaps/capas D modeladas, {len(pending)} pendientes",
                "DELTA_VALIDAR_D_LAYER_READONLY",
                "Bajo",
                "Ejecutar validadores D read-only solo si el target queda explícito.",
                str(D_CODEX_ROOT),
                "No escribir en D ni tratar logs ausentes como evidencia.",
                "Workbook actualizado; quedan validadores pendientes de ejecución.",
            ]
        )
    return rows


def add_table(ws, name: str, start_row: int, headers: list[str], rows: list[list]):
    end_row = max(start_row, start_row + len(rows))
    end_col = len(headers)
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def create_sheet(
    wb: Workbook,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[list],
    table_name: str,
    widths: list[int],
):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title).font = Font(bold=True, size=16, color=WHITE)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1, subtitle).font = Font(italic=True, color="44546A")
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=SUB_FILL)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for r_idx, row in enumerate(rows, 5):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_FILL)
            if isinstance(value, str):
                upper = value.upper()
                if upper in {"PASS", "OK", "SI", "NO_TOCAR", "MONTADA_READONLY", "MONTADA_GOBERNADA", "ACTUAL", "INDEXED", "CHRONOLOGY_INDEXED", "ARCHIVE_REVERSIBLE", "LEGACY_HEAVY_INDEXED", "REFERENCIA"}:
                    cell.fill = PatternFill("solid", fgColor=OK_FILL)
                elif upper in {"REVISAR", "OBSERVAR", "PENDIENTE", "REVISAR_OWNER", "REVISAR_SENSIBLE", "ENTRYPOINT_INCOMPLETE", "NO_MONTADA"}:
                    cell.fill = PatternFill("solid", fgColor=WARN_FILL)
                elif upper in {"FAIL", "MISSING"}:
                    cell.fill = PatternFill("solid", fgColor=BAD_FILL)
                if (value.startswith("C:\\") or value.startswith("D:\\")) and len(value) < 240:
                    cell.hyperlink = value
                    cell.style = "Hyperlink"

    add_table(ws, table_name, 4, headers, rows)
    ws.freeze_panes = "A5"
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 34
    return ws


def load_old_decisions() -> tuple[dict[str, str], dict[str, str]]:
    decisions: dict[str, str] = {}
    prompt_decisions: dict[str, str] = {}
    if not WORKBOOK.exists():
        return decisions, prompt_decisions
    try:
        old = load_workbook(WORKBOOK, data_only=False)
        if "Decision" in old.sheetnames:
            ws = old["Decision"]
            for row in range(5, ws.max_row + 1):
                key = ws.cell(row, 1).value
                value = ws.cell(row, 8).value
                if key and value:
                    decisions[str(key)] = str(value)
        for sheet_name in ("Prompt History", "Candidatos"):
            if sheet_name not in old.sheetnames:
                continue
            ws = old[sheet_name]
            for row in range(5, ws.max_row + 1):
                key = ws.cell(row, 1).value
                value = ws.cell(row, 13).value
                if key and value:
                    prompt_decisions[str(key)] = str(value)
    except Exception:
        pass
    return decisions, prompt_decisions


def main():
    now = datetime.now()
    stamp = now.strftime("%Y%m%d_%H%M%S")
    now_label = now.strftime("%Y-%m-%d %H:%M:%S")
    old_decisions, old_prompt_decisions = load_old_decisions()

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = None
    if WORKBOOK.exists():
        backup_path = BACKUP_DIR / f"{WORKBOOK.stem}_before_update_{stamp}.xlsx"
        shutil.copy2(WORKBOOK, backup_path)

    state_bytes = STATE_PATH.read_bytes()
    state = json.loads(state_bytes.decode("utf-8"))
    state_hash = hashlib.sha256(state_bytes).hexdigest().upper()
    backup_hash = sha256(STATE_BAK)
    backup_identical = "SI" if backup_hash == state_hash else "NO"
    persisted = state.get("electron-persisted-atom-state", {})
    env = persisted.get("environment", {})
    prompt_history = persisted.get("prompt-history", {})
    thread_permissions = persisted.get("heartbeat-thread-permissions-by-id", {})
    thread_hints = state.get("thread-workspace-root-hints", {})
    pinned = set(state.get("pinned-thread-ids", []) or [])

    config = tomllib.loads(CONFIG_PATH.read_text(encoding="utf-8")) if CONFIG_PATH.exists() else {}
    plugins = sorted(
        name for name, item in (config.get("plugins") or {}).items() if item.get("enabled") is True
    )
    projects = sorted((config.get("projects") or {}).keys())
    features = config.get("features") or {}
    desktop = config.get("desktop") or {}
    mcp_servers = sorted((config.get("mcp_servers") or {}).keys())

    drive_json = ps_json(
        "[System.IO.DriveInfo]::GetDrives() | "
        "Select-Object Name,DriveType,IsReady,VolumeLabel,DriveFormat,TotalSize,AvailableFreeSpace | "
        "ConvertTo-Json -Depth 3"
    )
    disk_json = ps_json(
        "Get-CimInstance -Namespace root\\Microsoft\\Windows\\Storage -ClassName MSFT_Disk | "
        "Select-Object Number,FriendlyName,BusType,Size,PartitionStyle,OperationalStatus,HealthStatus,IsOffline,IsReadOnly | "
        "ConvertTo-Json -Depth 4"
    )
    d_drive = next((item for item in as_list(drive_json) if item.get("Name") == "D:\\"), {})
    d_disk = next((item for item in as_list(disk_json) if item.get("FriendlyName") == "Msft Virtual Disk"), {})
    d_exists = Path(r"D:\\").exists()
    d_readonly = bool(d_disk.get("IsReadOnly"))
    if d_exists and d_readonly:
        d_status = "MONTADA_READONLY"
        d_mode_note = "Cabina D montada en modo solo lectura."
    elif d_exists:
        d_status = "MONTADA_GOBERNADA"
        d_mode_note = "Cabina D montada; escritura solo con gate humano exacto."
    else:
        d_status = "NO_MONTADA"
        d_mode_note = "Cabina D no disponible."

    text_patterns = [
        "OPENAI_API_KEY",
        "api_key",
        "client_secret",
        "BEGIN PRIVATE KEY",
        "Bearer ",
        "password",
        "access_token",
        "refresh_token",
        "openai_sk_like",
        "github_token_like",
        "jwt_like",
        "aws_access_key_like",
    ]

    def classify_prompt(group: str, text: str, sensitive_count: int):
        if group in {"global", "new-conversation"}:
            return "CORE_PROMPT_HISTORY", "NO_TOCAR", "Medio", "Base UI/prompt; no remover por limpieza."
        if group in pinned:
            return "PINNED_THREAD", "NO_TOCAR", "Medio", "Hilo fijado; requiere decisión humana específica."
        if group.startswith("task_"):
            return "CLEANUP_TASK_EPHEMERAL", "REMOVER_DRY_RUN", "Bajo", "Grupo task efímero; candidato seguro si postcheck pasa."
        if group in thread_permissions:
            return "PERMISSION_LINKED", "REVISAR_OWNER", "Medio/Alto", "Tiene permiso de hilo visible; revisar owner/recencia."
        if group in thread_hints:
            return "WORKSPACE_HINTED", "REVISAR_OWNER", "Medio", "Tiene pista de workspace; revisar antes de limpiar."
        if sensitive_count:
            return "CLEANUP_NO_PERMISSION", "REVISAR_SENSIBLE", "Medio", "Sin permiso visible pero con menciones sensibles."
        return "CLEANUP_NO_PERMISSION", "REMOVER_DRY_RUN", "Bajo/Medio", "Sin permiso de hilo visible; remover con backup."

    prompt_rows = []
    for group, entries in prompt_history.items():
        entries = entries if isinstance(entries, list) else [entries]
        text = "\n".join(str(item) for item in entries)
        counts = count_mentions(text)
        sensitive_count = sum(counts.values())
        cls, action, risk, note = classify_prompt(str(group), text, sensitive_count)
        prompt_rows.append(
            [
                str(group),
                len(entries),
                len(text),
                kb(len(text)),
                max([len(str(item)) for item in entries], default=0),
                sensitive_count,
                "SI" if str(group) in thread_permissions else "NO",
                "SI" if str(group) in thread_hints else "NO",
                None,
                cls,
                action,
                risk,
                old_prompt_decisions.get(str(group)),
                note,
            ]
        )
    prompt_rows.sort(key=lambda item: (0 if item[9].startswith("CLEANUP") else 1, -(item[3] or 0), item[0]))
    candidate_rows = [row[:] for row in prompt_rows if row[9].startswith("CLEANUP")]
    prompt_read_rows = []
    for group, entries in prompt_history.items():
        entries = entries if isinstance(entries, list) else [entries]
        group_text = "\n".join(str(item) for item in entries)
        counts = count_mentions(group_text)
        cls, action, risk, _note = classify_prompt(str(group), group_text, sum(counts.values()))
        for entry_index, entry in enumerate(entries, 1):
            original = str(entry)
            original_hash = hashlib.sha256(original.encode("utf-8")).hexdigest().upper()
            sanitized, redactions = sanitize_prompt_text(original)
            for chunk_index, chunk in enumerate(chunk_text(sanitized), 1):
                prompt_read_rows.append(
                    [
                        str(group),
                        entry_index,
                        chunk_index,
                        cls,
                        action,
                        risk,
                        len(original),
                        original_hash,
                        redactions,
                        chunk,
                    ]
                )
    ph_groups = len(prompt_rows)
    ph_entries = sum(row[1] for row in prompt_rows)
    ph_kb = round(sum(row[3] or 0 for row in prompt_rows), 2)
    safe_kb = round(sum(row[3] or 0 for row in candidate_rows if row[10] == "REMOVER_DRY_RUN"), 2)
    review_kb = round(sum(row[3] or 0 for row in prompt_rows if row[10].startswith("REVISAR")), 2)
    ph_text = json.dumps(prompt_history, ensure_ascii=False, sort_keys=True)
    repo_rows = discover_repositories(projects)
    branch_rows = discover_branch_rows(repo_rows)
    pr_rows = discover_pr_rows(repo_rows)
    cloud_rows = discover_cloud_environment_rows(env, ph_text)
    agent_rows = discover_agent_rows()
    universe_rows = discover_universe_rows()
    skill_rows = discover_skill_rows()
    recipe_rows = discover_recipe_rows()
    tool_rows = discover_tool_rows(config)
    connection_rows = discover_connection_rows(config, env)
    dataverse_source_rows = discover_dataverse_source_rows()
    environment_solution_rows = discover_environment_solution_rows()
    environment_agent_rows = discover_environment_agent_rows()
    workqueue_rows = discover_workqueue_rows()
    atomic_action_rows = discover_atomic_action_rows()
    master_atomic_rows = discover_master_atomic_matrix_rows()
    metadata_hydration_rows = discover_metadata_hydration_rows()
    d_surface_rows = discover_d_surface_rows(now_label)
    local_surface_rows = discover_local_surface_rows(now_label, d_status)
    workspace_rows = discover_workspace_rows(now_label, d_status)
    d_indexes_rows = discover_d_indexes_rows()
    d_matrix_index_rows = discover_d_matrix_index_rows()
    d_skill_registry_rows = discover_d_skill_registry_rows()
    d_subskill_rows = discover_d_subskill_rows()
    d_recipe_registry_rows = discover_d_recipe_registry_rows()
    d_tools_registry_rows = discover_d_tools_registry_rows()
    d_validator_rows = discover_d_validator_rows()
    d_agent_registry_rows = discover_d_agent_registry_rows()
    d_levels_rows = discover_d_levels_rows()
    d_routing_rows = discover_d_routing_rows()
    d_workpaper_rows = discover_d_workpaper_rows()
    d_order_rows = discover_d_order_rows()
    d_readback_rows = discover_d_readback_rows()
    d_agent_tool_skill_rows = discover_d_agent_tool_skill_rows()
    d_authority_rows = discover_d_authority_rows()
    d_governed_asset_rows = discover_d_governed_asset_rows()
    d_source_ref_rows = discover_d_source_ref_rows()
    d_registries_ledgers_rows = discover_d_registries_ledgers_rows()
    d_domain_matrix_rows = discover_d_domain_matrix_rows()
    d_coverage_audit_rows = discover_d_coverage_audit_rows()
    d_gap_rows = discover_d_workbook_gap_rows()
    atomic_delta_rows = discover_atomic_delta_rows(
        repo_rows,
        pr_rows,
        branch_rows,
        environment_agent_rows,
        workqueue_rows,
        cloud_rows,
        dataverse_source_rows,
        d_gap_rows,
    )

    env_label = env.get("label", "N/D")
    env_id = env.get("id", "N/D")

    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = [
        ["Archivo auditado", str(STATE_PATH), "Estado vivo de Codex/Electron"],
        ["Fecha workbook", now_label, "Actualizado localmente"],
        ["Tamaño global-state", STATE_PATH.stat().st_size, "Bytes actuales al actualizar workbook"],
        ["JSON valido", "PASS", "Parseo correcto"],
        ["Backup identical", backup_identical, "Backup del global-state contra archivo vivo"],
        ["Environment activo", env_label, env_id],
        ["Codex model", config.get("model", "N/D"), f"reasoning={config.get('model_reasoning_effort', 'N/D')}"],
        ["Sandbox / approval", f"{config.get('sandbox_mode', 'N/D')} / {config.get('approval_policy', 'N/D')}", "Config local"],
        ["Plugins habilitados", len(plugins), ", ".join(plugins[:8]) + ("..." if len(plugins) > 8 else "")],
        ["Trusted projects", len(projects), "Entradas [projects] en config.toml"],
        ["Workspace actual", str(ROOT), "Rama y estado en hoja Workspace Actual"],
        ["Superficies locales", len(local_surface_rows), "D, PROJEC CDX, .codex, CodexLocal, cronologia, archivo y repos"],
        ["D root", d_status, d_mode_note],
        ["D total GB", gb(d_drive.get("TotalSize")), "Capacidad visible actual"],
        ["D libre GB", gb(d_drive.get("AvailableFreeSpace")), "Espacio libre actual"],
        ["VHDX origen", str(VHDX_PATH), f"{gb(VHDX_PATH.stat().st_size) if VHDX_PATH.exists() else 'N/D'} GB físicos"],
        ["Prompt-history grupos", ph_groups, "Con lectura saneada en hoja Prompts Lectura"],
        ["Prompt-history entradas", ph_entries, "Conteo saneado"],
        ["Prompts lectura filas", len(prompt_read_rows), "Una fila por entrada/chunk con tokens reales redactados"],
        ["Repositorios inventariados", len(repo_rows), "GitHub local, D activo, trusted projects y workspace actual"],
        ["Ramas inventariadas", len(branch_rows), "Refs locales/remotas y ramas GitHub live cuando gh responde"],
        ["PRs inventariados", len(pr_rows), "GitHub live via gh; all states, limite 50 por repo"],
        ["Codex Cloud entornos", len(cloud_rows), "Confirmados localmente + menciones en prompts"],
        ["Agentes inventariados", len(agent_rows), "Fan-in, roster SDU, revision live y componentes Dataverse"],
        ["Agentes por entorno", len(environment_agent_rows), "SDU agent mappings + bots/copilots Dataverse"],
        ["Entornos/Soluciones", len(environment_solution_rows), "HUBDesarrollo, soluciones y conteos de tablas"],
        ["Colas Dataverse", len(workqueue_rows), "Workqueues vivas con backlog"],
        ["Universos inventariados", len(universe_rows), "Superficies canonicas + auditoria de relaciones"],
        ["Skills inventariadas", len(skill_rows), "Tabla unificada + escaneo SKILL.md"],
        ["Recetas inventariadas", len(recipe_rows), "Recipes locales y .agents"],
        ["Tools inventariadas", len(tool_rows), "Scripts, MCP servers y plugins habilitados"],
        ["Conexiones inventariadas", len(connection_rows), "Superficies, gates, mappings, MCP y plugins"],
        ["Dataverse fuentes", len(dataverse_source_rows), "Tablas SDU, source/evidence y lecturas vivas previas"],
        ["Matriz maestra atómica", len(master_atomic_rows), "Waves owner/reviewer/gates/surface/next_delta"],
        ["Acciones atómicas", len(atomic_action_rows), "Diccionario de acciones atómicas y mutabilidad"],
        ["Matriz hidratación", len(metadata_hydration_rows), "Rows metadata-only para Dataverse SDU"],
        ["D superficies", len(d_surface_rows), "Roots y capas D con conteos físicos"],
        ["D indexes/matrices", len(d_indexes_rows), "Índices CSV y matrices de D con shape"],
        ["D matriz index", len(d_matrix_index_rows), "MATRIX_INDEX D como mapa maestro operativo"],
        ["D skills registry", len(d_skill_registry_rows), "SKILL_USAGE_MATRIX + rutas físicas D"],
        ["D recipes registry", len(d_recipe_registry_rows), "Recipe/subrecipe/source indexes D"],
        ["D tools registry", len(d_tools_registry_rows), "TOOL_INDEX + permisos/fuentes D"],
        ["D agentes canonicos", len(d_agent_registry_rows), "AGENTS_INDEX + agents.json D"],
        ["D routing", len(d_routing_rows), "Rutas y handoff rules D"],
        ["D workpapers", len(d_workpaper_rows), "Workpaper index + carpetas no canónicas"],
        ["D readbacks", len(d_readback_rows), "Registry + readbacks físicos D"],
        ["D authority canon", len(d_authority_rows), "Archivos rectores en 02_AUTHORITY_CANON"],
        ["D governed assets", len(d_governed_asset_rows), "GOVERNED_ASSET_CANONICAL_INVENTORY"],
        ["D source refs", len(d_source_ref_rows), "SOURCE/ref markdown-json-yaml detectados en D"],
        ["D registries/ledgers", len(d_registries_ledgers_rows), "Indexes, registries y ledgers detectados"],
        ["D domain matrices", len(d_domain_matrix_rows), "Matrices/seeds por dominio D"],
        ["D coverage audit", len(d_coverage_audit_rows), "Auditoría path/file/id de cobertura D"],
        ["D workbook gaps", len(d_gap_rows), "Capa de gaps modelada y pendientes"],
        ["Deltas atómicos", len(atomic_delta_rows), "Siguiente capa operable por superficie"],
        ["Prompt-history KB aprox", ph_kb, "Tamaño textual aproximado"],
        ["Pool seguro dry-run KB", safe_kb, "Task/no-permission"],
        ["Pool bajo revisión KB", review_kb, "Permission/workspace/sensible"],
    ]
    ws = create_sheet(
        wb,
        "Resumen",
        "Configuración vigente, D recuperado, prompts leídos con saneamiento y opciones de decisión.",
        ["Metrica", "Valor", "Lectura"],
        summary_rows,
        "tblResumen",
        [28, 78, 70],
    )
    nav_start = len(summary_rows) + 7
    ws.cell(nav_start, 1, "Navegación").font = Font(bold=True, color=WHITE)
    ws.cell(nav_start, 1).fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws.cell(nav_start, 3, "Dictamen").font = Font(bold=True)
    ws.cell(nav_start + 1, 1, "Decision")
    ws.cell(nav_start + 1, 3, "Operable: SI. D está read-only; escritura requiere gate separado.")
    for index, sheet_name in enumerate(
        [
            "Decision",
            "Config Vigente",
            "Cabina D",
            "Workspace Actual",
            "Superficies Locales",
            "D Surface Summary",
            "D Indexes",
            "D Matrix Index",
            "D Skills Registry",
            "D Subskills",
            "D Recipes Registry",
            "D Tools Registry",
            "D Validators",
            "D Agent Registry",
            "D Levels",
            "D Routing",
            "D Workpapers",
            "D Orders",
            "D Readbacks",
            "D Agent Tool Skill",
            "D Authority Canon",
            "D Governed Assets",
            "D Source Refs",
            "D Registries Ledgers",
            "D Domain Matrices",
            "D Coverage Audit",
            "D Workbook Gaps",
            "Global State",
            "Prompt History",
            "Prompts Lectura",
            "Candidatos",
            "Repositorios",
            "Ramas",
            "PRs",
            "Codex Cloud Env",
            "Agentes",
            "Agentes Entorno",
            "Entornos Soluciones",
            "Colas",
            "Universos",
            "Skills",
            "Recetas",
            "Tools",
            "Conexiones",
            "Dataverse Fuentes",
            "Matriz Maestra",
            "Acciones Atomicas",
            "Matriz Hidratacion",
            "Deltas Atomicos",
            "Riesgos Gates",
            "Agentes Carriles",
            "Validacion",
            "Navegacion",
        ],
        nav_start + 2,
    ):
        ws.cell(index, 1, sheet_name)
        ws.cell(index, 1).hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(index, 1).style = "Hyperlink"

    decision_rows = [
        ["0_NO_TOCAR", "Conservar estado actual", 0, "Bajo", "Estabilidad total", "Ninguno", "No aplica", old_decisions.get("0_NO_TOCAR")],
        ["1_D_READONLY_ACTUAL", "Mantener D montado solo lectura", 0, "Bajo", "Auditoría/inventario seguro", "Actual", "Dismount VHDX", old_decisions.get("1_D_READONLY_ACTUAL", "ACTUAL")],
        ["2_D_READWRITE_GOVERNED", "Remontar D con escritura", 0, "Medio", "Solo si hay que operar en cabina D", "Orden humana exacta", "Dismount + attach readonly", old_decisions.get("2_D_READWRITE_GOVERNED")],
        ["3_SAFE_DRY_RUN", "Marcar/remover task_ephemeral + no-permission", safe_kb, "Bajo", "Primer movimiento atómico seguro", "Codex cerrado + backup", "Restaurar JSON backup", old_decisions.get("3_SAFE_DRY_RUN", "RECOMENDADO")],
        ["4_BALANCED_REVIEW", "Sumar buckets tras revisar recencia", review_kb, "Medio", "Reducción notable", "Lista exacta + owner/reviewer", "Restaurar JSON backup", old_decisions.get("4_BALANCED_REVIEW")],
        ["5_POWERFUL_PERMISSION_REVIEW", "Revisar permission/workspace", review_kb, "Medio/Alto", "Limpieza potente real", "Orden gobernada separada", "Restaurar JSON backup", old_decisions.get("5_POWERFUL_PERMISSION_REVIEW")],
    ]
    ws = create_sheet(
        wb,
        "Decision",
        "Opciones graduadas para operar D y alivianar Codex sin romper estado vivo.",
        ["Opcion", "Accion", "Reduccion estimada KB", "Riesgo", "Cuando elegir", "Gate", "Rollback", "Decision"],
        decision_rows,
        "tblDecision",
        [30, 48, 22, 18, 42, 32, 34, 18],
    )
    dv = DataValidation(type="list", formula1='"ACTUAL,RECOMENDADO,APROBADO,ESPERAR,RECHAZADO"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"H5:H{4 + len(decision_rows)}")

    config_rows = [
        ["environment.id", env.get("id"), "NO_TOCAR", "Environment activo de Codex Cloud."],
        ["environment.label", env.get("label"), "NO_TOCAR", "Label visible del entorno."],
        ["environment.workspace_dir", env.get("workspace_dir"), "NO_TOCAR", "Ruta cloud activa."],
        ["environment.codexCloudAccess", persisted.get("codexCloudAccess"), "NO_TOCAR", "Acceso cloud."],
        ["environment.task_count", env.get("task_count"), "OBSERVAR", "Tareas actuales del entorno."],
        ["environment.post_setup_cache_enabled", env.get("post_setup_cache_enabled"), "NO_TOCAR", "Cache post setup."],
        ["environment.secrets.count", len(env.get("secrets") or []), "OK", "Sin imprimir secretos."],
        ["codex.config.model", config.get("model"), "OK", "Modelo local configurado."],
        ["codex.config.reasoning", config.get("model_reasoning_effort"), "OK", "Esfuerzo de razonamiento."],
        ["codex.config.approval_policy", config.get("approval_policy"), "OBSERVAR", "Archivo config.toml."],
        ["codex.config.sandbox_mode", config.get("sandbox_mode"), "OBSERVAR", "Archivo config.toml."],
        ["codex.features.multi_agent", features.get("multi_agent"), "OK", "Multi-agente habilitado."],
        ["codex.features.memories", features.get("memories"), "OK", "Memorias habilitadas."],
        ["codex.desktop.locale", desktop.get("localeOverride"), "OK", "Interfaz local."],
        ["codex.desktop.shell", desktop.get("integratedTerminalShell"), "OK", "Shell integrada."],
        ["codex.plugins.enabled_count", len(plugins), "OK", "Plugins habilitados."],
        ["codex.projects.trusted_count", len(projects), "OK", "Proyectos trusted."],
        ["codex.mcp_servers.count", len(mcp_servers), "OK", "Servidores MCP configurados."],
        ["local.surface.D.status", d_status, "OK" if d_exists else "REVISAR", d_mode_note],
        ["local.surface.D.path", r"D:\\", "NO_TOCAR", "Root de cabina universal."],
        ["local.surface.D.label", d_drive.get("VolumeLabel"), "OK", "Etiqueta del volumen."],
        ["local.surface.D.format", d_drive.get("DriveFormat"), "OK", "Formato del volumen."],
        ["local.surface.D.total_gb", gb(d_drive.get("TotalSize")), "OK", "Tamaño virtual visible."],
        ["local.surface.D.free_gb", gb(d_drive.get("AvailableFreeSpace")), "OK", "Espacio libre visible."],
        ["local.surface.D.disk_readonly", d_readonly, "OBSERVAR", "Indicador del disco subyacente; no usar como permiso de escritura."],
        ["local.surface.D.vhdx", str(VHDX_PATH), "REFERENCIA", "Evidencia historica de recuperacion si existe."],
        ["local.surface.D.evidence_log", str(EVIDENCE_LOG), "REFERENCIA", "Bitácora historica de attach/recuperacion si existe."],
    ]
    for key, value in sorted((env.get("env_vars") or {}).items()):
        decision = "REVISAR" if key == "SOURCE_TREE_PATH" and str(value) == "CODEX_SOURCE_TREE_PATH" else "OK"
        note = "Valor literal sospechoso" if decision == "REVISAR" else "Variable runtime/version."
        config_rows.append([f"env_vars.{key}", safe(value), decision, note])
    for index, line in enumerate(as_list(env.get("setup")), 1):
        config_rows.append([f"setup[{index}]", line, "OK", "Script setup cloud vigente."])
    for index, line in enumerate(as_list(env.get("maintenance_setup")), 1):
        config_rows.append([f"maintenance_setup[{index}]", line, "OK", "Script mantenimiento cloud vigente."])

    ws = create_sheet(
        wb,
        "Config Vigente",
        "Environment, config.toml, variables, setup, mantenimiento y superficies locales.",
        ["Campo", "Valor", "Decision", "Nota"],
        config_rows,
        "tblConfig",
        [36, 92, 18, 58],
    )

    d_agents = Path(r"D:\AGENTS.md")
    d_manifest = Path(r"D:\MANIFEST.yaml")
    d_validator_agent = Path(r"D:\.agents\codex\tools\local_validate_agent_layer.ps1")
    d_validator_chain = Path(r"D:\.agents\codex\tools\local_validate_operational_chain.ps1")
    d_rows = [
        ["status", d_status, "PASS" if d_exists else "FAIL", d_mode_note],
        ["drive", r"D:\\", "PASS" if d_exists else "FAIL", "Letra visible actual."],
        ["label", d_drive.get("VolumeLabel"), "PASS", "Etiqueta del volumen."],
        ["format", d_drive.get("DriveFormat"), "PASS", "Formato del volumen."],
        ["total_gb", gb(d_drive.get("TotalSize")), "PASS", "Capacidad virtual visible."],
        ["free_gb", gb(d_drive.get("AvailableFreeSpace")), "PASS", "Libre actual."],
        ["disk_number", d_disk.get("Number"), "PASS", "Número de disco asociado."],
        ["disk_name", d_disk.get("FriendlyName"), "PASS", "Msft Virtual Disk."],
        ["is_readonly", d_readonly, "INFO" if d_exists else "OBSERVAR", "Estado reportado por el disco; no reemplaza el gate humano."],
        ["vhdx_path", str(VHDX_PATH), "INFO" if VHDX_PATH.exists() else "MISSING", "Referencia historica de recuperacion."],
        ["vhdx_size_gb", gb(VHDX_PATH.stat().st_size) if VHDX_PATH.exists() else None, "INFO" if VHDX_PATH.exists() else "MISSING", "Tamaño físico si la referencia existe."],
        ["AGENTS.md", str(d_agents), "PASS" if d_agents.exists() else "MISSING", "Fuente rectora local."],
        ["MANIFEST.yaml", str(d_manifest), "PASS" if d_manifest.exists() else "MISSING", "Manifest de cabina."],
        ["validator.agent_layer", str(d_validator_agent), "PASS" if d_validator_agent.exists() else "MISSING", "Validador local."],
        ["validator.operational_chain", str(d_validator_chain), "PASS" if d_validator_chain.exists() else "MISSING", "Validador cadena."],
        ["evidence_log", str(EVIDENCE_LOG), "INFO" if EVIDENCE_LOG.exists() else "MISSING", "Bitácora historica de attach/recuperacion."],
        ["next_gate", "D_READWRITE_GOVERNED", "PENDIENTE", "Solo si hace falta escribir; requiere orden humana exacta."],
    ]
    create_sheet(wb, "Cabina D", "Estado de la consola rectora D y frontera de escritura gobernada.", ["Campo", "Valor", "Estado", "Nota"], d_rows, "tblCabinaD", [30, 98, 18, 62])

    create_sheet(
        wb,
        "Workspace Actual",
        "Workspace de esta rama: ruta, Git, workbook vigente y cruces con superficies locales.",
        ["Campo", "Valor", "Estado", "Nota"],
        workspace_rows,
        "tblWorkspaceActual",
        [34, 110, 24, 70],
    )

    create_sheet(
        wb,
        "Superficies Locales",
        "Mapa operativo de consola, workbench, runtime, puente local, cronologia, archivo y repos canonicos.",
        [
            "Surface Id",
            "Path",
            "Clase",
            "Estado",
            "Existe",
            "README",
            "AGENTS",
            "MAPA",
            "Indice",
            "Git",
            "Dirs",
            "Files",
            "KB aprox",
            "Extensiones",
            "Last Scan",
            "Rol",
            "Regla",
        ],
        local_surface_rows,
        "tblSuperficiesLocales",
        [36, 92, 28, 30, 10, 10, 10, 10, 10, 10, 12, 12, 14, 50, 22, 62, 78],
    )

    create_sheet(
        wb,
        "D Surface Summary",
        "Roots principales de D con conteo fisico y estado de disponibilidad gobernada.",
        ["Surface Root", "Existe", "Dirs", "Files", "Extensions", "Last Scan", "Status", "Notes"],
        d_surface_rows,
        "tblDSurfaceSummary",
        [82, 10, 12, 12, 58, 22, 34, 70],
    )

    create_sheet(
        wb,
        "D Indexes",
        "Indices, matrices, registries y mapas CSV encontrados en D; muestra shape y si ya quedan consumidos por el workbook.",
        ["Root", "Index Name", "Path", "Kind", "Rows", "Columns", "Bytes", "Modified", "Consumed By Sheet", "Gap Status"],
        d_indexes_rows,
        "tblDIndexes",
        [70, 54, 96, 24, 12, 92, 14, 22, 20, 30],
    )

    create_sheet(
        wb,
        "D Matrix Index",
        "MATRIX_INDEX de D como mapa maestro de matrices operativas, readers y reglas de actualizacion.",
        ["Matrix Id", "Path", "Path Exists", "Rows", "Columns", "Scope", "Primary Reader", "Update Rule", "Source", "Source Path"],
        d_matrix_index_rows,
        "tblDMatrixIndex",
        [48, 96, 14, 10, 92, 38, 34, 90, 24, 96],
    )

    create_sheet(
        wb,
        "D Skills Registry",
        "SKILL_USAGE_MATRIX de D cruzada contra rutas fisicas SKILL.md en D.",
        [
            "Root Label",
            "Skill Id",
            "Canonical",
            "Source Path",
            "Has Skill Md",
            "Assigned Level",
            "Assigned Agents",
            "Use When",
            "Live Boundary",
            "Duplicate Group",
            "Workbook Status",
            "Source Index",
        ],
        d_skill_registry_rows,
        "tblDSkillsRegistry",
        [24, 42, 42, 110, 14, 38, 70, 90, 30, 28, 34, 96],
    )

    create_sheet(
        wb,
        "D Subskills",
        "Subskills y calidad de metadata en D para no perder capacidades finas.",
        [
            "Subskill Id",
            "Parent Skill",
            "Assigned Agents",
            "Use When",
            "Required Recipe",
            "Required Tool",
            "Validator",
            "Stop Condition",
            "Source Index",
            "Workbook Status",
        ],
        d_subskill_rows,
        "tblDSubskills",
        [44, 42, 70, 90, 58, 58, 70, 60, 96, 34],
    )

    create_sheet(
        wb,
        "D Recipes Registry",
        "Recipe, subrecipe y source recipe indexes de D modelados por ID estable.",
        ["Recipe Id", "Level/Parent", "Primary Agent", "Path", "Output/Purpose", "Bytes", "Modified", "Source Index", "Workbook Status"],
        d_recipe_registry_rows,
        "tblDRecipesRegistry",
        [58, 34, 46, 100, 70, 14, 22, 96, 34],
    )

    create_sheet(
        wb,
        "D Tools Registry",
        "TOOL_INDEX y permisos/fuentes TCU de D; distingue tool_id de archivo fisico.",
        ["Tool Id", "Level/Source", "Tool Type", "Path Or Command", "Physical Path Present", "Allowed Surface", "Blocked Surface", "Source Index", "Workbook Status"],
        d_tools_registry_rows,
        "tblDToolsRegistry",
        [52, 34, 24, 110, 18, 58, 58, 96, 34],
    )

    create_sheet(
        wb,
        "D Validators",
        "Validators y runners fisicos de D; inventariados pero no ejecutados en esta pasada.",
        ["Validator Name", "Path", "Linked Tool Id", "Linked Recipe", "Linked Subskill", "Present In Tools", "Present In Validacion", "Last Result", "Execution Status", "Kind"],
        d_validator_rows,
        "tblDValidators",
        [54, 100, 50, 42, 42, 16, 20, 24, 34, 22],
    )

    create_sheet(
        wb,
        "D Agent Registry",
        "AGENTS_INDEX y agents.json de D: roster canonico, perfil, workpaper y defaults operativos.",
        [
            "Agent Id",
            "Name",
            "Level Id",
            "Layer",
            "Primary Surface",
            "Profile Path",
            "Workpaper Path",
            "Mission",
            "Default Skills",
            "Default Recipes",
            "Default Tools",
            "Default Plugins",
            "Status",
            "Source Index",
        ],
        d_agent_registry_rows,
        "tblDAgentRegistry",
        [44, 46, 24, 32, 70, 100, 92, 90, 120, 120, 120, 34, 34, 96],
    )

    create_sheet(
        wb,
        "D Levels",
        "Subniveles de agentes en D y politica de lectura por nivel.",
        ["Level Id", "Name", "Folder", "Primary Role", "Agent Count", "Default Read", "Source Index"],
        d_levels_rows,
        "tblDLevels",
        [28, 34, 92, 70, 14, 70, 96],
    )

    create_sheet(
        wb,
        "D Routing",
        "routing.json de D: clases de orden, señales, agentes y handoff rules.",
        ["Kind", "Order/When", "Signals", "Agents", "Default Agent", "Must Include", "Notes", "Status", "Source Path"],
        d_routing_rows,
        "tblDRouting",
        [18, 42, 96, 70, 38, 120, 60, 34, 96],
    )

    create_sheet(
        wb,
        "D Workpapers",
        "Workpapers por agente: matriz canonica, conteo fisico y carpetas extra no canonicas.",
        [
            "Agent Id",
            "Level Id",
            "Workpaper Path",
            "File Count",
            "Canonical Agent",
            "Status",
            "Primary Surface",
            "Purpose",
            "Owner Agent",
            "Required Matrices",
            "Required Recipes",
            "Required Tools",
            "Required Validators",
            "Evidence Policy",
            "Validator",
            "Stop Condition",
            "Source Path",
        ],
        d_workpaper_rows,
        "tblDWorkpapers",
        [46, 26, 100, 12, 16, 34, 62, 90, 44, 90, 90, 90, 90, 42, 70, 54, 100],
    )

    create_sheet(
        wb,
        "D Orders",
        "Clases de orden gobernada y archivos de orden detectados en D.",
        [
            "Order Class",
            "Preparer Agent",
            "Reviewer Agent",
            "Approver Role",
            "Canon As Of",
            "Source Authority",
            "Required Fields",
            "Allowed Actions",
            "Blocked Actions",
            "Recipe",
            "Tool",
            "Evidence",
            "Validator",
            "Expiration Rule",
            "Stop Condition",
            "Source Path",
        ],
        d_order_rows,
        "tblDOrders",
        [42, 44, 44, 24, 18, 60, 120, 90, 100, 54, 54, 46, 90, 42, 54, 100],
    )

    create_sheet(
        wb,
        "D Readbacks",
        "Registry de readbacks/evidencia de D y archivos fisicos no registrados.",
        [
            "Evidence Id",
            "Readback Path",
            "Source Repo",
            "Repo Id",
            "Lane Id",
            "Agent Id",
            "Validator",
            "Status",
            "Proves",
            "Does Not Prove",
            "Next Gate",
            "Evidence Category",
            "Canonical Weight",
            "Source Path",
        ],
        d_readback_rows,
        "tblDReadbacks",
        [54, 110, 46, 36, 46, 42, 70, 32, 90, 90, 54, 28, 20, 100],
    )

    create_sheet(
        wb,
        "D Agent Tool Skill",
        "Cruce agente-skill-receta-tool-plugin-validator de D.",
        [
            "Agent Id",
            "Level Id",
            "Purpose",
            "Surface",
            "Skill Refs",
            "Recipe Refs",
            "Tool Refs",
            "Plugin Refs",
            "Workpaper Path",
            "Validator",
            "Default Mode",
            "Stop Condition",
            "Source Path",
        ],
        d_agent_tool_skill_rows,
        "tblDAgentToolSkill",
        [46, 24, 90, 110, 120, 120, 120, 42, 100, 80, 24, 70, 96],
    )

    create_sheet(
        wb,
        "D Authority Canon",
        "Archivos rectores de D/02_AUTHORITY_CANON con tamaño, shape y primera lectura.",
        ["Name", "Kind", "Path", "Bytes", "Modified", "Rows", "Columns", "Summary", "Source"],
        d_authority_rows,
        "tblDAuthorityCanon",
        [58, 24, 100, 14, 22, 10, 92, 90, 26],
    )

    create_sheet(
        wb,
        "D Governed Assets",
        "Inventario canonico de assets gobernados de D, owner, matriz, receta, tool y stop condition.",
        [
            "Asset Class",
            "Asset Id",
            "Owner Agent",
            "Authority Level",
            "Governing Matrix",
            "Required Recipe",
            "Required Tool",
            "Evidence",
            "Validator",
            "Coverage Status",
            "Stop Condition",
            "Source Path",
        ],
        d_governed_asset_rows,
        "tblDGovernedAssets",
        [28, 70, 44, 28, 92, 62, 62, 46, 70, 24, 60, 96],
    )

    create_sheet(
        wb,
        "D Source Refs",
        "Referencias fuente SOURCE_* y documentos estructurales D modelados como cobertura read-only.",
        [
            "Source Id",
            "Source Path",
            "Domain",
            "Source Type",
            "Recommended Sheet",
            "Priority",
            "Rows",
            "Columns",
            "Bytes",
            "Modified",
            "Summary",
            "Workbook Status",
        ],
        d_source_ref_rows,
        "tblDSourceRefs",
        [54, 110, 34, 24, 34, 12, 10, 92, 14, 22, 90, 34],
    )

    create_sheet(
        wb,
        "D Registries Ledgers",
        "Registries, ledgers e indexes de D con dominio, owner/reviewer cuando aparece y stop condition.",
        [
            "Artifact Id",
            "Type",
            "Path",
            "Domain",
            "Owner Agent",
            "Reviewer Agent",
            "Evidence Ref",
            "Status",
            "Stop Condition",
            "Rows",
            "Columns",
            "Priority",
            "Recommended Sheet",
        ],
        d_registries_ledgers_rows,
        "tblDRegistriesLedgers",
        [54, 18, 110, 34, 44, 44, 44, 30, 60, 10, 92, 12, 34],
    )

    create_sheet(
        wb,
        "D Domain Matrices",
        "Matrices, seeds y CSV por dominio de D: conexiones, dataverse, powerautomate, runtime y gobernanza.",
        [
            "Domain",
            "File",
            "Scope",
            "Surface",
            "Dev/Live Boundary",
            "Owner",
            "Gate",
            "Postcheck",
            "Rows",
            "Columns",
            "Priority",
            "Ruta Fuente",
        ],
        d_domain_matrix_rows,
        "tblDDomainMatrices",
        [34, 62, 24, 90, 54, 44, 54, 70, 10, 92, 12, 110],
    )

    create_sheet(
        wb,
        "D Coverage Audit",
        "Auditoría de cobertura D por path/file/id; convierte faltantes amplios en próximos deltas priorizados.",
        [
            "Candidate Path",
            "Candidate Type",
            "Domain",
            "Matched By Path",
            "Matched By File",
            "Matched By Id",
            "Missing Reason",
            "Next Delta",
            "Recommended Sheet",
            "Priority",
            "Bytes",
            "Modified",
        ],
        d_coverage_audit_rows,
        "tblDCoverageAudit",
        [120, 24, 34, 16, 16, 14, 34, 90, 34, 12, 14, 22],
    )

    create_sheet(
        wb,
        "D Workbook Gaps",
        "Gaps detectados por agentes y estado de absorcion dentro de esta actualizacion del workbook.",
        ["Gap Id", "Target Sheet", "Item Type", "Absolute Path", "Current Presence", "Expected Presence", "Proposed Action"],
        d_gap_rows,
        "tblDWorkbookGaps",
        [44, 34, 34, 100, 34, 76, 82],
    )

    global_rows = []
    for prefix, source in [("persisted", persisted), ("root", {k: v for k, v in state.items() if k != "electron-persisted-atom-state"})]:
        for key, value in source.items():
            typ = "array" if isinstance(value, list) else "object" if isinstance(value, dict) else "null" if value is None else type(value).__name__
            children = len(value) if isinstance(value, (list, dict)) else None
            size = json_size(value)
            action = "REVISAR_DRY_RUN" if key == "prompt-history" else "NO_TOCAR"
            risk = "Historial textual" if key == "prompt-history" else "Estado vivo"
            if key in {"active-workspace-roots", "electron-saved-workspace-roots", "electron-workspace-root-labels"}:
                risk = "Workspace vivo"
            global_rows.append([f"{prefix}.{key}", typ, children, size, kb(size), action, risk])
    global_rows.sort(key=lambda row: (-(row[4] or 0), row[0]))
    create_sheet(wb, "Global State", "Secciones root y persisted atom state, tamaño aproximado y acción recomendada.", ["Seccion", "Tipo", "Hijos", "Bytes aprox", "KB aprox", "Accion recomendada", "Riesgo"], global_rows, "tblGlobalState", [58, 16, 12, 16, 14, 24, 26])

    prompt_headers = ["Grupo", "Entradas", "Chars aprox", "KB aprox", "Max entrada chars", "Menciones sensibles", "Permiso hilo", "Workspace hint", "Updated at", "Clase", "Accion recomendada", "Riesgo", "Decision humana", "Nota"]
    ws = create_sheet(wb, "Prompt History", "No incluye contenidos de prompts. Solo conteos, tamaños, flags y decisión.", prompt_headers, prompt_rows, "tblPromptHistory", [44, 12, 14, 12, 18, 18, 14, 16, 20, 28, 24, 18, 18, 70])
    dv = DataValidation(type="list", formula1='"CONSERVAR,REMOVER_DRY_RUN,REVISAR_OWNER,REVISAR_SENSIBLE,NO_TOCAR"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"M5:M{4 + len(prompt_rows)}")
    create_sheet(
        wb,
        "Prompts Lectura",
        "Texto de prompts leído desde global-state. Tokens reales y asignaciones sensibles se redactan; se conserva hash original.",
        [
            "Grupo",
            "Entrada",
            "Chunk",
            "Clase",
            "Accion",
            "Riesgo",
            "Chars original",
            "SHA256 original",
            "Redacciones",
            "Texto saneado",
        ],
        prompt_read_rows,
        "tblPromptsLectura",
        [44, 10, 10, 26, 22, 18, 14, 68, 14, 120],
    )
    create_sheet(wb, "Candidatos", "Lista operable para decidir Wave 2 sin tocar contenidos.", prompt_headers, candidate_rows, "tblCandidatos", [44, 12, 14, 12, 18, 18, 14, 16, 20, 28, 24, 18, 18, 70])

    create_sheet(
        wb,
        "Repositorios",
        "Inventario local de repositorios y proyectos trusted. Lectura Git sin cambios.",
        [
            "Fuente",
            "Nombre",
            "Ruta",
            "Existe",
            "Es Git",
            "Git root",
            "Branch",
            "HEAD",
            "Remote origin",
            "Dirty count",
            "Decision",
            "Nota",
        ],
        repo_rows,
        "tblRepositorios",
        [24, 34, 92, 12, 12, 92, 24, 16, 70, 14, 18, 48],
    )

    create_sheet(
        wb,
        "Ramas",
        "Ramas locales, remote-tracking y ramas GitHub live por repositorio.",
        [
            "Fuente",
            "Repo",
            "Nombre Local",
            "Git Root",
            "Branch",
            "Actual",
            "Upstream",
            "Ahead",
            "Behind",
            "Commit",
            "Commit Date",
            "Autor",
            "Protected",
            "Default",
            "Estado",
            "Nota",
        ],
        branch_rows,
        "tblRamas",
        [24, 42, 32, 92, 48, 12, 40, 10, 10, 18, 28, 28, 14, 14, 24, 70],
    )

    create_sheet(
        wb,
        "PRs",
        "Pull requests leidos con GitHub live via gh; incluye abiertos, draft e historicos recientes por repo.",
        [
            "Fuente",
            "Repo",
            "Numero",
            "Estado",
            "Draft",
            "Head",
            "Base",
            "Titulo",
            "Autor",
            "Created At",
            "Updated At",
            "Merge State",
            "Review Decision",
            "Accion",
            "URL/Nota",
        ],
        pr_rows,
        "tblPRs",
        [26, 42, 12, 16, 10, 38, 22, 80, 24, 26, 26, 20, 22, 24, 90],
    )

    create_sheet(
        wb,
        "Codex Cloud Env",
        "Entornos Codex Cloud confirmados localmente y menciones detectadas en prompts.",
        [
            "Fuente",
            "Environment Id",
            "Label/Repo",
            "Workspace",
            "Repos",
            "Network",
            "Tasks",
            "Secrets",
            "Cache",
            "Setup",
            "Maintenance",
            "Estado",
            "Nota",
        ],
        cloud_rows,
        "tblCodexCloudEnv",
        [24, 38, 42, 22, 50, 16, 10, 10, 12, 70, 70, 24, 60],
    )

    create_sheet(
        wb,
        "Agentes",
        "Agentes institucionales, SDU, corte ejecutora y componentes Dataverse con evidencia local/live previa.",
        [
            "Fuente",
            "Agente",
            "Rol",
            "Superficie",
            "Canonical ID",
            "Estado",
            "Evidencia",
            "Riesgo",
            "Proximo Delta",
            "Ruta Fuente",
        ],
        agent_rows,
        "tblAgentes",
        [34, 36, 28, 42, 56, 28, 70, 54, 54, 92],
    )

    create_sheet(
        wb,
        "Agentes Entorno",
        "Agentes por entorno y solucion: SDU agent mappings y bots/copilots Power Platform.",
        [
            "Fuente",
            "Entorno",
            "Environment URL",
            "Environment Id",
            "Solucion",
            "Tabla",
            "Agente/Bot",
            "Canonical/Schema",
            "Owner Agent",
            "Reviewer Agent",
            "Superficies",
            "Estado",
            "Ultima Reconciliacion",
            "Row Id",
            "Riesgo/State",
            "Ruta Fuente",
        ],
        environment_agent_rows,
        "tblAgentesEntorno",
        [30, 24, 56, 38, 36, 34, 46, 58, 34, 34, 90, 24, 28, 42, 34, 92],
    )

    create_sheet(
        wb,
        "Entornos Soluciones",
        "Entorno Dataverse/Power Platform, soluciones y conteos de tablas vivos previamente observados.",
        [
            "Fuente",
            "Entorno",
            "Environment URL",
            "Environment Id",
            "Tipo",
            "Nombre",
            "Friendly Name",
            "Version",
            "Managed",
            "Component/Count",
            "Estado",
            "Nota",
            "Ruta Fuente",
        ],
        environment_solution_rows,
        "tblEntornosSoluciones",
        [32, 24, 56, 38, 20, 44, 54, 18, 14, 18, 24, 70, 92],
    )

    create_sheet(
        wb,
        "Colas",
        "Workqueues Dataverse SDU con backlog y accion atomica recomendada.",
        [
            "Fuente",
            "Entorno",
            "Environment URL",
            "Queue",
            "Workqueue Id",
            "Statecode",
            "Statuscode",
            "Retry Limit",
            "Requeue Limit",
            "Priority Type",
            "Queue Type",
            "Backlog",
            "Estado",
            "Accion",
            "Ruta Fuente",
        ],
        workqueue_rows,
        "tblColas",
        [30, 24, 56, 44, 42, 12, 12, 14, 16, 16, 16, 12, 24, 28, 92],
    )

    create_sheet(
        wb,
        "Universos",
        "Superficies canonicas y auditoria de relaciones entre universos locales, repos y cabinas.",
        ["Fuente", "Universo", "Origen", "Tipo", "Destino", "Dependencia", "Estado", "Evidencia"],
        universe_rows,
        "tblUniversos",
        [28, 34, 92, 22, 92, 72, 24, 70],
    )

    create_sheet(
        wb,
        "Skills",
        "Skills disponibles desde tabla unificada, raices locales y cache de plugins.",
        ["RootLabel", "Kind", "Family", "Canonical", "Alias", "Purpose", "SourcePath", "Estado", "Notes"],
        skill_rows,
        "tblSkills",
        [24, 16, 36, 38, 34, 90, 100, 24, 70],
    )

    create_sheet(
        wb,
        "Recetas",
        "Recetas operativas locales y .agents, con indice y ruta de fuente.",
        ["Fuente", "Nombre", "Tipo", "Ruta", "Bytes", "Modified", "Estado", "Resumen", "Indice/Fuente"],
        recipe_rows,
        "tblRecetas",
        [24, 42, 22, 100, 14, 18, 22, 90, 90],
    )

    create_sheet(
        wb,
        "Tools",
        "Herramientas locales, scripts, servidores MCP y plugins habilitados.",
        ["Fuente", "Nombre", "Tipo", "Ruta", "Bytes", "Modified", "Estado", "Resumen"],
        tool_rows,
        "tblTools",
        [24, 44, 20, 100, 14, 18, 24, 90],
    )

    create_sheet(
        wb,
        "Conexiones",
        "Superficies de conexion, gates, mappings agente-repo, MCP, plugins y Codex Cloud.",
        [
            "Fuente",
            "Tipo",
            "Surface/Conexion",
            "Provider",
            "Gate",
            "Estado",
            "Live Enabled",
            "Owner Agent",
            "Reviewer Agent",
            "Evidencia",
            "Stop Condition",
            "Ruta Fuente",
        ],
        connection_rows,
        "tblConexiones",
        [34, 24, 44, 26, 42, 28, 16, 34, 34, 90, 52, 92],
    )

    create_sheet(
        wb,
        "Dataverse Fuentes",
        "Tablas Dataverse que alimentan o pueden alimentar el workbook; incluye rehidratacion live read-only previa.",
        [
            "Fuente",
            "Rol",
            "Logical Name",
            "Entity Set",
            "Solution/Environment",
            "Estado",
            "Conteo/Id",
            "Canonical/Schema",
            "Evidencia",
            "Nota",
            "Ruta Fuente",
        ],
        dataverse_source_rows,
        "tblDataverseFuentes",
        [40, 34, 36, 38, 46, 28, 42, 70, 46, 90, 92],
    )

    create_sheet(
        wb,
        "Matriz Maestra",
        "Matriz atomica rectora: waves, agentes, superficie, gates, evidencia, rollback, stop condition y next delta.",
        [
            "Wave Id",
            "Agente Owner",
            "Agente Reviewer",
            "Owner Humano",
            "Decision Humana",
            "Approval Status",
            "Execution Actor",
            "Surface",
            "Read Scope",
            "Write Scope",
            "Lock Key",
            "Estado Operativo",
            "Audit Status",
            "Environment Classification",
            "Permission Scope",
            "Candidate Count",
            "Allowed Action",
            "Blocked Action",
            "Evidencia",
            "Evidence Sink",
            "Validador",
            "Postcheck",
            "Rollback",
            "Stop Condition",
            "Next Delta",
            "Ruta Fuente",
        ],
        master_atomic_rows,
        "tblMatrizMaestra",
        [24, 22, 22, 24, 18, 28, 36, 38, 70, 54, 54, 26, 30, 34, 28, 18, 42, 60, 92, 60, 62, 70, 62, 46, 54, 92],
    )

    create_sheet(
        wb,
        "Acciones Atomicas",
        "Diccionario operativo de acciones atomicas: mutabilidad, lock, rollback y validacion requerida.",
        [
            "Action Type",
            "Purpose",
            "Target Surface",
            "Mutating",
            "Requires Lock",
            "Rollback/Compensation",
            "Validation Focus",
            "Terminal State",
            "Notes",
            "Ruta Fuente",
        ],
        atomic_action_rows,
        "tblAccionesAtomicas",
        [32, 60, 46, 12, 14, 60, 54, 18, 70, 92],
    )

    create_sheet(
        wb,
        "Matriz Hidratacion",
        "Matriz metadata-only para hidratar Dataverse SDU por canonical_id, target, validator y stop condition.",
        [
            "Wave Id",
            "Metadata Id",
            "Tenant Id",
            "Environment",
            "Surface",
            "Source Artifact",
            "Target Entity Set",
            "Target Logical Name",
            "Canonical Id",
            "Display Name",
            "Classification",
            "Status",
            "Owner",
            "Evidence Ref",
            "Rollback",
            "Postcheck",
            "Validator",
            "Stop Condition",
            "Notes",
            "Ruta Fuente",
        ],
        metadata_hydration_rows,
        "tblMatrizHidratacion",
        [28, 28, 38, 24, 22, 62, 34, 34, 58, 46, 22, 26, 28, 72, 70, 82, 62, 54, 70, 92],
    )

    create_sheet(
        wb,
        "Deltas Atomicos",
        "Siguiente movimiento operable por superficie: PRs, repos dirty, colas, Cloud y Dataverse rehidratado.",
        [
            "Orden",
            "Superficie",
            "Target",
            "Detalle",
            "Accion",
            "Riesgo",
            "Postcheck",
            "Evidencia",
            "Stop Condition",
            "Estado",
        ],
        atomic_delta_rows,
        "tblDeltasAtomicos",
        [10, 24, 44, 72, 34, 18, 70, 92, 70, 36],
    )

    risk_rows = [
        ["NO_TOCAR", "environment", "Estado vivo de entorno cloud/local", "Si un cambio toca environment, detener."],
        ["NO_TOCAR", "secrets/auth/cap_sid/SQLite", "Secretos y runtime interno", "Cualquier secreto o auth requiere gate separado."],
        ["NO_TOCAR", "heartbeat-thread-permissions-by-id", "Permisos de hilos", "No borrar permisos por inferencia."],
        ["RIESGO", "prompt-history", "Historial textual sensible, no secreto estructural", "No reemplazar texto globalmente."],
        ["GATE", "D read/write", "D es consola rectora local gobernada", "Escritura requiere orden humana, target, rollback y postcheck."],
        ["GATE", "Wave 2 dry-run", "Primer movimiento seguro", "Debe producir lista exacta sin escribir."],
        ["GATE", "Wave 3 escritura", "Requiere backup y Codex cerrado", "Si hash cambia durante operación, abortar."],
        ["POSTCHECK", "JSON", "Debe parsear después", "Si falla parseo, restaurar backup."],
        ["POSTCHECK", "D root", "Debe conservar AGENTS/MANIFEST visibles", "Si desaparece D o cambia frontera, detener y revalidar."],
        ["POSTCHECK", "environment", "Debe quedar idéntico", "Si cambia id/label/repo/secrets, rollback."],
    ]
    create_sheet(wb, "Riesgos Gates", "Fronteras para mantener limpieza atómica y cabina D segura.", ["Tipo", "Elemento", "Motivo", "Stop condition"], risk_rows, "tblRiesgos", [18, 34, 58, 68])

    lane_rows = [
        ["parallel_readonly_scouts", "rey.control_plane_orchestrator", "court.seshat_evidence", "read_only", "lock.readonly.scout", "agent_final_readback", "OK"],
        ["d_drive_governed_console", "rey.control_plane_orchestrator", "court.anubis_gate", "D:/AGENTS|D:/MANIFEST|D:/REPO_SCOPE", "lock.d.governed", str(D_ROOT), d_status],
        ["global_state_safe_dry_run", "rey.control_plane_orchestrator", "court.seshat_evidence", "prompt-history candidates", "lock.global_state", "dry-run manifest", "PREPARADO"],
    ]
    create_sheet(wb, "Agentes Carriles", "Carriles gobernados usados para D y limpieza futura.", ["lane_id", "lead_agent", "reviewer_agent", "read_scope", "lock_key", "evidence", "status"], lane_rows, "tblParallel", [32, 34, 28, 46, 28, 60, 16])

    env_text = json.dumps(env, ensure_ascii=False, sort_keys=True)
    ph_text = json.dumps(prompt_history, ensure_ascii=False, sort_keys=True)
    other_text = json.dumps({k: v for k, v in persisted.items() if k not in {"environment", "prompt-history"}}, ensure_ascii=False, sort_keys=True)
    secret_rows = []
    for zone, text in [("environment", env_text), ("prompt-history", ph_text), ("other-state", other_text)]:
        counts = count_mentions(text)
        for pattern in text_patterns:
            secret_rows.append([zone, pattern, counts.get(pattern, 0), "token_real_pattern" if pattern.endswith("_like") else "mencion_textual"])

    validation_rows = [
        ["STATE_PATH", str(STATE_PATH), "INFO"],
        ["STATE_SHA256", state_hash, "INFO"],
        ["STATE_SIZE_BYTES", STATE_PATH.stat().st_size, "INFO"],
        ["BACKUP_PATH", str(STATE_BAK), "INFO" if STATE_BAK.exists() else "MISSING"],
        ["BACKUP_SHA256", backup_hash or "N/D", "INFO" if backup_hash else "MISSING"],
        ["BACKUP_IDENTICAL", backup_identical, "PASS" if backup_identical == "SI" else "OBSERVAR"],
        ["JSON_PARSE", "VALID", "PASS"],
        ["ENVIRONMENT_LABEL", env_label, "PASS" if env_label != "N/D" else "OBSERVAR"],
        ["SECRET_COUNT", len(env.get("secrets") or []), "PASS"],
        ["PROMPT_HISTORY_GROUPS", ph_groups, "INFO"],
        ["PROMPT_HISTORY_ENTRIES", ph_entries, "INFO"],
        ["PROMPTS_LECTURA_ROWS", len(prompt_read_rows), "INFO"],
        ["REPOSITORIES_ROWS", len(repo_rows), "INFO"],
        ["BRANCH_ROWS", len(branch_rows), "INFO"],
        ["PR_ROWS", len(pr_rows), "INFO"],
        ["CODEX_CLOUD_ENV_ROWS", len(cloud_rows), "INFO"],
        ["AGENTS_ROWS", len(agent_rows), "INFO"],
        ["ENVIRONMENT_AGENT_ROWS", len(environment_agent_rows), "INFO"],
        ["ENVIRONMENT_SOLUTION_ROWS", len(environment_solution_rows), "INFO"],
        ["WORKQUEUE_ROWS", len(workqueue_rows), "INFO"],
        ["UNIVERSES_ROWS", len(universe_rows), "INFO"],
        ["SKILLS_ROWS", len(skill_rows), "INFO"],
        ["RECIPES_ROWS", len(recipe_rows), "INFO"],
        ["TOOLS_ROWS", len(tool_rows), "INFO"],
        ["CONNECTIONS_ROWS", len(connection_rows), "INFO"],
        ["DATAVERSE_SOURCE_ROWS", len(dataverse_source_rows), "INFO"],
        ["MASTER_ATOMIC_MATRIX_ROWS", len(master_atomic_rows), "INFO"],
        ["ATOMIC_ACTION_ROWS", len(atomic_action_rows), "INFO"],
        ["METADATA_HYDRATION_ROWS", len(metadata_hydration_rows), "INFO"],
        ["ATOMIC_DELTA_ROWS", len(atomic_delta_rows), "INFO"],
        ["D_SURFACE_ROWS", len(d_surface_rows), "INFO"],
        ["D_INDEX_ROWS", len(d_indexes_rows), "INFO"],
        ["D_MATRIX_INDEX_ROWS", len(d_matrix_index_rows), "INFO"],
        ["D_SKILL_REGISTRY_ROWS", len(d_skill_registry_rows), "INFO"],
        ["D_SUBSKILL_ROWS", len(d_subskill_rows), "INFO"],
        ["D_RECIPE_REGISTRY_ROWS", len(d_recipe_registry_rows), "INFO"],
        ["D_TOOL_REGISTRY_ROWS", len(d_tools_registry_rows), "INFO"],
        ["D_VALIDATOR_ROWS", len(d_validator_rows), "INFO"],
        ["D_AGENT_REGISTRY_ROWS", len(d_agent_registry_rows), "INFO"],
        ["D_ROUTING_ROWS", len(d_routing_rows), "INFO"],
        ["D_WORKPAPER_ROWS", len(d_workpaper_rows), "INFO"],
        ["D_ORDER_ROWS", len(d_order_rows), "INFO"],
        ["D_READBACK_ROWS", len(d_readback_rows), "INFO"],
        ["D_AGENT_TOOL_SKILL_ROWS", len(d_agent_tool_skill_rows), "INFO"],
        ["D_AUTHORITY_ROWS", len(d_authority_rows), "INFO"],
        ["D_GOVERNED_ASSET_ROWS", len(d_governed_asset_rows), "INFO"],
        ["D_SOURCE_REF_ROWS", len(d_source_ref_rows), "INFO"],
        ["D_REGISTRIES_LEDGERS_ROWS", len(d_registries_ledgers_rows), "INFO"],
        ["D_DOMAIN_MATRIX_ROWS", len(d_domain_matrix_rows), "INFO"],
        ["D_COVERAGE_AUDIT_ROWS", len(d_coverage_audit_rows), "INFO"],
        ["D_WORKBOOK_GAP_ROWS", len(d_gap_rows), "INFO"],
        ["LOCAL_SURFACE_ROWS", len(local_surface_rows), "INFO"],
        ["WORKSPACE_ROWS", len(workspace_rows), "INFO"],
        ["D_VALIDATORS_EXECUTED", "NO", "INFO"],
        ["D_DRIVE_PRESENT", d_exists, "PASS" if d_exists else "FAIL"],
        ["D_DRIVE_STATUS", d_status, "PASS" if d_exists else "FAIL"],
        ["D_DRIVE_READONLY", d_readonly, "INFO"],
        ["D_AGENTS", str(d_agents), "PASS" if d_agents.exists() else "MISSING"],
        ["D_MANIFEST", str(d_manifest), "PASS" if d_manifest.exists() else "MISSING"],
        ["D_EVIDENCE_LOG", str(EVIDENCE_LOG), "PASS" if EVIDENCE_LOG.exists() else "MISSING"],
        ["WORKBOOK_BACKUP", str(backup_path) if backup_path else "N/D", "PASS" if backup_path else "INFO"],
    ]
    ws = create_sheet(wb, "Validacion", "Hashes, parseo, escaneo saneado, D read-only y postchecks.", ["Check", "Resultado", "Estado"], validation_rows, "tblValidation", [30, 100, 18])
    start = 4 + len(validation_rows) + 3
    ws.cell(start - 1, 1, "Escaneo De Patrones Saneado").font = Font(bold=True, color=WHITE)
    ws.cell(start - 1, 1).fill = PatternFill("solid", fgColor=HEADER_FILL)
    for col, header in enumerate(["Zona", "Patron", "Conteo textual", "Tipo"], 1):
        cell = ws.cell(start, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.border = BORDER
    for row_index, row in enumerate(secret_rows, start + 1):
        for col_index, value in enumerate(row, 1):
            cell = ws.cell(row_index, col_index, value)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    add_table(ws, "tblSecretScan", start, ["Zona", "Patron", "Conteo textual", "Tipo"], secret_rows)

    nav_rows = [
        ["Archivo", "Global State", str(STATE_PATH), "Vivo / ignorado por Git"],
        ["Archivo", "Global State Backup", str(STATE_BAK), "Rollback inmediato" if STATE_BAK.exists() else "No encontrado"],
        ["Archivo", "Codex config", str(CONFIG_PATH), "Config técnica"],
        ["Archivo", "Codex AGENTS", r"C:\Users\enzo1\.codex\AGENTS.md", "Prompt global visible"],
        ["Archivo", "PROJEC CDX AGENTS", str(ROOT / "AGENTS.md"), "Instrucción local"],
        ["Superficie", "Workspace actual", str(ROOT), f"{len(workspace_rows)} filas en Workspace Actual"],
        ["Superficie", "Mapa superficies locales", "Superficies Locales", f"{len(local_surface_rows)} superficies normalizadas"],
        ["Superficie", "Cabina D root", r"D:\\", d_status if d_exists else "No montada"],
        ["Superficie", "CodexLocal live entry", str(HOME_CODEX_LOCAL), "Entrada viva liviana; requiere indice si se promueve"],
        ["Superficie", "CodexLocal legacy", str(DOCUMENTS_CODEX_LOCAL), "Workspace legado pesado"],
        ["Superficie", "Documents Codex chronology", str(DOCUMENTS_CODEX), "Cronologia documental"],
        ["Superficie", "Documents CodexArchives", str(DOCUMENTS_CODEX_ARCHIVES), "Archivo reversible"],
        ["Superficie", "GitHub repos canonicos", str(GITHUB_ROOT), "Raiz de repositorios"],
        ["Archivo", "D AGENTS", str(d_agents), "PASS" if d_agents.exists() else "MISSING"],
        ["Archivo", "D MANIFEST", str(d_manifest), "PASS" if d_manifest.exists() else "MISSING"],
        ["Archivo", "D validator agent", str(d_validator_agent), "PASS" if d_validator_agent.exists() else "MISSING"],
        ["Archivo", "D validator chain", str(d_validator_chain), "PASS" if d_validator_chain.exists() else "MISSING"],
        ["Directorio", "D Codex agent root", str(D_CODEX_ROOT), f"{len(d_surface_rows)} superficies D resumidas"],
        ["Directorio", "D matrices root", str(D_MATRICES_ROOT), f"{len(d_domain_matrix_rows)} matrices/seeds/csv de dominio"],
        ["Directorio", "D governance registry", str(D_GOVERNANCE_REGISTRY), "Incluido en auditoria de cobertura"],
        ["Directorio", "D authority canon", str(D_AUTHORITY_CANON), f"{len(d_authority_rows)} archivos rectores"],
        ["Directorio", "D dataverse data", str(D_DATAVERSE_DATA), "Seeds/ref data incluidos en cobertura"],
        ["Directorio", "D validation", str(D_VALIDATION_ROOT), "Reportes incluidos como evidencia read-only"],
        ["Archivo", "VHDX referencia historica", str(VHDX_PATH), "Referencia" if VHDX_PATH.exists() else "MISSING"],
        ["Archivo", "Attach/recovery evidence log", str(EVIDENCE_LOG), "INFO" if EVIDENCE_LOG.exists() else "MISSING"],
        ["Hoja", "Workspace Actual", "Workspace Actual", f"{len(workspace_rows)} filas"],
        ["Hoja", "Superficies Locales", "Superficies Locales", f"{len(local_surface_rows)} filas"],
        ["Hoja", "Prompts Lectura", "Prompts Lectura", f"{len(prompt_read_rows)} filas saneadas"],
        ["Hoja", "Repositorios", "Repositorios", f"{len(repo_rows)} filas"],
        ["Hoja", "Ramas", "Ramas", f"{len(branch_rows)} filas"],
        ["Hoja", "PRs", "PRs", f"{len(pr_rows)} filas"],
        ["Hoja", "Codex Cloud Env", "Codex Cloud Env", f"{len(cloud_rows)} filas"],
        ["Hoja", "Agentes", "Agentes", f"{len(agent_rows)} filas"],
        ["Hoja", "Agentes Entorno", "Agentes Entorno", f"{len(environment_agent_rows)} filas"],
        ["Hoja", "Entornos Soluciones", "Entornos Soluciones", f"{len(environment_solution_rows)} filas"],
        ["Hoja", "Colas", "Colas", f"{len(workqueue_rows)} filas"],
        ["Hoja", "Universos", "Universos", f"{len(universe_rows)} filas"],
        ["Hoja", "Skills", "Skills", f"{len(skill_rows)} filas"],
        ["Hoja", "Recetas", "Recetas", f"{len(recipe_rows)} filas"],
        ["Hoja", "Tools", "Tools", f"{len(tool_rows)} filas"],
        ["Hoja", "Conexiones", "Conexiones", f"{len(connection_rows)} filas"],
        ["Hoja", "Dataverse Fuentes", "Dataverse Fuentes", f"{len(dataverse_source_rows)} filas"],
        ["Hoja", "Matriz Maestra", "Matriz Maestra", f"{len(master_atomic_rows)} filas"],
        ["Hoja", "Acciones Atomicas", "Acciones Atomicas", f"{len(atomic_action_rows)} filas"],
        ["Hoja", "Matriz Hidratacion", "Matriz Hidratacion", f"{len(metadata_hydration_rows)} filas"],
        ["Hoja", "Deltas Atomicos", "Deltas Atomicos", f"{len(atomic_delta_rows)} filas"],
        ["Hoja", "D Surface Summary", "D Surface Summary", f"{len(d_surface_rows)} filas"],
        ["Hoja", "D Indexes", "D Indexes", f"{len(d_indexes_rows)} filas"],
        ["Hoja", "D Matrix Index", "D Matrix Index", f"{len(d_matrix_index_rows)} filas"],
        ["Hoja", "D Skills Registry", "D Skills Registry", f"{len(d_skill_registry_rows)} filas"],
        ["Hoja", "D Subskills", "D Subskills", f"{len(d_subskill_rows)} filas"],
        ["Hoja", "D Recipes Registry", "D Recipes Registry", f"{len(d_recipe_registry_rows)} filas"],
        ["Hoja", "D Tools Registry", "D Tools Registry", f"{len(d_tools_registry_rows)} filas"],
        ["Hoja", "D Validators", "D Validators", f"{len(d_validator_rows)} filas"],
        ["Hoja", "D Agent Registry", "D Agent Registry", f"{len(d_agent_registry_rows)} filas"],
        ["Hoja", "D Routing", "D Routing", f"{len(d_routing_rows)} filas"],
        ["Hoja", "D Workpapers", "D Workpapers", f"{len(d_workpaper_rows)} filas"],
        ["Hoja", "D Orders", "D Orders", f"{len(d_order_rows)} filas"],
        ["Hoja", "D Readbacks", "D Readbacks", f"{len(d_readback_rows)} filas"],
        ["Hoja", "D Agent Tool Skill", "D Agent Tool Skill", f"{len(d_agent_tool_skill_rows)} filas"],
        ["Hoja", "D Authority Canon", "D Authority Canon", f"{len(d_authority_rows)} filas"],
        ["Hoja", "D Governed Assets", "D Governed Assets", f"{len(d_governed_asset_rows)} filas"],
        ["Hoja", "D Source Refs", "D Source Refs", f"{len(d_source_ref_rows)} filas"],
        ["Hoja", "D Registries Ledgers", "D Registries Ledgers", f"{len(d_registries_ledgers_rows)} filas"],
        ["Hoja", "D Domain Matrices", "D Domain Matrices", f"{len(d_domain_matrix_rows)} filas"],
        ["Hoja", "D Coverage Audit", "D Coverage Audit", f"{len(d_coverage_audit_rows)} filas"],
        ["Hoja", "D Workbook Gaps", "D Workbook Gaps", f"{len(d_gap_rows)} filas"],
        ["Dataverse", "Source artifacts", "mon_sdu_source_artifacts", "Tabla viva confirmada para artefactos fuente"],
        ["Dataverse", "Evidences", "mon_sdu_evidences", "Tabla viva confirmada para evidencias"],
        ["Dataverse", "Agent mapping", "mon_sdu_agent_connection_mappings", "Tabla viva/postcheck para roster SDU"],
        ["Dataverse", "Work queues", "workqueues / workqueueitems", "Tablas vivas para cola SDU"],
        ["Workbook", "Este workbook", str(WORKBOOK), "Decision"],
        ["Workbook", "Backup anterior", str(backup_path) if backup_path else "N/D", "Rollback workbook" if backup_path else "N/D"],
    ]
    create_sheet(wb, "Navegacion", "Entradas, fuentes y rutas de decisión.", ["Tipo", "Nombre", "Ruta", "Estado"], nav_rows, "tblNavegacion", [18, 30, 100, 28])

    list_rows = []
    for value in ["CONSERVAR", "REMOVER_DRY_RUN", "REVISAR_OWNER", "REVISAR_SENSIBLE", "NO_TOCAR", "ACTUAL", "APROBADO", "ESPERAR", "RECHAZADO"]:
        list_rows.append(["Decision humana", value])
    for value in [
        "PASS",
        "INFO",
        "OBSERVAR",
        "PENDIENTE",
        "MISSING",
        "FAIL",
        "VERIFICAR_CLOUD_LIVE",
        "LIVE_TABLE_CONFIRMED",
        "LIVE_ROWS_CONFIRMED",
        "POSTCHECKED",
        "LOCAL_INVENTORY_GATED",
        "METADATA_ONLY",
        "GITHUB_LIVE_READ",
        "REVISAR_ABIERTO",
        "REVISAR_DRAFT",
        "DELTA_PRIORIZAR_BACKLOG",
        "DELTA_TRIAGE_CAMBIOS",
        "OK_BASE_REHIDRATADA",
        "APPROVED_FOR_PREPARATION",
        "APPROVED_FOR_READ_ONLY_SCOUTING",
        "LOCAL_PASS_CLOUD_TASK_PENDING",
        "METADATA_ONLY_PREPARED",
        "PENDING_W1_CLOSEOUT",
        "PENDING_TARGET_ONLY",
        "ADDED_CURRENT_RUN",
        "MODELED_CURRENT_RUN",
        "MODELED_DIRECTLY",
        "MODELED_IN_COVERAGE_SHEET",
        "PENDING_FINE_MODEL",
        "PENDING_EXECUTION",
        "AVAILABLE_READONLY",
        "MONTADA_GOBERNADA",
        "NO_MONTADA",
        "ENTRYPOINT_INCOMPLETE",
        "LEGACY_HEAVY_INDEXED",
        "CHRONOLOGY_INDEXED",
        "ARCHIVE_REVERSIBLE",
        "INDEXED",
        "REFERENCIA",
        "ACTUAL",
        "MISSING_EXPECTED_OR_NOT_APPLICABLE",
        "NOT_EXECUTED_READONLY_INVENTORY",
    ]:
        list_rows.append(["Estado", value])
    create_sheet(wb, "Listas", "Valores para dropdowns y criterios.", ["Lista", "Valor"], list_rows, "tblListas", [28, 42])

    wb.properties.title = "Codex Global State Decision Workbook"
    wb.properties.subject = "Configuración vigente Codex, Cabina D, global-state y limpieza gobernada"
    wb.properties.creator = "Codex"
    wb.properties.lastModifiedBy = "Codex"
    wb.properties.modified = now
    WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK)

    verified = load_workbook(WORKBOOK, data_only=False)
    formula_errors = []
    for ws in verified.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}:
                    formula_errors.append([ws.title, cell.coordinate, cell.value])
    print(
        json.dumps(
            {
                "workbook": str(WORKBOOK),
                "backup": str(backup_path) if backup_path else None,
                "sheet_count": len(verified.sheetnames),
                "sheets": verified.sheetnames,
                "prompt_rows": len(prompt_rows),
                "candidate_rows": len(candidate_rows),
                "branch_rows": len(branch_rows),
                "pr_rows": len(pr_rows),
                "agent_rows": len(agent_rows),
                "environment_agent_rows": len(environment_agent_rows),
                "environment_solution_rows": len(environment_solution_rows),
                "workqueue_rows": len(workqueue_rows),
                "universe_rows": len(universe_rows),
                "skill_rows": len(skill_rows),
                "recipe_rows": len(recipe_rows),
                "tool_rows": len(tool_rows),
                "connection_rows": len(connection_rows),
                "dataverse_source_rows": len(dataverse_source_rows),
                "master_atomic_rows": len(master_atomic_rows),
                "atomic_action_rows": len(atomic_action_rows),
                "metadata_hydration_rows": len(metadata_hydration_rows),
                "atomic_delta_rows": len(atomic_delta_rows),
                "d_surface_rows": len(d_surface_rows),
                "d_indexes_rows": len(d_indexes_rows),
                "d_matrix_index_rows": len(d_matrix_index_rows),
                "d_skill_registry_rows": len(d_skill_registry_rows),
                "d_subskill_rows": len(d_subskill_rows),
                "d_recipe_registry_rows": len(d_recipe_registry_rows),
                "d_tools_registry_rows": len(d_tools_registry_rows),
                "d_validator_rows": len(d_validator_rows),
                "d_agent_registry_rows": len(d_agent_registry_rows),
                "d_routing_rows": len(d_routing_rows),
                "d_workpaper_rows": len(d_workpaper_rows),
                "d_order_rows": len(d_order_rows),
                "d_readback_rows": len(d_readback_rows),
                "d_agent_tool_skill_rows": len(d_agent_tool_skill_rows),
                "d_authority_rows": len(d_authority_rows),
                "d_governed_asset_rows": len(d_governed_asset_rows),
                "d_source_ref_rows": len(d_source_ref_rows),
                "d_registries_ledgers_rows": len(d_registries_ledgers_rows),
                "d_domain_matrix_rows": len(d_domain_matrix_rows),
                "d_coverage_audit_rows": len(d_coverage_audit_rows),
                "d_gap_rows": len(d_gap_rows),
                "d_present": d_exists,
                "d_readonly": d_readonly,
                "state_sha": state_hash,
                "formula_errors": formula_errors,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
