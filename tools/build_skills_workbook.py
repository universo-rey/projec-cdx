from __future__ import annotations

import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

WORKSPACE_ROOT = Path(__file__).resolve().parents[1]
INVENTORY_ROOT = WORKSPACE_ROOT / "inventarios"
CSV_PATH = INVENTORY_ROOT / "SKILLS_UNIFIED_TABLE.csv"
XLSX_PATH = INVENTORY_ROOT / "SKILLS_UNIFIED_TABLE.xlsx"


def read_rows():
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def nice_title(text: str) -> str:
    return text.replace(".codex/skills", ".codex skills").replace(
        ".agents/skills", ".agents skills"
    )


def autosize(ws, rows, columns):
    widths = {col: len(col) + 2 for col in columns}
    for row in rows:
        for col in columns:
            value = row.get(col, "")
            value = "" if value is None else str(value)
            candidate = len(value) + 2
            if candidate > widths[col]:
                widths[col] = min(candidate, 72)
    for idx, col in enumerate(columns, start=1):
        ws.column_dimensions[chr(64 + idx)].width = widths[col]


def build_workbook():
    rows = read_rows()
    if not rows:
        raise RuntimeError(f"No rows found in {CSV_PATH}")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_registry = wb.create_sheet("Registro")
    ws_lists = wb.create_sheet("Listas")

    # Theme
    navy = "17324D"
    teal = "1E6F78"
    gold = "C89B3C"
    pale = "F3F6F8"
    soft = "EAF2F6"
    border = Side(style="thin", color="C9D3DA")
    thin_border = Border(left=border, right=border, top=border, bottom=border)
    header_fill = PatternFill("solid", fgColor=navy)
    band_fill = PatternFill("solid", fgColor=teal)
    accent_fill = PatternFill("solid", fgColor=gold)
    soft_fill = PatternFill("solid", fgColor=pale)
    table_fill = PatternFill("solid", fgColor=soft)

    title_font = Font(name="Aptos", size=16, bold=True, color="FFFFFF")
    heading_font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
    label_font = Font(name="Aptos", size=10, bold=True, color="17324D")
    body_font = Font(name="Aptos", size=10, color="1F2933")
    mono_font = Font(name="Cascadia Mono", size=9, color="1F2933")

    # ---------- Resumen ----------
    ws_summary.sheet_view.showGridLines = False
    ws_summary.freeze_panes = "A6"
    ws_summary.merge_cells("A1:H1")
    ws_summary["A1"] = "Tabla unica de skills"
    ws_summary["A1"].fill = header_fill
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 24

    ws_summary.merge_cells("A2:H2")
    ws_summary["A2"] = "Mapa operativo consolidado de .codex, .agents y plugin local."
    ws_summary["A2"].font = Font(name="Aptos", size=10, italic=True, color="334E68")
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    meta = [
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Fuente", str(CSV_PATH.relative_to(WORKSPACE_ROOT))),
        ("Roots", "3"),
    ]
    for i, (label, value) in enumerate(meta, start=4):
        ws_summary[f"A{i}"] = label
        ws_summary[f"A{i}"].font = label_font
        ws_summary[f"B{i}"] = value
        ws_summary[f"B{i}"].font = body_font
        ws_summary[f"A{i}"].fill = soft_fill
        ws_summary[f"B{i}"].fill = soft_fill
        ws_summary[f"A{i}"].border = thin_border
        ws_summary[f"B{i}"].border = thin_border

    counts = Counter(row["RootLabel"] for row in rows)
    summary_headers = ["Root", "Count"]
    for cell, value in zip(("D4", "E4"), summary_headers):
        ws_summary[cell] = value
        ws_summary[cell].fill = band_fill
        ws_summary[cell].font = heading_font
        ws_summary[cell].alignment = Alignment(horizontal="center")
        ws_summary[cell].border = thin_border

    for idx, root in enumerate(sorted(counts), start=5):
        ws_summary[f"D{idx}"] = root
        ws_summary[f"E{idx}"] = counts[root]
        for cell in (f"D{idx}", f"E{idx}"):
            ws_summary[cell].border = thin_border
            ws_summary[cell].font = body_font
            ws_summary[cell].fill = table_fill if idx % 2 else soft_fill

    families = defaultdict(int)
    for row in rows:
        families[row["Family"]] += 1

    ws_summary["G4"] = "Family"
    ws_summary["H4"] = "Count"
    for cell in ("G4", "H4"):
        ws_summary[cell].fill = band_fill
        ws_summary[cell].font = heading_font
        ws_summary[cell].alignment = Alignment(horizontal="center")
        ws_summary[cell].border = thin_border
    for idx, family in enumerate(sorted(families), start=5):
        ws_summary[f"G{idx}"] = family
        ws_summary[f"H{idx}"] = families[family]
        for cell in (f"G{idx}", f"H{idx}"):
            ws_summary[cell].border = thin_border
            ws_summary[cell].font = body_font
            ws_summary[cell].fill = table_fill if idx % 2 else soft_fill

    ws_summary["A8"] = "Criterio"
    ws_summary["A8"].fill = accent_fill
    ws_summary["A8"].font = heading_font
    ws_summary["A8"].border = thin_border
    criteria = [
        "Canonical = nombre visible del SKILL.md o del manifesto del plugin.",
        "Alias = carpeta o nombre alternativo cuando existe drift.",
        "RootLabel = raiz efectiva donde vive la capacidad.",
        "Kind = skill o plugin.",
    ]
    for i, text in enumerate(criteria, start=9):
        ws_summary[f"A{i}"] = f"- {text}"
        ws_summary[f"A{i}"].font = body_font
    ws_summary.column_dimensions["A"].width = 48
    ws_summary.column_dimensions["B"].width = 20
    ws_summary.column_dimensions["D"].width = 38
    ws_summary.column_dimensions["E"].width = 12
    ws_summary.column_dimensions["G"].width = 38
    ws_summary.column_dimensions["H"].width = 12

    # ---------- Registro ----------
    ws_registry.sheet_view.showGridLines = False
    headers = [
        "RootLabel",
        "Kind",
        "Family",
        "Canonical",
        "Alias",
        "Purpose",
        "SourcePath",
        "Notes",
    ]
    ws_registry.append(headers)
    for cell in ws_registry[1]:
        cell.fill = header_fill
        cell.font = heading_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in rows:
        ws_registry.append([row.get(h, "") for h in headers])

    ws_registry.freeze_panes = "A2"
    ws_registry.auto_filter.ref = f"A1:H{ws_registry.max_row}"
    for row in ws_registry.iter_rows(min_row=2, max_row=ws_registry.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.font = mono_font if cell.column in (1, 2, 3, 4, 5, 8) else body_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx in range(2, ws_registry.max_row + 1):
        fill = soft_fill if idx % 2 == 0 else table_fill
        for col in range(1, 9):
            ws_registry.cell(idx, col).fill = fill
    ws_registry.column_dimensions["A"].width = 18
    ws_registry.column_dimensions["B"].width = 12
    ws_registry.column_dimensions["C"].width = 36
    ws_registry.column_dimensions["D"].width = 34
    ws_registry.column_dimensions["E"].width = 26
    ws_registry.column_dimensions["F"].width = 80
    ws_registry.column_dimensions["G"].width = 54
    ws_registry.column_dimensions["H"].width = 36

    table = Table(displayName="SkillsUnifiedTable", ref=f"A1:H{ws_registry.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_registry.add_table(table)

    # Data validation
    ws_lists.sheet_view.showGridLines = False
    root_labels = sorted(counts)
    kinds = sorted(set(row["Kind"] for row in rows))
    families_sorted = sorted(set(row["Family"] for row in rows))
    ws_lists["A1"] = "Roots"
    ws_lists["B1"] = "Kinds"
    ws_lists["C1"] = "Families"
    for cell in ("A1", "B1", "C1"):
        ws_lists[cell].fill = header_fill
        ws_lists[cell].font = heading_font
        ws_lists[cell].border = thin_border
    for idx, value in enumerate(root_labels, start=2):
        ws_lists[f"A{idx}"] = value
    for idx, value in enumerate(kinds, start=2):
        ws_lists[f"B{idx}"] = value
    for idx, value in enumerate(families_sorted, start=2):
        ws_lists[f"C{idx}"] = value
    ws_lists.column_dimensions["A"].width = 24
    ws_lists.column_dimensions["B"].width = 16
    ws_lists.column_dimensions["C"].width = 40

    # Named ranges via direct formulas on the list sheet
    root_range = f"=Listas!$A$2:$A${len(root_labels)+1}"
    kind_range = f"=Listas!$B$2:$B${len(kinds)+1}"
    family_range = f"=Listas!$C$2:$C${len(families_sorted)+1}"

    root_dv = DataValidation(type="list", formula1=root_range, allow_blank=True)
    kind_dv = DataValidation(type="list", formula1=kind_range, allow_blank=True)
    family_dv = DataValidation(type="list", formula1=family_range, allow_blank=True)
    ws_registry.add_data_validation(root_dv)
    ws_registry.add_data_validation(kind_dv)
    ws_registry.add_data_validation(family_dv)
    root_dv.add(f"A2:A{ws_registry.max_row}")
    kind_dv.add(f"B2:B{ws_registry.max_row}")
    family_dv.add(f"C2:C{ws_registry.max_row}")

    # Apply a thin border to list entries
    for sheet in (ws_lists,):
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
                    cell.font = body_font

    wb.properties.creator = "Codex"
    wb.properties.title = "Tabla unica de skills"
    wb.properties.subject = "Inventario consolidado de skills y plugin local"
    wb.properties.description = "Workbook generado desde SKILLS_UNIFIED_TABLE.csv"

    wb.save(XLSX_PATH)


if __name__ == "__main__":
    build_workbook()
    print(f"Wrote: {XLSX_PATH}")
