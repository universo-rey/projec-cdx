from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(r"C:\Users\enzo1\.codex")
WORKSPACE = Path(__file__).parents[1]
XLSX_PATH = WORKSPACE / "CODEX_ROOT_INVENTORY.xlsx"
CSV_PATH = WORKSPACE / "CODEX_ROOT_INVENTORY.csv"
JSON_PATH = WORKSPACE / "CODEX_ROOT_MOVE_PLAN.json"

TOP_LEVEL_KEEP = {
    "README.md",
    "README.reference.md",
    "AGENTS.md",
    "AGENTS.reference.md",
    "MAPA_MAESTRO.md",
    "CARTOGRAFIA_COMPLETA.md",
    "CODEX_ATOMIC_RUNTIME_MACHINE_V2_1.md",
    "config.toml",
    "tge.config.toml",
    "version.json",
    "installation_id",
    ".personality_migration",
}

DIRECTORY_KEEP = {
    ".sandbox",
    ".sandbox-bin",
    ".sandbox-secrets",
    ".tmp",
    "ambient-suggestions",
    "app-server-local",
    "archived_sessions",
    "attachments",
    "automations",
    "backups",
    "browser",
    "cache",
    "computer-use",
    "computer-use-turn-ended",
    "environment",
    "generated_images",
    "log",
    "matrices",
    "memories",
    "memories_extensions",
    "node_repl",
    "packages",
    "pets",
    "plugins",
    "private",
    "process_manager",
    "profiles",
    "readbacks",
    "rules",
    "secrets",
    "sessions",
    "skills",
    "skills_archived",
    "skills_backups",
    "sqlite",
    "tmp",
    "tmp-appserver-schema",
    "vendor_imports",
    "workpapers",
    "worktree-archives",
    "worktrees",
    "worktrees-disabled",
}


def classify(entry: Path) -> dict[str, str]:
    name = entry.name
    is_dir = entry.is_dir()
    size = str(entry.stat().st_size) if entry.exists() and entry.is_file() else ""
    lwt = (
        datetime.fromtimestamp(entry.stat().st_mtime).isoformat(timespec="seconds")
        if entry.exists()
        else ""
    )
    current_path = str(entry)

    if is_dir:
        if name == "backups":
            return {
                "Name": name,
                "Kind": "folder",
                "CurrentPath": current_path,
                "SurfaceRole": "historical-backup-surface",
                "Status": "KEEP",
                "SuggestedDestination": "",
                "Risk": "low",
                "Reason": "Historical and versioned backup surface; compare generations before any cleanup. Not canonical.",
                "LastWriteTime": lwt,
                "Size": "",
            }
        if name == "cache":
            return {
                "Name": name,
                "Kind": "folder",
                "CurrentPath": current_path,
                "SurfaceRole": "support-cache-surface",
                "Status": "KEEP",
                "SuggestedDestination": "",
                "Risk": "low",
                "Reason": "Support cache surface; live runtime cache stays canonical at root while runtime refreshes it.",
                "LastWriteTime": lwt,
                "Size": "",
            }
        if name == "secrets":
            return {
                "Name": name,
                "Kind": "folder",
                "CurrentPath": current_path,
                "SurfaceRole": "sensitive-secrets-surface",
                "Status": "KEEP",
                "SuggestedDestination": "",
                "Risk": "high",
                "Reason": "Canonical sensitive surface; only names and structure are inventoried here.",
                "LastWriteTime": lwt,
                "Size": "",
            }
        if name == ".sandbox-secrets":
            return {
                "Name": name,
                "Kind": "folder",
                "CurrentPath": current_path,
                "SurfaceRole": "sandbox-sensitive-surface",
                "Status": "KEEP",
                "SuggestedDestination": "",
                "Risk": "high",
                "Reason": "Sandbox sensitive surface; keep as a dedicated sensitive bucket.",
                "LastWriteTime": lwt,
                "Size": "",
            }
        if name == "private":
            return {
                "Name": name,
                "Kind": "folder",
                "CurrentPath": current_path,
                "SurfaceRole": "private-sensitive-surface",
                "Status": "KEEP",
                "SuggestedDestination": "",
                "Risk": "high",
                "Reason": "Private reserved surface; do not treat as cleanup or canon candidate.",
                "LastWriteTime": lwt,
                "Size": "",
            }
        return {
            "Name": name,
            "Kind": "folder",
            "CurrentPath": current_path,
            "SurfaceRole": "top-level-surface",
            "Status": "KEEP",
            "SuggestedDestination": "",
            "Risk": "low",
            "Reason": "Surface root already exposed by README/map structure.",
            "LastWriteTime": lwt,
            "Size": "",
        }

    if name in TOP_LEVEL_KEEP:
        role = "root-doc" if name.endswith(".md") else "runtime-core"
        if name.endswith(".md"):
            role = "visible-entry" if name in {"README.md", "AGENTS.md"} else "map-or-reference"
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": role,
            "Status": "KEEP",
            "SuggestedDestination": "",
            "Risk": "low" if name.endswith(".md") else "medium",
            "Reason": "Root-owned document or runtime anchor.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if re.match(r"^\.{2}codex-global-state\.json\.tmp-", name):
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "temp-state",
            "Status": "MOVE_SAFE",
            "SuggestedDestination": r"tmp\global-state",
            "Risk": "low",
            "Reason": "Temporary global-state snapshot; better under tmp.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if re.match(r"^\.codex-global-state\.json\.bak", name):
        destination = ROOT / "backups" / "global-state" / name
        if destination.exists():
            return {
                "Name": name,
                "Kind": "file",
                "CurrentPath": current_path,
                "SurfaceRole": "backup-state",
                "Status": "BLOCKED_REVIEW",
                "SuggestedDestination": r"backups\global-state",
                "Risk": "medium",
                "Reason": "A backup with the same name already exists in backups/global-state; compare generations before moving or overwriting.",
                "LastWriteTime": lwt,
                "Size": size,
            }
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "backup-state",
            "Status": "MOVE_SAFE",
            "SuggestedDestination": r"backups\global-state",
            "Risk": "low",
            "Reason": "Global-state backup belongs under backups/global-state.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if re.match(r"^AGENTS\.md\.bak", name):
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "backup-doc",
            "Status": "MOVE_SAFE",
            "SuggestedDestination": r"backups\agents",
            "Risk": "low",
            "Reason": "AGENTS backup belongs under backups/agents.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if re.match(r"^(config\.toml|tge\.config\.toml)\.bak", name):
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "backup-config",
            "Status": "MOVE_SAFE",
            "SuggestedDestination": r"backups\config",
            "Risk": "low",
            "Reason": "Config backups belong under backups/config.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if re.match(r"^state_5\.sqlite\.bak", name):
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "backup-sqlite",
            "Status": "MOVE_SAFE",
            "SuggestedDestination": r"backups\sqlite",
            "Risk": "low",
            "Reason": "SQLite backups belong under backups/sqlite.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if name.startswith(("pending-", "repair-", "restart-")) and name.lower().endswith(
        (".ps1", ".log")
    ):
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "maintenance-artifact",
            "Status": "MOVE_SAFE",
            "SuggestedDestination": r"workpapers\maintenance",
            "Risk": "low",
            "Reason": "Historical maintenance scripts/logs fit workpapers/maintenance.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if name.lower().startswith("sandbox") and name.lower().endswith(".log"):
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "log",
            "Status": "MOVE_SAFE",
            "SuggestedDestination": r"log\sandbox",
            "Risk": "low",
            "Reason": "Sandbox logs belong under log/sandbox.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if name == "models_cache.json":
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "live-runtime-cache",
            "Status": "KEEP",
            "SuggestedDestination": "",
            "Risk": "medium",
            "Reason": "Live Codex model cache refreshed by runtime; keep in root as the canonical working cache.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if name in {"auth.json", "cap_sid"}:
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "secret-or-credential",
            "Status": "BLOCKED_REVIEW",
            "SuggestedDestination": r"secrets",
            "Risk": "high",
            "Reason": "Sensitive/auth material needs explicit confirmation before moving.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if name.startswith("chrome-native-hosts") and name.endswith(".json"):
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "browser-host-config",
            "Status": "BLOCKED_REVIEW",
            "SuggestedDestination": r"browser",
            "Risk": "medium",
            "Reason": "May be bound to browser/native host registration; validate before moving.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    if name.endswith((".sqlite", ".sqlite-shm", ".sqlite-wal")) or name in {
        "history.jsonl",
        "session_index.jsonl",
        "transcription-history.jsonl",
        "goals_1.sqlite",
        "goals_1.sqlite-shm",
        "goals_1.sqlite-wal",
        "logs_2.sqlite",
        "logs_2.sqlite-shm",
        "logs_2.sqlite-wal",
        "memories_1.sqlite",
        "memories_1.sqlite-shm",
        "memories_1.sqlite-wal",
        "state_5.sqlite",
        "state_5.sqlite-shm",
        "state_5.sqlite-wal",
        ".codex-global-state.json",
        "version.json",
        "installation_id",
    }:
        return {
            "Name": name,
            "Kind": "file",
            "CurrentPath": current_path,
            "SurfaceRole": "runtime-state",
            "Status": "KEEP",
            "SuggestedDestination": "",
            "Risk": "medium",
            "Reason": "Live runtime/state artifact; do not move without exact dependency validation.",
            "LastWriteTime": lwt,
            "Size": size,
        }

    return {
        "Name": name,
        "Kind": "file",
        "CurrentPath": current_path,
        "SurfaceRole": "uncategorized",
        "Status": "BLOCKED_REVIEW",
        "SuggestedDestination": "",
        "Risk": "medium",
        "Reason": "Needs manual review to confirm whether it belongs in a child surface.",
        "LastWriteTime": lwt,
        "Size": size,
    }


def build_workbook(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_reg = wb.create_sheet("Registro")
    ws_lists = wb.create_sheet("Listas")

    navy = "17324D"
    teal = "1E6F78"
    pale = "F3F6F8"
    soft = "EAF2F6"
    border = Side(style="thin", color="C9D3DA")
    thin_border = Border(left=border, right=border, top=border, bottom=border)
    header_fill = PatternFill("solid", fgColor=navy)
    band_fill = PatternFill("solid", fgColor=teal)
    soft_fill = PatternFill("solid", fgColor=pale)
    table_fill = PatternFill("solid", fgColor=soft)

    title_font = Font(name="Aptos", size=16, bold=True, color="FFFFFF")
    heading_font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=10, color="1F2933")
    mono_font = Font(name="Cascadia Mono", size=9, color="1F2933")

    # Summary
    ws_summary.sheet_view.showGridLines = False
    ws_summary.freeze_panes = "A5"
    ws_summary.merge_cells("A1:H1")
    ws_summary["A1"] = "Inventario de raiz .codex"
    ws_summary["A1"].fill = header_fill
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(horizontal="center")

    ws_summary.merge_cells("A2:H2")
    ws_summary["A2"] = "Clasificacion de superficie y plan de reubicacion segura."
    ws_summary["A2"].font = Font(name="Aptos", size=10, italic=True, color="334E68")
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    meta = [
        ("Root", str(ROOT)),
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total items", str(len(rows))),
    ]
    for idx, (label, value) in enumerate(meta, start=4):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"B{idx}"] = value
        for c in (f"A{idx}", f"B{idx}"):
            ws_summary[c].fill = soft_fill
            ws_summary[c].border = thin_border
        ws_summary[f"A{idx}"].font = body_font
        ws_summary[f"B{idx}"].font = body_font

    action_counts = Counter(r["Status"] for r in rows)
    role_counts = Counter(r["SurfaceRole"] for r in rows)
    note_row = max(4 + len(action_counts), 4 + len(role_counts)) + 2

    ws_summary["D4"] = "Status"
    ws_summary["E4"] = "Count"
    for c in ("D4", "E4"):
        ws_summary[c].fill = band_fill
        ws_summary[c].font = heading_font
        ws_summary[c].border = thin_border
    for idx, key in enumerate(sorted(action_counts), start=5):
        ws_summary[f"D{idx}"] = key
        ws_summary[f"E{idx}"] = action_counts[key]
        for c in (f"D{idx}", f"E{idx}"):
            ws_summary[c].fill = table_fill if idx % 2 else soft_fill
            ws_summary[c].border = thin_border

    ws_summary["G4"] = "SurfaceRole"
    ws_summary["H4"] = "Count"
    for c in ("G4", "H4"):
        ws_summary[c].fill = band_fill
        ws_summary[c].font = heading_font
        ws_summary[c].border = thin_border
    for idx, key in enumerate(sorted(role_counts), start=5):
        ws_summary[f"G{idx}"] = key
        ws_summary[f"H{idx}"] = role_counts[key]
        for c in (f"G{idx}", f"H{idx}"):
            ws_summary[c].fill = table_fill if idx % 2 else soft_fill
            ws_summary[c].border = thin_border

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 48
    ws_summary.column_dimensions["D"].width = 18
    ws_summary.column_dimensions["E"].width = 10
    ws_summary.column_dimensions["G"].width = 28
    ws_summary.column_dimensions["H"].width = 10

    ws_summary.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    ws_summary.cell(note_row, 1).value = "Operating note"
    ws_summary.cell(note_row, 1).fill = band_fill
    ws_summary.cell(note_row, 1).font = heading_font
    ws_summary.cell(note_row, 1).border = thin_border
    ws_summary.cell(note_row, 1).alignment = Alignment(horizontal="left")

    note_text = (
        "Only a second exact copy in cache/ or backups/ counts as a live duplicate. "
        "backups/ is historical and versioned, not canonical. models_cache.json stays "
        "in the root as the live runtime cache while Codex refreshes it. auth.json "
        "and cap_sid remain sensitive and are not candidates for cleanup without an "
        "explicit review order."
    )
    ws_summary.merge_cells(
        start_row=note_row + 1, start_column=1, end_row=note_row + 3, end_column=8
    )
    ws_summary.cell(note_row + 1, 1).value = note_text
    ws_summary.cell(note_row + 1, 1).fill = soft_fill
    ws_summary.cell(note_row + 1, 1).font = body_font
    ws_summary.cell(note_row + 1, 1).border = thin_border
    ws_summary.cell(note_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")

    # Registry
    headers = [
        "Name",
        "Kind",
        "CurrentPath",
        "SurfaceRole",
        "Status",
        "SuggestedDestination",
        "Risk",
        "Reason",
        "LastWriteTime",
        "Size",
    ]
    ws_reg.append(headers)
    for cell in ws_reg[1]:
        cell.fill = header_fill
        cell.font = heading_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws_reg.append([row.get(h, "") for h in headers])

    ws_reg.freeze_panes = "A2"
    ws_reg.auto_filter.ref = f"A1:J{ws_reg.max_row}"
    for ridx in range(2, ws_reg.max_row + 1):
        fill = soft_fill if ridx % 2 == 0 else table_fill
        for cidx in range(1, len(headers) + 1):
            cell = ws_reg.cell(ridx, cidx)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = mono_font if cidx in (1, 2, 4, 5, 7, 10) else body_font

    widths = {
        "A": 36,
        "B": 10,
        "C": 56,
        "D": 24,
        "E": 16,
        "F": 32,
        "G": 14,
        "H": 60,
        "I": 20,
        "J": 16,
    }
    for col, width in widths.items():
        ws_reg.column_dimensions[col].width = width

    table = Table(displayName="CodexRootInventory", ref=f"A1:J{ws_reg.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_reg.add_table(table)

    # Lists
    ws_lists.sheet_view.showGridLines = False
    status_values = ["KEEP", "MOVE_SAFE", "BLOCKED_REVIEW"]
    roles = sorted(set(role_counts))
    ws_lists["A1"] = "Status"
    ws_lists["B1"] = "SurfaceRole"
    for c in ("A1", "B1"):
        ws_lists[c].fill = band_fill
        ws_lists[c].font = heading_font
        ws_lists[c].border = thin_border
    for idx, value in enumerate(status_values, start=2):
        ws_lists[f"A{idx}"] = value
    for idx, value in enumerate(roles, start=2):
        ws_lists[f"B{idx}"] = value
    ws_lists.column_dimensions["A"].width = 18
    ws_lists.column_dimensions["B"].width = 28

    # Save workbook
    wb.properties.creator = "Codex"
    wb.properties.title = "Inventario de raiz .codex"
    wb.properties.subject = "Clasificacion de la raiz de .codex"
    wb.properties.description = (
        "Inventario, plan de movimiento y clasificacion de la raiz de .codex"
    )
    wb.save(XLSX_PATH)


def main() -> None:
    entries = sorted(ROOT.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
    rows = [classify(entry) for entry in entries]

    with CSV_PATH.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "Name",
                "Kind",
                "CurrentPath",
                "SurfaceRole",
                "Status",
                "SuggestedDestination",
                "Risk",
                "Reason",
                "LastWriteTime",
                "Size",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    plan = {
        "root": str(ROOT),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "operating_note": (
            "Only a second exact copy in cache/ or backups/ counts as a live duplicate; "
            "backups/ is historical and versioned; models_cache.json stays canonical in "
            "root as the live runtime cache; auth.json and cap_sid remain sensitive and "
            "blocked from cleanup without explicit review."
        ),
        "move_safe": [
            {
                "name": row["Name"],
                "source": row["CurrentPath"],
                "destination": row["SuggestedDestination"],
                "surface_role": row["SurfaceRole"],
                "reason": row["Reason"],
            }
            for row in rows
            if row["Status"] == "MOVE_SAFE"
        ],
        "blocked_review": [
            {
                "name": row["Name"],
                "source": row["CurrentPath"],
                "destination": row["SuggestedDestination"],
                "surface_role": row["SurfaceRole"],
                "reason": row["Reason"],
            }
            for row in rows
            if row["Status"] == "BLOCKED_REVIEW"
        ],
    }

    with JSON_PATH.open("w", encoding="utf-8") as fh:
        json.dump(plan, fh, indent=2)

    build_workbook(rows)
    print(f"Wrote: {CSV_PATH}")
    print(f"Wrote: {JSON_PATH}")
    print(f"Wrote: {XLSX_PATH}")


if __name__ == "__main__":
    main()
