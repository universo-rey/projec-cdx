from __future__ import annotations

import argparse
import csv
import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


HOME = Path(r"C:\Users\enzo1")
PROJEC = HOME / "PROJEC CDX"
CODEX = HOME / ".codex"
AGENTS = HOME / ".agents"
GITHUB = HOME / "Documents" / "GitHub"
CABINA = GITHUB / "cabina-universal-d"

OUTDIR = PROJEC / "outputs" / "universe_relationship_audit_20260614"
OUTDIR.mkdir(parents=True, exist_ok=True)

CSV_OUT = OUTDIR / "UNIVERSE_RELATIONSHIP_AUDIT.csv"
MD_OUT = OUTDIR / "UNIVERSE_RELATIONSHIP_AUDIT.md"
XLSX_OUT = OUTDIR / "UNIVERSE_RELATIONSHIP_AUDIT.xlsx"

CABINA_OUTDIR = PROJEC / "outputs" / "cabina_relationship_audit_20260614"
CABINA_OUTDIR.mkdir(parents=True, exist_ok=True)

CABINA_DEEP_CSV = CABINA_OUTDIR / "CABINA_RELATIONSHIP_AUDIT_DEEP.csv"
CABINA_DEEP_MD = CABINA_OUTDIR / "CABINA_RELATIONSHIP_AUDIT_DEEP.md"
CABINA_DEEP_XLSX = CABINA_OUTDIR / "CABINA_RELATIONSHIP_AUDIT_DEEP.xlsx"

CABINA_EXEC_CSV = CABINA_OUTDIR / "CABINA_RELATIONSHIP_AUDIT_EXECUTIVE.csv"
CABINA_EXEC_MD = CABINA_OUTDIR / "CABINA_RELATIONSHIP_AUDIT_EXECUTIVE.md"
CABINA_EXEC_XLSX = CABINA_OUTDIR / "CABINA_RELATIONSHIP_AUDIT_EXECUTIVE.xlsx"

ROOT_INVENTORY_CSV = PROJEC / "inventarios" / "CODEX_ROOT_INVENTORY.csv"
SKILLS_TABLE_CSV = PROJEC / "inventarios" / "SKILLS_UNIFIED_TABLE.csv"

STATE_ORDER = ["activo", "referencia", "derivado", "interno", "listado"]
STATE_COLORS = {
    "activo": "D8F3DC",
    "referencia": "DDEBFF",
    "derivado": "FFF3CD",
    "interno": "E2E3E5",
    "listado": "EADCF8",
}

CABINA_SKIP_DIRS = {
    ".git",
    ".venv",
    "__pycache__",
    "node_modules",
    "tmp",
    "cache",
    ".cache",
}

CABINA_PRIORITY_FILES = [
    CABINA / "README.md",
    CABINA / "AGENTS.md",
    CABINA / "MANIFEST.yaml",
    CABINA / "MAPA_HUMANO.md",
    CABINA / "README.reference.md",
    CABINA / "02_AUTHORITY_CANON" / "CURRENT_STATE.md",
    CABINA / "00_CONTROL_PLANE_INGRESS" / "ROUTING.json",
    CABINA / "01_GOVERNANCE_REGISTRY" / "README.md",
    CABINA / "01_GOVERNANCE_REGISTRY" / "REPOSITORIES.csv",
    CABINA / "01_GOVERNANCE_REGISTRY" / "GITHUB_BASE_WORK_MATRIX.csv",
    CABINA / "01_GOVERNANCE_REGISTRY" / "OWNER_MATRIX.csv",
    CABINA / "01_GOVERNANCE_REGISTRY" / "MIGRATION_MAP_D.csv",
    CABINA / "02_AUTHORITY_CANON" / "README.md",
    CABINA / "02_AUTHORITY_CANON" / "ACTIVE_GOVERNED_EXECUTION_BY_DEFAULT_POLICY_20260603.md",
    CABINA / ".agents" / "codex" / "README.md",
    CABINA / ".agents" / "codex" / "agents.json",
    CABINA / ".agents" / "codex" / "routing.json",
    CABINA / ".agents" / "codex" / "matrices" / "CAPABILITY_SOURCE_INVENTORY.csv",
    CABINA / ".agents" / "codex" / "matrices" / "CAPABILITY_IMPORT_DECISION_MATRIX.csv",
    CABINA / ".agents" / "codex" / "matrices" / "PARALLEL_OPERATION_CRITERIA_MATRIX.csv",
    CABINA / ".agents" / "codex" / "matrices" / "ORDER_PREPARATION_ASSIGNMENT_MATRIX.csv",
    CABINA / ".agents" / "codex" / "matrices" / "SKILL_REFERENCE_SOURCE_MATRIX.csv",
    CABINA / ".agents" / "codex" / "matrices" / "REPO_RUNTIME_ALIGNMENT_MATRIX.csv",
    CABINA / ".agents" / "codex" / "matrices" / "ACTIVE_EXECUTION_CAPABILITY_MATRIX_20260603.csv",
    CABINA / "governance" / "canon" / "ACTIVE_GOVERNED_EXECUTION_BY_DEFAULT_POLICY_20260603.md",
    CABINA / "governance" / "canon" / "ACTIVE_GOVERNED_EXECUTION_BY_DEFAULT_POLICY_20260603.csv",
    CABINA / "readbacks" / "README.md",
    CABINA / "validation" / "README.md",
    CABINA / "matrices" / "README.md",
    CABINA / "recipes" / "README.md",
    CABINA / "scripts" / "README.md",
    CABINA / "docs" / "README.md",
    CABINA / "docs" / "powerautomate" / "WORK_QUEUE_FLOW_OPERATING_DESIGN.md",
    CABINA / "powerplatform" / "workqueues" / "workqueue.manifest.yml",
    CABINA / "local-agent-bridge" / "README.md",
    CABINA / "80_REFERENCIAS_TECNICAS" / "README.md",
    CABINA / "08_READBACKS" / "README.md",
]

EXECUTIVE_ROW_LIMIT = 120


def norm(text: object) -> str:
    if text is None:
        return ""
    return str(text).strip()


def rel(path: Path) -> str:
    return str(path).replace("\\", "/")


def classify_file_type(path: Path) -> str:
    name = path.name.lower()
    suffix = path.suffix.lower()
    if name in {"readme.md", "readme.reference.md", "reference.md"}:
        return "reference"
    if name in {"mapa_maestro.md", "mapa.md"}:
        return "map"
    if name in {"agents.md", "agents.reference.md"}:
        return "governance"
    if suffix == ".xlsx":
        return "workbook"
    if suffix == ".csv":
        return "inventory"
    if suffix == ".json":
        return "data"
    if suffix in {".ps1", ".py", ".mjs", ".js"}:
        return "tool"
    if suffix == ".zip":
        return "package"
    if suffix == ".md":
        return "doc"
    return "file"


def classify_state(path: Path, kind: str, universe: str) -> str:
    p = rel(path).lower()
    name = path.name.lower()
    if universe == ".codex":
        if kind in {"reference", "map", "governance"}:
            return "referencia"
        if "workpapers" in p or "skills" in p or "matrices" in p or "readbacks" in p or "rules" in p or "sessions" in p or "plugins" in p or "worktrees" in p or "environment" in p or "log" in p:
            return "activo"
        if "tmp" in p or "cache" in p or "backup" in p or "archived" in p or "sqlite" in p or "secrets" in p or "private" in p or "node_modules" in p:
            return "interno"
        return "activo"
    if universe == "PROJEC CDX":
        if "outputs" in p or "workbook" in name or "zip" in name:
            return "derivado"
        if kind in {"reference", "map", "governance"}:
            return "referencia"
        if "tools" in p or "atomic" in p:
            return "activo"
        return "activo"
    if universe in {"Projects", "CodexLocal", "Documents\\GitHub"}:
        if name in {"readme.md", "mapa_maestro.md", "agents.md"} or kind in {"reference", "map", "governance"}:
            return "referencia"
        if "archive" in p or "backup" in p or "tmp" in p:
            return "interno"
        return "activo"
    if universe in {".agents", "cabina-universal-d"}:
        if "archive" in p or "backup" in p or "tmp" in p:
            return "interno"
        if kind in {"reference", "map", "governance"}:
            return "referencia"
        return "activo"
    if "output" in p or "outputs" in p:
        return "derivado"
    return "listado"


def state_for_status(status: str) -> str:
    value = (status or "").strip().upper()
    if value == "KEEP":
        return "activo"
    if value == "MOVE_SAFE":
        return "interno"
    if value in {"ACTIVO", "REFERENCIA", "DERIVADO", "INTERNO", "LISTADO"}:
        return value.lower()
    return "listado"


def destination_for_path(path: Path, root_readme: str):
    current = path if path.is_dir() else path.parent
    while True:
        for name in ("README.md", "MAPA_MAESTRO.md", "MAPA.md", "README.reference.md", "reference.md", "AGENTS.md", "AGENTS.reference.md"):
            candidate = current / name
            if candidate.exists():
                return rel(candidate)
        if current == current.parent:
            break
        current = current.parent
    return root_readme


def add_row(rows, universe, origin, typ, destination, dependency, state, source):
    rows.append(
        {
            "Universo": universe,
            "Origen": origin,
            "Tipo": typ,
            "Destino": destination,
            "Dependencia": dependency,
            "Estado": state,
            "Fuente": source,
        }
    )


def add_tree_rows(
    rows,
    universe,
    root_path: Path,
    root_readme: str,
    dependency: str,
    source: str,
    max_depth: int = 1,
    max_files_per_dir: int = 8,
    skip_dir_names: set[str] | None = None,
):
    if not root_path.exists():
        return

    base_depth = len(root_path.parts)
    skip_dir_names = skip_dir_names or set()
    for current, dirs, files in os.walk(root_path):
        current_path = Path(current)
        depth = len(current_path.parts) - base_depth
        dirs[:] = [d for d in dirs if d not in skip_dir_names]
        if depth > max_depth:
            dirs[:] = []
            continue

        typ = "root" if depth == 0 else "folder"
        add_row(
            rows,
            universe,
            rel(current_path),
            typ,
            destination_for_path(current_path, root_readme),
            dependency,
            classify_state(current_path, typ, universe),
            source,
        )

        for filename in sorted(files)[:max_files_per_dir]:
            file_path = current_path / filename
            typ = classify_file_type(file_path)
            add_row(
                rows,
                universe,
                rel(file_path),
                typ,
                destination_for_path(file_path, root_readme),
                dependency,
                classify_state(file_path, typ, universe),
                source,
            )

        if depth >= max_depth:
            dirs[:] = []


def add_priority_files(rows, universe, root_readme: str, dependency: str, source: str):
    for path in CABINA_PRIORITY_FILES:
        if not path.exists():
            continue
        kind = classify_file_type(path)
        add_row(
            rows,
            universe,
            rel(path),
            kind,
            destination_for_path(path, root_readme),
            dependency,
            classify_state(path, kind, universe),
            source,
        )


def add_codex_inventory(rows):
    if not ROOT_INVENTORY_CSV.exists():
        return
    with ROOT_INVENTORY_CSV.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            current = norm(row.get("CurrentPath"))
            name = norm(row.get("Name"))
            kind = norm(row.get("Kind")) or "file"
            status = norm(row.get("Status"))
            suggested = norm(row.get("SuggestedDestination"))
            universe = ".codex"
            typ = "folder" if kind == "folder" else "file"
            if name.endswith(".md"):
                typ = "reference" if "README" in name or "reference" in name.lower() else "doc"
            destination = suggested or rel(CODEX / "README.reference.md")
            if not suggested and typ == "folder":
                destination = rel(CODEX / "README.md")
            dependency = rel(CODEX / "README.md")
            state = state_for_status(status)
            source = "CODEX_ROOT_INVENTORY.csv"
            add_row(rows, universe, current or rel(CODEX / name), typ, destination, dependency, state, source)


def add_skills(rows):
    if not SKILLS_TABLE_CSV.exists():
        return
    with SKILLS_TABLE_CSV.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            root_label = norm(row.get("RootLabel"))
            kind = norm(row.get("Kind")) or "skill"
            canonical = norm(row.get("Canonical"))
            alias = norm(row.get("Alias"))
            source_path = norm(row.get("SourcePath"))
            family = norm(row.get("Family"))
            universe = ".codex" if root_label.startswith(".codex") else ".agents"
            origin = source_path or canonical or alias
            destination = rel(CODEX / "skills" / "README.md") if root_label.startswith(".codex") else rel(AGENTS / "README.md")
            if "plugins" in root_label:
                destination = rel(AGENTS / "README.md")
            dependency = "STACK.md / AGENTS_INDEX.csv"
            state = "activo"
            typ = kind
            source = "SKILLS_UNIFIED_TABLE.csv"
            add_row(rows, universe, origin, typ, destination, dependency, state, source)


def scan_github_roots(rows):
    if not GITHUB.exists():
        return
    for repo in sorted([p for p in GITHUB.iterdir() if p.is_dir()]):
        universe = "Documents\\GitHub"
        root_readme = rel(repo / "README.md")
        dependency = rel(PROJEC / "README.md")
        state = "activo" if (repo / ".git").exists() or (repo / "README.md").exists() else "listado"
        add_row(rows, universe, rel(repo), "root", root_readme, dependency, state, "scan:Documents\\GitHub")
        add_tree_rows(rows, universe, repo, root_readme, dependency, "scan:Documents\\GitHub", max_depth=2, max_files_per_dir=5)


def build_rows():
    rows = []

    # Core .codex inventory and relations
    add_codex_inventory(rows)

    # Skills and plugins registry
    add_skills(rows)

    # Root surfaces
    root_specs = [
        (".agents", AGENTS, rel(AGENTS / "README.md"), rel(HOME / "README.md"), ".agents/README.md", 2, 6),
        ("PROJEC CDX", PROJEC, rel(PROJEC / "README.md"), rel(CODEX / "README.reference.md"), "PROJEC CDX/README.md", 3, 8),
        ("Projects", HOME / "Projects", rel(HOME / "Projects" / "README.md"), rel(CODEX / "README.md"), "Projects/README.md", 2, 6),
        ("CodexLocal", HOME / "CodexLocal", rel(HOME / "CodexLocal" / "README.md"), rel(CODEX / "README.md"), "CodexLocal/README.md", 2, 6),
        ("Sandboxes", HOME / "Sandboxes", rel(HOME / "Sandboxes" / "README.md"), rel(HOME / "Projects" / "README.md"), "Sandboxes/README.md", 2, 5),
        ("Intake", HOME / "Intake", rel(HOME / "Intake" / "README.md"), rel(HOME / "Projects" / "README.md"), "Intake/README.md", 2, 5),
        ("cabina-universal-d", HOME / "Documents" / "GitHub" / "cabina-universal-d", rel(HOME / "Documents" / "GitHub" / "cabina-universal-d" / "README.md"), rel(PROJEC / "README.md"), "cabina-universal-d/README.md", 3, 6),
    ]

    for universe, root_path, root_readme, dependency, source, max_depth, max_files in root_specs:
        if not root_path.exists():
            continue
        add_row(rows, universe, rel(root_path), "root", root_readme, dependency, "activo", source)
        add_tree_rows(rows, universe, root_path, root_readme, dependency, source, max_depth=max_depth, max_files_per_dir=max_files)

    # Extra nested surfaces in PROJEC CDX.
    proj_outputs = PROJEC / "outputs"
    if proj_outputs.exists():
        add_tree_rows(rows, "PROJEC CDX", proj_outputs, rel(proj_outputs / "README.md"), rel(PROJEC / "README.md"), "scan:PROJEC CDX/outputs", max_depth=2, max_files_per_dir=8)
    proj_docs = PROJEC / "docs" / "superpowers"
    if proj_docs.exists():
        add_tree_rows(rows, "PROJEC CDX", proj_docs, rel(proj_docs / "README.md"), rel(PROJEC / "README.md"), "scan:PROJEC CDX/docs/superpowers", max_depth=2, max_files_per_dir=6)

    # GitHub repo roots.
    scan_github_roots(rows)

    # Home-level listed roots that Projects README calls out.
    listed_roots = [
        HOME / "Docs Authoring",
        HOME / "ESCRIBANIA BITSCH",
        HOME / "Modo On",
    ]
    for root_path in listed_roots:
        if root_path.exists():
            add_row(rows, "home-listed", rel(root_path), "root", rel(root_path / "README.md"), rel(HOME / "Projects" / "README.md"), "listado", "Projects/README.md")
            add_tree_rows(rows, "home-listed", root_path, rel(root_path / "README.md"), rel(HOME / "Projects" / "README.md"), "scan:home-listed", max_depth=2, max_files_per_dir=5)

    # Dedupe while preserving order.
    seen = set()
    deduped = []
    for row in rows:
        key = tuple(row[k] for k in ("Universo", "Origen", "Tipo", "Destino", "Dependencia", "Estado", "Fuente"))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped


def cabina_depth(path_value: str) -> int:
    try:
        return len(Path(path_value).resolve().relative_to(CABINA.resolve()).parts)
    except Exception:
        return 999


def executive_score(row: dict) -> int:
    origin = norm(row.get("Origen")).lower()
    destination = norm(row.get("Destino")).lower()
    dependency = norm(row.get("Dependencia")).lower()
    source = norm(row.get("Fuente")).lower()
    typ = norm(row.get("Tipo")).lower()
    state = norm(row.get("Estado")).lower()
    text = " ".join([origin, destination, dependency, source, typ, state])

    if row.get("Universo") != "cabina-universal-d":
        return -999

    score = 0
    depth = cabina_depth(origin)

    if typ == "root":
        score += 100
    if depth == 1:
        score += 35
    elif depth == 2:
        score += 18
    elif depth == 3:
        score += 8

    if typ in {"reference", "map", "governance"}:
        score += 28
    elif typ in {"folder", "doc", "file"}:
        score += 6

    if state == "activo":
        score += 14
    elif state == "referencia":
        score += 12
    elif state == "interno":
        score += 0
    elif state == "derivado":
        score -= 8
    elif state == "listado":
        score -= 15

    keywords = [
        "manifest.yaml",
        "mapa_humano.md",
        "current_state.md",
        "routing.json",
        "readme.md",
        "readme.reference.md",
        "agents.md",
        "gates",
        "governance",
        "canon",
        "matrix",
        "matrices",
        "readback",
        "validator",
        "recipe",
        "skill",
        "tools",
        "atomic",
        "workqueue",
        "powerplatform",
        "local-agent-bridge",
        "00_control_plane_ingress",
        "01_governance_registry",
        "02_authority_canon",
        "03_corte_ejecutora_del_rey",
        "80_referencias_tecnicas",
        "08_readbacks",
    ]
    for keyword in keywords:
        if keyword in text:
            score += 6

    if "outputs" in origin or "tmp" in origin or "cache" in origin:
        score -= 12

    return score


def is_executive_row(row: dict) -> bool:
    score = executive_score(row)
    if score < 34:
        return False
    if norm(row.get("Estado")).lower() not in {"activo", "referencia"}:
        return False
    if norm(row.get("Tipo")).lower() in {"file", "doc"} and cabina_depth(row["Origen"]) > 3:
        return False
    return True


def build_cabina_rows():
    rows = []
    if not CABINA.exists():
        return rows

    root_readme = rel(CABINA / "README.md")
    dependency = rel(CABINA / "README.md")
    add_row(rows, "cabina-universal-d", rel(CABINA), "root", root_readme, dependency, "activo", "cabina-universal-d/README.md")
    add_priority_files(rows, "cabina-universal-d", root_readme, dependency, "cabina-universal-d/priority")
    add_tree_rows(
        rows,
        "cabina-universal-d",
        CABINA,
        root_readme,
        dependency,
        "cabina-universal-d/recursive",
        max_depth=4,
        max_files_per_dir=10,
        skip_dir_names=CABINA_SKIP_DIRS,
    )

    seen = set()
    deduped = []
    for row in rows:
        key = tuple(row[k] for k in ("Universo", "Origen", "Tipo", "Destino", "Dependencia", "Estado", "Fuente"))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def build_cabina_executive_rows():
    rows = build_cabina_rows()
    executive = [row for row in rows if is_executive_row(row)]
    executive.sort(key=lambda row: (
        -executive_score(row),
        cabina_depth(row["Origen"]),
        0 if row["Tipo"] == "root" else 1,
        row["Tipo"],
        row["Origen"],
    ))
    return executive[:EXECUTIVE_ROW_LIMIT]


def write_csv(rows, out_path: Path):
    headers = ["Universo", "Origen", "Tipo", "Destino", "Dependencia", "Estado", "Fuente"]
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_md(rows, out_path: Path, title: str, intro: str):
    headers = ["Universo", "Origen", "Tipo", "Destino", "Dependencia", "Estado", "Fuente"]
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        fh.write(f"# {title}\n\n")
        fh.write(f"{intro}\n\n")
        fh.write("| " + " | ".join(headers) + " |\n")
        fh.write("|" + "|".join(["---"] * len(headers)) + "|\n")
        for row in rows:
            fh.write("| " + " | ".join(row.get(h, "").replace("|", "\\|") for h in headers) + " |\n")


def autosize(ws, headers, rows):
    widths = {h: len(h) + 2 for h in headers}
    for row in rows:
        for h in headers:
            widths[h] = max(widths[h], min(len(norm(row.get(h))) + 2, 80))
    for idx, h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + idx)].width = widths[h]


def write_xlsx(rows, out_path: Path, title: str, subtitle: str):
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_summary = wb.create_sheet("Resumen")
    ws_table = wb.create_sheet("Relaciones")
    ws_lists = wb.create_sheet("Listas")

    navy = "17324D"
    teal = "1E6F78"
    gold = "C89B3C"
    pale = "F3F6F8"
    soft = "EAF2F6"
    border = Side(style="thin", color="C9D3DA")
    thin_border = Border(left=border, right=border, top=border, bottom=border)
    header_fill = PatternFill("solid", fgColor=navy)
    band_fill = PatternFill("solid", fgColor=teal)
    accent_fill = PatternFill("solid", fgColor=gold)
    soft_fill = PatternFill("solid", fgColor=pale)
    table_fill = PatternFill("solid", fgColor=soft)

    title_font = Font(name="Aptos", size=16, bold=True, color="FFFFFF")
    heading_font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=10, color="1F2933")
    mono_font = Font(name="Cascadia Mono", size=9, color="1F2933")
    card_title_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    card_value_font = Font(name="Aptos", size=18, bold=True, color="FFFFFF")

    headers = ["Universo", "Origen", "Tipo", "Destino", "Dependencia", "Estado", "Fuente"]
    counts_universe = Counter(row["Universo"] for row in rows)
    counts_state = Counter(row["Estado"] for row in rows)
    counts_type = Counter(row["Tipo"] for row in rows)

    # Dashboard
    ws_dash.sheet_view.showGridLines = False
    ws_dash.freeze_panes = "A4"
    ws_dash.merge_cells("A1:L1")
    ws_dash["A1"] = title
    ws_dash["A1"].fill = header_fill
    ws_dash["A1"].font = title_font
    ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 24
    ws_dash.merge_cells("A2:L2")
    ws_dash["A2"] = subtitle
    ws_dash["A2"].alignment = Alignment(horizontal="center")

    cards = [
        ("A4:C5", "Total relaciones", len(rows), header_fill),
        ("D4:F5", "Activas", counts_state.get("activo", 0), PatternFill("solid", fgColor=STATE_COLORS["activo"])),
        ("G4:I5", "Referencias", counts_state.get("referencia", 0), PatternFill("solid", fgColor=STATE_COLORS["referencia"])),
        ("J4:L5", "Derivadas", counts_state.get("derivado", 0), PatternFill("solid", fgColor=STATE_COLORS["derivado"])),
        ("A7:C8", "Internas", counts_state.get("interno", 0), PatternFill("solid", fgColor=STATE_COLORS["interno"])),
        ("D7:F8", "Listadas", counts_state.get("listado", 0), PatternFill("solid", fgColor=STATE_COLORS["listado"])),
    ]
    for rng, label, value, fill in cards:
        ws_dash.merge_cells(rng)
        cell = ws_dash[rng.split(":")[0]]
        cell.value = f"{label}\n{value}"
        cell.fill = fill
        cell.font = card_value_font if value is not None else card_title_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_dash["A10"] = "Leyenda"
    ws_dash["A10"].fill = accent_fill
    ws_dash["A10"].font = card_title_font
    ws_dash["A10"].border = thin_border
    legend = [("activo", 2), ("referencia", 3), ("derivado", 4), ("interno", 5), ("listado", 6)]
    for state, col in legend:
        cell = ws_dash.cell(10, col)
        cell.value = state
        cell.fill = PatternFill("solid", fgColor=STATE_COLORS[state])
        cell.font = body_font
        cell.border = thin_border

    # Data blocks for charts
    ws_dash["A13"] = "Universo"
    ws_dash["B13"] = "Conteo"
    ws_dash["D13"] = "Estado"
    ws_dash["E13"] = "Conteo"
    ws_dash["G13"] = "Tipo"
    ws_dash["H13"] = "Conteo"
    for cell in ("A13", "B13", "D13", "E13", "G13", "H13"):
        ws_dash[cell].fill = band_fill
        ws_dash[cell].font = heading_font
        ws_dash[cell].border = thin_border
        ws_dash[cell].alignment = Alignment(horizontal="center")

    universe_order = sorted(counts_universe.items(), key=lambda kv: (-kv[1], kv[0]))
    for idx, (key, value) in enumerate(universe_order, start=14):
        ws_dash[f"A{idx}"] = key
        ws_dash[f"B{idx}"] = value
    for idx, key in enumerate(STATE_ORDER, start=14):
        ws_dash[f"D{idx}"] = key
        ws_dash[f"E{idx}"] = counts_state.get(key, 0)
    type_order = sorted(counts_type.items(), key=lambda kv: (-kv[1], kv[0]))[:12]
    top_type_names = [k for k, _ in type_order]
    for idx, (key, value) in enumerate(type_order, start=14):
        ws_dash[f"G{idx}"] = key
        ws_dash[f"H{idx}"] = value

    # Charts
    universe_chart = BarChart()
    universe_chart.type = "bar"
    universe_chart.style = 10
    universe_chart.title = "Relaciones por universo"
    universe_chart.y_axis.title = "Universo"
    universe_chart.x_axis.title = "Relaciones"
    data = Reference(ws_dash, min_col=2, min_row=13, max_row=13 + len(universe_order))
    cats = Reference(ws_dash, min_col=1, min_row=14, max_row=13 + len(universe_order))
    universe_chart.add_data(data, titles_from_data=True)
    universe_chart.set_categories(cats)
    universe_chart.height = 7.5
    universe_chart.width = 12
    universe_chart.legend = None
    ws_dash.add_chart(universe_chart, "J13")

    state_chart = PieChart()
    state_chart.title = "Estados"
    data = Reference(ws_dash, min_col=5, min_row=13, max_row=18)
    cats = Reference(ws_dash, min_col=4, min_row=14, max_row=18)
    state_chart.add_data(data, titles_from_data=True)
    state_chart.set_categories(cats)
    state_chart.height = 7.5
    state_chart.width = 9
    ws_dash.add_chart(state_chart, "J28")

    type_chart = BarChart()
    type_chart.type = "col"
    type_chart.style = 11
    type_chart.title = "Top tipos"
    type_chart.y_axis.title = "Relaciones"
    type_chart.x_axis.title = "Tipo"
    data = Reference(ws_dash, min_col=8, min_row=13, max_row=13 + len(type_order))
    cats = Reference(ws_dash, min_col=7, min_row=14, max_row=13 + len(type_order))
    type_chart.add_data(data, titles_from_data=True)
    type_chart.set_categories(cats)
    type_chart.height = 7.5
    type_chart.width = 12
    ws_dash.add_chart(type_chart, "A28")

    for col, width in {"A": 30, "B": 12, "D": 20, "E": 12, "G": 26, "H": 12, "J": 14, "K": 14, "L": 14}.items():
        ws_dash.column_dimensions[col].width = width

    # Summary
    ws_summary.sheet_view.showGridLines = False
    ws_summary.freeze_panes = "A6"
    ws_summary.merge_cells("A1:G1")
    ws_summary["A1"] = title
    ws_summary["A1"].fill = header_fill
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 24
    ws_summary.merge_cells("A2:G2")
    ws_summary["A2"] = subtitle
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    meta = [
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Filas", str(len(rows))),
        ("Fuentes", "CODEX_ROOT_INVENTORY.csv, SKILLS_UNIFIED_TABLE.csv, scans recursivos"),
    ]
    for idx, (label, value) in enumerate(meta, start=4):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"B{idx}"] = value
        ws_summary[f"A{idx}"].fill = soft_fill
        ws_summary[f"B{idx}"].fill = soft_fill
        ws_summary[f"A{idx}"].border = thin_border
        ws_summary[f"B{idx}"].border = thin_border

    def write_count_block(start_col, title, counts, order=None):
        col1 = chr(64 + start_col)
        col2 = chr(64 + start_col + 1)
        ws_summary[f"{col1}4"] = title
        ws_summary[f"{col2}4"] = "Count"
        for cell in (f"{col1}4", f"{col2}4"):
            ws_summary[cell].fill = band_fill
            ws_summary[cell].font = heading_font
            ws_summary[cell].alignment = Alignment(horizontal="center")
            ws_summary[cell].border = thin_border
        keys = order if order else sorted(counts)
        for idx, key in enumerate(keys, start=5):
            ws_summary[f"{col1}{idx}"] = key
            ws_summary[f"{col2}{idx}"] = counts.get(key, 0)
            for cell in (f"{col1}{idx}", f"{col2}{idx}"):
                ws_summary[cell].border = thin_border
                ws_summary[cell].fill = table_fill if idx % 2 else soft_fill
                ws_summary[cell].font = body_font

    write_count_block(4, "Universo", counts_universe)
    write_count_block(7, "Estado", counts_state, STATE_ORDER)
    write_count_block(10, "Tipo", counts_type, top_type_names + [k for k in sorted(counts_type) if k not in set(top_type_names)])

    ws_summary["A8"] = "Criterio"
    ws_summary["A8"].fill = accent_fill
    ws_summary["A8"].font = heading_font
    ws_summary["A8"].border = thin_border
    criteria = [
        "Origen = ruta o nodo observado.",
        "Tipo = folder, file, workbook, skill, plugin, tool, reference u otros.",
        "Destino = entry visible o superficie de consumo.",
        "Dependencia = documento, raiz o inventario que gobierna el nodo.",
        "Estado = activo, referencia, derivado, interno o listado.",
    ]
    for i, text in enumerate(criteria, start=9):
        ws_summary[f"A{i}"] = f"- {text}"
        ws_summary[f"A{i}"].font = body_font

    ws_summary.column_dimensions["A"].width = 26
    ws_summary.column_dimensions["B"].width = 66
    for col in ("D", "G", "J"):
        ws_summary.column_dimensions[col].width = 24
    for col in ("E", "H", "K"):
        ws_summary.column_dimensions[col].width = 12

    # Main table
    ws_table.sheet_view.showGridLines = False
    ws_table.freeze_panes = "A2"
    ws_table.append(headers)
    for cell in ws_table[1]:
        cell.fill = header_fill
        cell.font = heading_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in rows:
        ws_table.append([row.get(h, "") for h in headers])

    for r in range(2, ws_table.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws_table.cell(r, c)
            cell.border = thin_border
            cell.fill = soft_fill if r % 2 == 0 else table_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = mono_font if c in (1, 2, 3, 4, 5, 6) else body_font

    ws_table.auto_filter.ref = f"A1:G{ws_table.max_row}"
    table = Table(displayName="UniverseRelationshipAudit", ref=f"A1:G{ws_table.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws_table.add_table(table)

    data_range = f"A2:G{ws_table.max_row}"
    for state, color in STATE_COLORS.items():
        fill = PatternFill("solid", fgColor=color)
        rule = FormulaRule(formula=[f'$F2="{state}"'], fill=fill, stopIfTrue=True)
        ws_table.conditional_formatting.add(data_range, rule)

    autosize(ws_table, headers, rows)

    # Lists
    ws_lists.sheet_view.showGridLines = False
    ws_lists["A1"] = "Universos"
    ws_lists["B1"] = "Tipos"
    ws_lists["C1"] = "Estados"
    for cell in ("A1", "B1", "C1"):
        ws_lists[cell].fill = header_fill
        ws_lists[cell].font = heading_font
        ws_lists[cell].border = thin_border
    for idx, value in enumerate(sorted(counts_universe), start=2):
        ws_lists[f"A{idx}"] = value
    for idx, value in enumerate([k for k, _ in type_order] + [k for k in sorted(counts_type) if k not in {x for x, _ in type_order}], start=2):
        ws_lists[f"B{idx}"] = value
    for idx, value in enumerate(STATE_ORDER, start=2):
        ws_lists[f"C{idx}"] = value
    ws_lists.column_dimensions["A"].width = 28
    ws_lists.column_dimensions["B"].width = 24
    ws_lists.column_dimensions["C"].width = 24

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["deep", "executive", "both"], default="both")
    args = parser.parse_args()

    deep_rows = build_cabina_rows()
    executive_rows = build_cabina_executive_rows()

    if args.mode in {"deep", "both"}:
        write_csv(deep_rows, CABINA_DEEP_CSV)
        write_md(
            deep_rows,
            CABINA_DEEP_MD,
            "Tabla profunda de relaciones de cabina-universal-d",
            "Relaciones de auditoria cabina-only: origen -> tipo -> destino -> dependencia -> estado.",
        )
        write_xlsx(
            deep_rows,
            CABINA_DEEP_XLSX,
            "Dashboard cabina-universal-d profundo",
            "Cabina-only con cobertura profunda, filtros duros y colores por estado.",
        )

    if args.mode in {"executive", "both"}:
        write_csv(executive_rows, CABINA_EXEC_CSV)
        write_md(
            executive_rows,
            CABINA_EXEC_MD,
            "Tabla ejecutiva de relaciones de cabina-universal-d",
            "Relaciones clave cabina-only: solo los nodos y dependencias más importantes.",
        )
        write_xlsx(
            executive_rows,
            CABINA_EXEC_XLSX,
            "Dashboard cabina-universal-d ejecutivo",
            "Vista reducida de cabina-only con los enlaces operativos más importantes.",
        )

    print(f"Cabina deep rows: {len(deep_rows)}")
    print(f"Cabina executive rows: {len(executive_rows)}")
    if args.mode in {"deep", "both"}:
        print(CABINA_DEEP_CSV)
        print(CABINA_DEEP_MD)
        print(CABINA_DEEP_XLSX)
    if args.mode in {"executive", "both"}:
        print(CABINA_EXEC_CSV)
        print(CABINA_EXEC_MD)
        print(CABINA_EXEC_XLSX)


if __name__ == "__main__":
    main()
